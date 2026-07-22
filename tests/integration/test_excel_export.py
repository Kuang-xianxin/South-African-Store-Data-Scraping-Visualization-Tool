from __future__ import annotations

from dataclasses import replace
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from takealot_ops.exports.excel import export_excel
from takealot_ops.metrics.service import DashboardDataset


SHEET_NAMES = [
    "运营总览",
    "单品分析",
    "商品数据",
    "异常商品",
    "每日汇总",
    "销售明细",
    "流量快照",
    "指标说明",
    "数据质量",
]


def test_excel_contains_all_eight_required_sheets(
    tmp_path: Path, dashboard_dataset: DashboardDataset
) -> None:
    destination = export_excel(dashboard_dataset, tmp_path / "report.xlsx")

    workbook = load_workbook(destination, read_only=False, data_only=False)
    try:
        assert workbook.sheetnames == SHEET_NAMES
    finally:
        workbook.close()


def test_excel_has_no_vba_project_and_opens_with_openpyxl(
    tmp_path: Path, dashboard_dataset: DashboardDataset
) -> None:
    destination = export_excel(dashboard_dataset, tmp_path / "report.xlsx")

    workbook = load_workbook(destination, keep_vba=False)
    try:
        assert workbook.vba_archive is None
        assert not workbook._external_links
    finally:
        workbook.close()


def test_excel_key_totals_match_dataset(
    tmp_path: Path, dashboard_dataset: DashboardDataset
) -> None:
    destination = export_excel(dashboard_dataset, tmp_path / "report.xlsx")

    workbook = load_workbook(destination, data_only=False)
    try:
        overview = workbook["运营总览"]
        latest_date = dashboard_dataset.product_daily["metric_date"].max()
        latest_views = dashboard_dataset.product_daily.loc[
            dashboard_dataset.product_daily["metric_date"] == latest_date,
            "page_views_30_days",
        ].sum(min_count=1)
        assert overview["A3"].value == "近7天订购件数"
        assert overview["A4"].value == dashboard_dataset.store_daily["ordered_units"].sum()
        assert overview["C3"].value == "近30天浏览量（商品汇总）"
        assert overview["C4"].value == latest_views
        assert overview["E3"].value == "近7天订购销售额"
        assert overview["E4"].value == dashboard_dataset.store_daily["ordered_revenue"].sum()
        assert overview["G3"].value == "异常记录数"
        assert overview["G4"].value == len(dashboard_dataset.anomalies)
        assert overview["A5"].value.startswith("口径提示：")
        assert "有效件数按本地销售状态规则计算" in overview["A5"].value
        assert "需完成销售状态规则确认" not in overview["A5"].value
    finally:
        workbook.close()


def test_excel_overview_limits_recent_kpis_and_trend_to_seven_days(
    tmp_path: Path, dashboard_dataset: DashboardDataset
) -> None:
    dates = pd.date_range("2026-07-15", periods=8, freq="D").date
    store_daily = pd.DataFrame(
        {
            "metric_date": dates,
            "ordered_units": [100, 1, 2, 3, 4, 5, 6, 7],
            "effective_units": [100, 1, 2, 3, 4, 5, 6, 7],
            "ordered_revenue": [1000, 10, 20, 30, 40, 50, 60, 70],
        }
    )
    dataset = replace(dashboard_dataset, store_daily=store_daily)

    destination = export_excel(dataset, tmp_path / "report.xlsx")

    workbook = load_workbook(destination, data_only=False)
    try:
        overview = workbook["运营总览"]
        assert overview["A4"].value == 28
        assert overview["E4"].value == 280
        assert [overview.cell(row, 1).value.date() for row in range(9, 16)] == list(dates[-7:])
        assert overview["A16"].value is None
    finally:
        workbook.close()


def test_excel_has_filters_freezes_charts_and_conditional_formatting(
    tmp_path: Path, dashboard_dataset: DashboardDataset
) -> None:
    destination = export_excel(dashboard_dataset, tmp_path / "report.xlsx")

    workbook = load_workbook(destination)
    try:
        overview = workbook["运营总览"]
        product = workbook["单品分析"]
        assert overview.freeze_panes is not None
        assert overview.auto_filter.ref
        assert len(overview._charts) == 1
        assert overview._charts[0].anchor._from.row == 8
        assert product.freeze_panes is not None
        assert product.auto_filter.ref
        assert len(product._charts) == 0
        assert product["A4"].value == "日期"
        assert workbook["商品数据"].auto_filter.ref
        assert any(len(sheet.conditional_formatting) for sheet in workbook.worksheets)
        assert workbook["销售明细"]["A1"].value == "商品每日销售明细/汇总"
    finally:
        workbook.close()


def test_excel_handles_empty_frames_and_preserves_blank_unknowns(
    tmp_path: Path, empty_dashboard_dataset: DashboardDataset
) -> None:
    destination = export_excel(empty_dashboard_dataset, tmp_path / "empty.xlsx")

    workbook = load_workbook(destination, data_only=False)
    try:
        assert workbook.sheetnames == SHEET_NAMES
        assert workbook["运营总览"]["A4"].value is None
        assert workbook["流量快照"].max_row == 3
    finally:
        workbook.close()


def test_excel_uses_readable_details_and_real_product_fields(
    tmp_path: Path, dashboard_dataset: DashboardDataset
) -> None:
    destination = export_excel(dashboard_dataset, tmp_path / "readable.xlsx")

    workbook = load_workbook(destination, data_only=False)
    try:
        product = workbook["商品数据"]
        assert product["E4"].value == "示例商品 A"
        assert product["I4"].value == "可购买"
        assert product["A4"].number_format == "@"

        analysis = workbook["单品分析"]
        assert analysis["D5"].value == "示例商品 A"
        assert analysis["P5"].value == "可购买"

        anomaly = workbook["异常商品"]
        anomaly_headers = [cell.value for cell in anomaly[3]]
        assert "详情" not in anomaly_headers
        assert anomaly["C4"].value == "示例商品 B"
        assert anomaly["D4"].value == "商品不可购买"

        quality = workbook["数据质量"]
        quality_headers = [cell.value for cell in quality[3]]
        assert "详情" not in quality_headers
        assert quality["F4"].value == "SKU 缺失"
        for sheet in (anomaly, quality):
            assert not any(
                isinstance(cell.value, str) and cell.value.lstrip().startswith(("{", "["))
                for row in sheet.iter_rows(min_row=4)
                for cell in row
            )
    finally:
        workbook.close()


def test_excel_leaves_effective_units_blank_when_sale_status_is_unknown(
    tmp_path: Path, dashboard_dataset: DashboardDataset
) -> None:
    unknown = dashboard_dataset.quality_events.iloc[0].copy()
    unknown["event_id"] = "unknown-status"
    unknown["event_type"] = "unknown_sale_status"
    unknown["offer_id"] = "offer-a"
    unknown["details"] = {"sale_statuses": ["New status"]}
    quality_events = pd.concat(
        [dashboard_dataset.quality_events, unknown.to_frame().T], ignore_index=True
    )
    dataset = replace(dashboard_dataset, quality_events=quality_events)

    destination = export_excel(dataset, tmp_path / "unknown-status.xlsx")
    workbook = load_workbook(destination, data_only=False)
    try:
        assert workbook["运营总览"]["C10"].value is None
        assert workbook["单品分析"]["G6"].value is None
        assert workbook["数据质量"]["G5"].value == "New status"
    finally:
        workbook.close()


def test_excel_preserves_long_numeric_sku_without_scientific_notation(
    tmp_path: Path, dashboard_dataset: DashboardDataset
) -> None:
    product_daily = dashboard_dataset.product_daily.copy(deep=True)
    offer_current = dashboard_dataset.offer_current.copy(deep=True)
    product_daily.loc[product_daily["offer_id"] == "offer-a", "sku"] = "9902240858421"
    offer_current.loc[offer_current["offer_id"] == "offer-a", "sku"] = "9902240858421"
    dataset = replace(
        dashboard_dataset,
        product_daily=product_daily,
        offer_current=offer_current,
    )

    destination = export_excel(dataset, tmp_path / "numeric-sku.xlsx")
    workbook = load_workbook(destination, data_only=False)
    try:
        assert workbook["单品分析"]["C5"].value == 9902240858421
        assert workbook["单品分析"]["C5"].number_format == "0"
        assert workbook["商品数据"]["C4"].value == 9902240858421
        assert workbook["商品数据"]["C4"].number_format == "0"
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "untrusted_text",
    [
        '=HYPERLINK("https://example.invalid", "click")',
        "+SUM(1,1)",
        "-2+3",
        "@SUM(1,1)",
    ],
)
def test_excel_writes_untrusted_text_as_literal_cells(
    tmp_path: Path,
    dashboard_dataset: DashboardDataset,
    untrusted_text: str,
) -> None:
    product_daily = dashboard_dataset.product_daily.copy(deep=True)
    product_daily.at[0, "sku"] = untrusted_text
    unsafe_dataset = replace(dashboard_dataset, product_daily=product_daily)

    destination = export_excel(unsafe_dataset, tmp_path / "safe.xlsx")
    workbook = load_workbook(destination, data_only=False)
    try:
        cell = workbook["销售明细"]["C4"]
        assert cell.data_type == "s"
        assert cell.value == f"'{untrusted_text}"
    finally:
        workbook.close()
