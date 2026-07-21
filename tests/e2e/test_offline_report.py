from __future__ import annotations

from zipfile import ZipFile

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from .release_support import ReleaseFixture


def test_offline_html_has_no_network_dependency(release_fixture: ReleaseFixture) -> None:
    requests: list[str] = []
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        page.on("request", lambda request: requests.append(request.url))
        page.goto(release_fixture.html_path.resolve().as_uri(), wait_until="load")
        page.wait_for_selector('[data-report-ready="true"]', timeout=15_000)
        browser.close()

    assert requests
    assert all(url.startswith("file:") for url in requests)


def test_generated_workbook_reopens_without_repair(
    release_fixture: ReleaseFixture,
) -> None:
    dimensions: list[tuple[str, int, int]] = []
    for _ in range(2):
        workbook = load_workbook(release_fixture.excel_path, data_only=False)
        dimensions = [
            (sheet.title, sheet.max_row, sheet.max_column)
            for sheet in workbook.worksheets
        ]
        assert all(rows >= 3 and columns >= 1 for _, rows, columns in dimensions)
        assert workbook["运营总览"]._charts
        assert workbook["单品分析"]._charts
        workbook.close()

    assert len(dimensions) == 8
    with ZipFile(release_fixture.excel_path) as archive:
        names = set(archive.namelist())
    assert any(name.startswith("xl/charts/chart") for name in names)
    assert any(name.startswith("xl/drawings/_rels/") for name in names)
