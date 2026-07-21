from __future__ import annotations

import json
import re
from collections import defaultdict

from openpyxl import load_workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session

from takealot_ops.storage.models import DailyProductMetric, OfferSnapshot, SaleItem

from .release_support import ReleaseFixture, business_state, reseed_and_rebuild


def test_ten_sku_totals_match_database_html_and_excel(
    release_fixture: ReleaseFixture,
) -> None:
    engine = create_engine(release_fixture.settings.database_url)
    with Session(engine) as session:
        database_totals = dict(
            session.execute(
                select(DailyProductMetric.sku, func.sum(DailyProductMetric.ordered_units))
                .where(DailyProductMetric.sku.is_not(None))
                .group_by(DailyProductMetric.sku)
            ).all()
        )
    engine.dispose()

    document = release_fixture.html_path.read_text(encoding="utf-8")
    match = re.search(
        r'<script id="dashboard-data" type="application/json">(.*?)</script>',
        document,
        flags=re.DOTALL,
    )
    assert match is not None
    html_payload = json.loads(match.group(1))
    html_totals: defaultdict[str, int] = defaultdict(int)
    for row in html_payload["product_daily"]:
        html_totals[row["sku"]] += int(row["ordered_units"] or 0)

    workbook = load_workbook(release_fixture.excel_path, data_only=False, read_only=False)
    excel_totals: defaultdict[str, int] = defaultdict(int)
    for row in workbook["销售明细"].iter_rows(min_row=4, values_only=True):
        if row[2] is not None:
            excel_totals[str(row[2])] += int(row[3] or 0)
    workbook.close()

    assert len(database_totals) == 10
    assert dict(html_totals) == database_totals
    assert dict(excel_totals) == database_totals


def test_traffic_snapshots_match_source_offers(release_fixture: ReleaseFixture) -> None:
    engine = create_engine(release_fixture.settings.database_url)
    with Session(engine) as session:
        rows = session.scalars(
            select(OfferSnapshot).order_by(
                OfferSnapshot.snapshot_date, OfferSnapshot.offer_id
            )
        ).all()
    engine.dispose()
    actual = {
        (row.snapshot_date, row.offer_id): row.page_views_30_days
        for row in rows
    }
    assert len(actual) == 31 * 10
    assert actual == release_fixture.source_traffic


def test_repeated_collection_is_idempotent(release_fixture: ReleaseFixture) -> None:
    before = business_state(release_fixture.settings.database_url)

    reseed_and_rebuild(release_fixture)

    after = business_state(release_fixture.settings.database_url)
    assert after == before


def test_sast_boundaries_status_update_unknown_status_and_missing_traffic(
    release_fixture: ReleaseFixture,
) -> None:
    engine = create_engine(release_fixture.settings.database_url)
    with Session(engine) as session:
        before = session.get(SaleItem, "boundary-before")
        after = session.get(SaleItem, "boundary-after")
        repeated = session.get(SaleItem, "repeated-item")
        unknown = session.get(SaleItem, "unknown-item")
        missing = session.scalars(
            select(OfferSnapshot).where(OfferSnapshot.offer_id == "offer-09")
        ).all()
    engine.dispose()
    assert before is not None and before.sales_day.isoformat() == "2026-07-20"
    assert after is not None and after.sales_day.isoformat() == "2026-07-21"
    assert repeated is not None and repeated.sale_status == "excluded"
    assert unknown is not None and unknown.sale_status == "new-marketplace-status"
    assert len(missing) == 31
    assert all(row.page_views_30_days is None for row in missing)
