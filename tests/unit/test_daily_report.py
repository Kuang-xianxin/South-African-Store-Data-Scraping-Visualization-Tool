from __future__ import annotations

from datetime import UTC, date, datetime, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from takealot_ops.erp import daily_report as daily_report_module
from takealot_ops.erp.daily_report import (
    DailyReportConflictError,
    DailyReportInputError,
    backfill_daily_inventory_snapshots,
    backfill_stock_continuity_reviews,
    capture_daily_report,
    confirm_entry,
    create_deadline_snapshot,
    daily_report_payload,
    delete_operator_note,
    dismiss_stock_alert,
    eliminate_stock_alert,
    export_operations_workbook,
    operations_business_date,
    record_daily_report_failure,
    reminder_payload,
    revert_confirmation,
    reopen_stock_alert,
    save_manual_candidate,
    save_operator_note,
    update_operator_note,
)
from takealot_ops.storage.migrations import create_schema
from takealot_ops.storage.models import (
    DailyInventorySnapshot,
    DailyReportAudit,
    DailyReportObservation,
    DailyReportRun,
    DailyReportResolution,
    ErpUser,
    OfferCurrent,
    SaleItem,
)


REPORT_DATE = date(2026, 7, 24)


def _report_capture_time(
    business_date: date,
    hour: int,
    minute: int = 0,
) -> datetime:
    inventory_date = business_date + timedelta(days=1)
    return datetime(
        inventory_date.year,
        inventory_date.month,
        inventory_date.day,
        hour,
        minute,
        tzinfo=UTC,
    )


def test_capture_business_date_uses_beijing_ten_to_ten_cycle() -> None:
    assert operations_business_date(
        datetime(2026, 7, 25, 2, 5, tzinfo=UTC)
    ) == date(2026, 7, 24)
    assert operations_business_date(
        datetime(2026, 7, 25, 10, 0, tzinfo=UTC)
    ) == date(2026, 7, 24)
    assert operations_business_date(
        datetime(2026, 7, 26, 1, 59, tzinfo=UTC)
    ) == date(2026, 7, 24)
    assert operations_business_date(
        datetime(2026, 7, 26, 2, 0, tzinfo=UTC)
    ) == date(2026, 7, 25)


def _seed(engine) -> None:
    with Session(engine) as session, session.begin():
        session.add(
            ErpUser(
                id=1,
                username="operator",
                display_name="Operator",
                password_hash="unused",
                role="operator",
                active=True,
                created_at=datetime(2026, 7, 24, 1),
                updated_at=datetime(2026, 7, 24, 1),
            )
        )
        session.add_all(
            [
                OfferCurrent(
                    offer_id="offer-a",
                    sku="9900000000001",
                    title="Product A",
                    image_url="https://example.invalid/product-a.png",
                    captured_at=datetime(2026, 7, 24, 2, tzinfo=UTC),
                    page_views_30_days=50,
                    takealot_available_stock=9,
                ),
                OfferCurrent(
                    offer_id="offer-b",
                    sku="9900000000002",
                    title="Product B",
                    image_url=None,
                    captured_at=datetime(2026, 7, 24, 2, tzinfo=UTC),
                    page_views_30_days=20,
                    takealot_available_stock=4,
                ),
            ]
        )
        session.add(
            SaleItem(
                order_item_id="sale-a",
                order_date=datetime(2026, 7, 24, 1, tzinfo=UTC),
                sales_day=REPORT_DATE,
                offer_id="offer-a",
                sku="9900000000001",
                quantity=1,
                raw_payload={},
            )
        )


def _engine():
    engine = create_engine("sqlite://")
    create_schema(engine)
    _seed(engine)
    with Session(engine) as session, session.begin():
        source_run_id = "inventory-source-2026-07-24"
        captured_at = datetime(2026, 7, 24, 2, 5)
        session.add(
            DailyReportRun(
                run_id=source_run_id,
                business_date=REPORT_DATE - timedelta(days=1),
                slot="morning",
                captured_at=captured_at,
                status="success",
                counts={"products": 2},
                created_at=captured_at,
            )
        )
        session.add_all(
            [
                DailyInventorySnapshot(
                    inventory_date=REPORT_DATE,
                    offer_id="offer-a",
                    run_id=source_run_id,
                    captured_at=captured_at,
                    platform_stock=9,
                    stock_source="takealot_available_stock",
                ),
                DailyInventorySnapshot(
                    inventory_date=REPORT_DATE,
                    offer_id="offer-b",
                    run_id=source_run_id,
                    captured_at=captured_at,
                    platform_stock=4,
                    stock_source="takealot_available_stock",
                ),
            ]
        )
    return engine


def test_capture_keeps_morning_and_evening_versions_and_requires_review() -> None:
    engine = _engine()
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2, 37),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 7
        session.get(SaleItem, "sale-a").quantity = 2
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="evening",
        captured_at=_report_capture_time(REPORT_DATE, 10),
    )

    payload = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in payload["items"] if row["offer_id"] == "offer-a")
    assert payload["counts"]["current_stock_total"] == 11
    assert payload["counts"]["current_stock_missing"] == 0
    assert product["image_url"] == "https://example.invalid/product-a.png"
    assert product["morning"]["ordered_units"] == 1
    assert product["morning"]["platform_stock"] == 9
    assert product["capture_versions"][0]["label"] == (
        "早间采集（实际 07-25 10:37）"
    )
    assert payload["capture_status"]["morning"]["captured_at"] == (
        "2026-07-25T02:37:00"
    )
    assert product["evening"]["ordered_units"] == 2
    assert product["evening"]["platform_stock"] == 9
    assert product["status"] == "needs_review"
    assert product["differences"] == ["ordered_units"]
    assert product["review_issues"] == [
        {
            "type": "capture_difference",
            "fields": product["differences"],
        }
    ]
    assert product["stock_context"] is None
    pending = next(
        row
        for row in payload["pending_actions"]
        if row["offer_id"] == "offer-a"
    )
    assert pending["image_url"] == "https://example.invalid/product-a.png"
    unchanged = next(row for row in payload["items"] if row["offer_id"] == "offer-b")
    assert unchanged["status"] == "ready"


def test_pre_close_capture_updates_sales_without_replacing_morning_inventory() -> None:
    engine = _engine()
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2, 5),
    )
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="evening",
        captured_at=_report_capture_time(REPORT_DATE, 10),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 1
        session.get(SaleItem, "sale-a").quantity = 3
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="pre_close",
        captured_at=datetime(2026, 7, 26, 1, tzinfo=UTC),
    )

    payload = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in payload["items"] if row["offer_id"] == "offer-a")

    assert payload["capture_status"]["pre_close"]["status"] == "success"
    assert product["capture_versions"][-1]["slot"] == "pre_close"
    assert product["capture_versions"][-1]["label"] == (
        "周期末采集（实际 07-26 09:00）"
    )
    assert product["capture_versions"][-1]["values"]["ordered_units"] == 3
    assert product["capture_versions"][-1]["values"]["platform_stock"] == 9
    assert product["status"] == "needs_review"
    assert product["differences"] == ["ordered_units"]


def test_reports_before_pre_close_activation_are_not_marked_missing() -> None:
    payload = daily_report_payload(_engine(), REPORT_DATE)

    assert payload["capture_status"]["pre_close"]["status"] == "not_applicable"
    assert all(
        issue["slot"] != "pre_close" for issue in payload["capture_issues"]
    )


def test_capture_issue_history_defaults_to_three_days_and_supports_a_custom_range(
) -> None:
    engine = _engine()
    included_date = REPORT_DATE - timedelta(days=2)
    older_date = REPORT_DATE - timedelta(days=3)
    record_daily_report_failure(
        engine,
        business_date=included_date,
        slot="morning",
        captured_at=_report_capture_time(included_date, 2, 5),
        reason="included capture failure",
    )
    record_daily_report_failure(
        engine,
        business_date=older_date,
        slot="morning",
        captured_at=_report_capture_time(older_date, 2, 5),
        reason="older capture failure",
    )

    payload = daily_report_payload(engine, REPORT_DATE)

    assert payload["capture_issue_range"]["selected_start"] == included_date.isoformat()
    assert payload["capture_issue_range"]["selected_end"] == REPORT_DATE.isoformat()
    assert all("business_date" in issue for issue in payload["capture_issues"])
    assert all(
        included_date.isoformat()
        <= issue["business_date"]
        <= REPORT_DATE.isoformat()
        for issue in payload["capture_issues"]
    )
    assert any(
        issue["reason"] == "included capture failure"
        for issue in payload["capture_issues"]
    )
    assert not any(
        issue["reason"] == "older capture failure"
        for issue in payload["capture_issues"]
    )

    older_payload = daily_report_payload(
        engine,
        REPORT_DATE,
        capture_start=older_date,
        capture_end=older_date,
    )

    assert older_payload["capture_issue_range"]["selected_start"] == (
        older_date.isoformat()
    )
    assert older_payload["capture_issue_range"]["selected_end"] == older_date.isoformat()
    assert older_payload["capture_issues"]
    assert all(
        issue["business_date"] == older_date.isoformat()
        for issue in older_payload["capture_issues"]
    )
    assert any(
        issue["reason"] == "older capture failure"
        for issue in older_payload["capture_issues"]
    )


def test_capture_issue_history_rejects_an_inverted_range() -> None:
    with pytest.raises(
        DailyReportInputError,
        match="数据完整性说明开始日期不能晚于结束日期",
    ):
        daily_report_payload(
            _engine(),
            REPORT_DATE,
            capture_start=REPORT_DATE,
            capture_end=REPORT_DATE - timedelta(days=1),
        )


def test_every_manual_refresh_in_the_ten_to_ten_cycle_is_compared() -> None:
    engine = _engine()
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2, 5),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 7
        session.get(SaleItem, "sale-a").quantity = 2
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="manual",
        captured_at=_report_capture_time(REPORT_DATE, 6, 0),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 9
        session.get(SaleItem, "sale-a").quantity = 1
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="manual",
        captured_at=_report_capture_time(REPORT_DATE, 7, 0),
    )

    payload = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in payload["items"] if row["offer_id"] == "offer-a")

    assert product["status"] == "needs_review"
    assert product["differences"] == ["ordered_units"]
    assert product["current"]["ordered_units"] == 1
    assert product["current"]["platform_stock"] == 9
    assert [row["slot"] for row in product["capture_versions"]] == [
        "morning",
        "manual",
        "manual",
    ]
    assert [issue["type"] for issue in product["review_issues"]] == [
        "capture_difference"
    ]
    assert len([run for run in payload["runs"] if run["slot"] == "manual"]) == 2


def test_later_capture_stock_is_not_written_over_the_1005_snapshot() -> None:
    engine = _engine()
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2, 5),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 7
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="evening",
        captured_at=_report_capture_time(REPORT_DATE, 10, 0),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 6
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="manual",
        captured_at=_report_capture_time(REPORT_DATE, 11, 0),
    )

    payload = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in payload["items"] if row["offer_id"] == "offer-a")

    assert product["status"] == "ready"
    assert product["differences"] == []
    assert product["review_issues"] == []
    assert product["current"]["platform_stock"] == 9
    assert [
        version["values"]["platform_stock"]
        for version in product["capture_versions"]
    ] == [9, 9, 9]
    assert reminder_payload(engine, REPORT_DATE)["count"] == 0


def test_delayed_same_day_capture_fills_missing_morning_stock_and_marks_resolved(
) -> None:
    engine = create_engine("sqlite://")
    create_schema(engine)
    _seed(engine)

    record_daily_report_failure(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2, 6),
        reason="Offers HTTP 403",
    )
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="manual",
        captured_at=_report_capture_time(REPORT_DATE, 2, 44),
    )

    payload = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in payload["items"] if row["offer_id"] == "offer-a")
    assert payload["counts"]["current_stock_total"] == 13
    assert payload["counts"]["current_stock_missing"] == 0
    assert product["current"]["platform_stock"] == 9
    assert product["status"] == "awaiting_evening"
    assert product["missing_capture"] is False
    assert product["missing_reason"] is None
    assert all(issue["kind"] != "product" for issue in payload["capture_issues"])
    assert any(issue["slot"] == "morning" for issue in payload["capture_issues"])

    context = payload["comparison_history"][-1]["inventory_context"]
    assert context["inventory_date"] == "2026-07-25"
    assert context["captured_at"] == "2026-07-25T02:44:00"
    assert context["source_slot"] == "manual"
    assert context["delayed"] is True
    assert context["resolved_after_missing"] is True
    assert context["complete"] is True
    assert context["missing_count"] == 0
    assert "早间库存漏爬已解决" in context["note"]
    assert "北京时间 2026-07-25 10:44:00" in context["note"]
    assert "周期内首次完整库存" in context["note"]
    assert context["exception_note"] == context["note"]

    manual_run = next(run for run in payload["runs"] if run["slot"] == "manual")
    assert manual_run["counts"]["reported_inventory_missing"] == 0
    assert manual_run["counts"]["reported_inventory_resolved"] is True
    assert manual_run["counts"]["reported_inventory_date"] == "2026-07-25"


def test_first_complete_delayed_inventory_is_not_replaced_by_later_captures() -> None:
    engine = create_engine("sqlite://")
    create_schema(engine)
    _seed(engine)

    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="manual",
        captured_at=_report_capture_time(REPORT_DATE, 2, 44),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 7
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="evening",
        captured_at=_report_capture_time(REPORT_DATE, 10),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 5
        session.get(OfferCurrent, "offer-b").takealot_available_stock = None
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="manual",
        captured_at=_report_capture_time(REPORT_DATE, 11),
    )

    payload = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in payload["items"] if row["offer_id"] == "offer-a")
    context = payload["comparison_history"][-1]["inventory_context"]

    assert product["current"]["platform_stock"] == 9
    assert [
        version["values"]["platform_stock"]
        for version in product["capture_versions"]
    ] == [9, 9, 9]
    assert context["source_slot"] == "manual"
    assert context["captured_at"] == "2026-07-25T02:44:00"
    assert context["complete"] is True
    assert "北京时间 2026-07-25 10:44:00" in context["note"]
    assert context["exception_note"] == context["note"]


def test_incomplete_inventory_is_replaced_once_by_the_first_complete_capture() -> None:
    engine = create_engine("sqlite://")
    create_schema(engine)
    _seed(engine)
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-b").takealot_available_stock = None
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="manual",
        captured_at=_report_capture_time(REPORT_DATE, 2, 20),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 8
        session.get(OfferCurrent, "offer-b").takealot_available_stock = 4
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="manual",
        captured_at=_report_capture_time(REPORT_DATE, 2, 40),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 7
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="evening",
        captured_at=_report_capture_time(REPORT_DATE, 10),
    )

    payload = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in payload["items"] if row["offer_id"] == "offer-a")
    context = payload["comparison_history"][-1]["inventory_context"]
    assert product["current"]["platform_stock"] == 8
    assert context["captured_at"] == "2026-07-25T02:40:00"
    assert context["complete"] is True
    assert context["missing_count"] == 0


def test_current_stock_total_follows_latest_offer_inventory_and_preserves_missing() -> None:
    engine = _engine()
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2, 5),
    )
    with Session(engine) as session, session.begin():
        offer_a = session.get(OfferCurrent, "offer-a")
        offer_b = session.get(OfferCurrent, "offer-b")
        assert offer_a is not None
        assert offer_b is not None
        offer_a.takealot_available_stock = 12
        offer_b.takealot_available_stock = 6

    refreshed_payload = daily_report_payload(engine, REPORT_DATE)

    assert refreshed_payload["counts"]["current_stock_total"] == 18
    assert refreshed_payload["counts"]["current_stock_missing"] == 0
    report_item = next(
        item
        for item in refreshed_payload["items"]
        if item["offer_id"] == "offer-a"
    )
    assert report_item["current"]["platform_stock"] == 9

    with Session(engine) as session, session.begin():
        offer_b = session.get(OfferCurrent, "offer-b")
        assert offer_b is not None
        offer_b.takealot_available_stock = None
        offer_b.total_stock = None

    payload = daily_report_payload(engine, REPORT_DATE)

    assert payload["counts"]["current_stock_total"] is None
    assert payload["counts"]["current_stock_missing"] == 1


def test_inventory_backfill_attaches_next_morning_stock_to_report_day() -> None:
    engine = create_engine("sqlite://")
    create_schema(engine)
    _seed(engine)
    with Session(engine) as session, session.begin():
        source_run = DailyReportRun(
            run_id="old-morning-source",
            business_date=REPORT_DATE - timedelta(days=1),
            slot="morning",
            captured_at=datetime(2026, 7, 24, 2, 5),
            status="success",
            counts={"products": 1},
            created_at=datetime(2026, 7, 24, 2, 5),
        )
        report_run = DailyReportRun(
            run_id="old-next-day-report",
            business_date=REPORT_DATE,
            slot="morning",
            captured_at=datetime(2026, 7, 25, 2, 5),
            status="success",
            counts={"products": 1},
            created_at=datetime(2026, 7, 25, 2, 5),
        )
        session.add_all([source_run, report_run])
        session.add_all(
            [
                DailyReportObservation(
                    run_id=source_run.run_id,
                    offer_id="offer-a",
                    sku="9900000000001",
                    title="Product A",
                    page_views_30_days=50,
                    ordered_units=1,
                    platform_stock=9,
                    stock_source="takealot_available_stock",
                ),
                DailyReportObservation(
                    run_id=report_run.run_id,
                    offer_id="offer-a",
                    sku="9900000000001",
                    title="Product A",
                    page_views_30_days=50,
                    ordered_units=1,
                    platform_stock=7,
                    stock_source="takealot_available_stock",
                ),
            ]
        )

    result = backfill_daily_inventory_snapshots(engine, through=REPORT_DATE)

    assert result.snapshots_created == 2
    assert result.observations_updated == 2
    with Session(engine) as session:
        snapshot = session.scalar(
            select(DailyInventorySnapshot).where(
                DailyInventorySnapshot.inventory_date
                == REPORT_DATE + timedelta(days=1),
                DailyInventorySnapshot.offer_id == "offer-a",
            )
        )
        report_observation = session.scalar(
            select(DailyReportObservation).where(
                DailyReportObservation.run_id == "old-next-day-report",
                DailyReportObservation.offer_id == "offer-a",
            )
        )
        assert snapshot is not None
        assert snapshot.platform_stock == 7
        assert report_observation is not None
        assert report_observation.platform_stock == 7
        assert report_observation.stock_source == "next_morning_1005"
        report_run = session.get(DailyReportRun, "old-next-day-report")
        assert report_run is not None
        assert report_run.counts["reported_inventory_date"] == "2026-07-25"
        assert report_run.counts["reported_inventory_snapshots"] == 1
        assert report_run.counts["reported_inventory_missing"] == 0
    second = backfill_daily_inventory_snapshots(engine, through=REPORT_DATE)
    assert second.snapshots_created == 0
    assert second.observations_updated == 0


def test_operator_notes_support_create_update_delete_and_keep_audit_history() -> None:
    engine = _engine()
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2, 5),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 7
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="evening",
        captured_at=_report_capture_time(REPORT_DATE, 10),
    )

    save_operator_note(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        note="先核对晚间库存来源",
        issue_type="capture_difference",
        user_id=1,
    )
    save_operator_note(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        note="运营补充：等待仓库盘点结果",
        issue_type="general",
        user_id=1,
    )

    initial = next(
        row
        for row in daily_report_payload(engine, REPORT_DATE)["items"]
        if row["offer_id"] == "offer-a"
    )
    first_note_id = initial["operator_notes"][0]["id"]
    second_note_id = initial["operator_notes"][1]["id"]
    update_operator_note(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        note_id=first_note_id,
        note="已核对晚间库存来源",
        issue_type="stock_continuity",
        user_id=1,
    )
    with pytest.raises(
        DailyReportInputError,
        match="删除问题备注必须填写删除原因",
    ):
        delete_operator_note(
            engine,
            business_date=REPORT_DATE,
            offer_id="offer-a",
            note_id=first_note_id,
            note="",
            user_id=1,
        )
    delete_operator_note(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        note_id=second_note_id,
        note="",
        user_id=1,
    )

    product = next(
        row
        for row in daily_report_payload(engine, REPORT_DATE)["items"]
        if row["offer_id"] == "offer-a"
    )
    assert product["status"] == "ready"
    assert product["operator_note"] == "已核对晚间库存来源"
    assert [
        (
            note["id"],
            note["issue_type"],
            note["note"],
            note["user_name"],
            note["updated_by"],
        )
        for note in product["operator_notes"]
    ] == [
        (
            first_note_id,
            "stock_continuity",
            "已核对晚间库存来源",
            "Operator",
            "Operator",
        ),
    ]
    with Session(engine) as session:
        actions = list(
            session.scalars(
                select(DailyReportAudit.action)
                .where(DailyReportAudit.offer_id == "offer-a")
                .order_by(DailyReportAudit.id)
            )
        )
    assert actions[-4:] == [
        "operator_note",
        "operator_note",
        "operator_note_updated",
        "operator_note_deleted",
    ]
    handled = daily_report_payload(engine, REPORT_DATE)["handled_actions"]
    assert [row["action_type"] for row in handled[:4]] == [
        "operator_note_deleted",
        "operator_note_updated",
        "operator_note",
        "operator_note",
    ]
    assert handled[0]["note"] is None
    assert handled[0]["detail"]["deleted_note"] == (
        "运营补充：等待仓库盘点结果"
    )
    assert handled[1]["note"] == "已核对晚间库存来源"
    assert handled[1]["detail"]["before_note"] == "先核对晚间库存来源"
    assert handled[1]["detail"]["after_note"] == "已核对晚间库存来源"


def test_confirmed_entry_does_not_reopen_for_page_view_or_stock_changes() -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )
    confirm_entry(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        source="latest",
        note="库存和订单一致，采用本周期最新值",
        user_id=1,
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").page_views_30_days = 99
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 7
    capture = capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="manual",
        captured_at=_report_capture_time(REPORT_DATE, 11, 0),
    )

    product = next(
        row
        for row in daily_report_payload(engine, REPORT_DATE)["items"]
        if row["offer_id"] == "offer-a"
    )
    assert capture.reopened_count == 0
    assert product["status"] == "confirmed"
    assert product["differences"] == []
    assert product["current"]["platform_stock"] == 9


def test_page_view_change_is_kept_but_does_not_require_merge() -> None:
    engine = _engine()
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2, 5),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").page_views_30_days = 58
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="evening",
        captured_at=_report_capture_time(REPORT_DATE, 10),
    )

    payload = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in payload["items"] if row["offer_id"] == "offer-a")
    assert product["morning"]["page_views_30_days"] == 50
    assert product["evening"]["page_views_30_days"] == 58
    assert product["current"]["page_views_30_days"] == 58
    assert product["status"] == "ready"
    assert product["differences"] == []
    assert reminder_payload(engine, REPORT_DATE)["count"] == 0


def test_manual_candidate_and_confirm_are_separate_audited_states() -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )
    save_manual_candidate(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        values={"ordered_units": 3},
        reason="platform_delay",
        note="平台订单延迟，人工核对为3件",
        user_id=1,
    )
    before = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in before["items"] if row["offer_id"] == "offer-a")
    assert product["status"] == "needs_review"
    assert product["manual"]["ordered_units"] == 3
    assert product["current"]["ordered_units"] == 1

    save_manual_candidate(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        values={"ordered_units": 4, "platform_stock": 8},
        reason="stock_adjustment",
        note="第二次核对后改为订单4、库存8",
        user_id=1,
    )
    revised = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in revised["items"] if row["offer_id"] == "offer-a")
    assert product["manual"]["ordered_units"] == 4
    assert product["manual"]["platform_stock"] == 8
    assert product["manual_reason"] == "stock_adjustment"
    assert product["manual_note"] == "第二次核对后改为订单4、库存8"
    with Session(engine) as session:
        edits = list(
            session.scalars(
                select(DailyReportAudit)
                .where(
                    DailyReportAudit.offer_id == "offer-a",
                    DailyReportAudit.action == "manual_candidate",
                )
                .order_by(DailyReportAudit.id)
            )
        )
    assert len(edits) == 2
    assert edits[1].payload["before"]["ordered_units"] == 3
    assert edits[1].payload["after"]["ordered_units"] == 4
    assert edits[1].payload["after"]["platform_stock"] == 8

    confirm_entry(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        source="manual",
        note="采用人工核对订单，库存沿用晚间值",
        user_id=1,
    )
    after = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in after["items"] if row["offer_id"] == "offer-a")
    assert product["status"] == "confirmed"
    assert product["final"]["ordered_units"] == 4
    assert product["final"]["platform_stock"] == 8
    handled = after["handled_actions"]
    assert [row["action_type"] for row in handled[:3]] == [
        "confirmation",
        "manual_candidate",
        "manual_candidate",
    ]
    assert handled[0]["note"] == "采用人工核对订单，库存沿用晚间值"
    assert all(
        row["image_url"] == "https://example.invalid/product-a.png"
        for row in handled[:3]
    )
    assert handled[1]["note"] == "第二次核对后改为订单4、库存8"
    assert handled[1]["detail"]["before_values"]["ordered_units"] == 3
    assert handled[1]["detail"]["after_values"]["ordered_units"] == 4


def test_manual_candidate_accepts_no_note_and_keeps_confirmation_note_required() -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )

    save_manual_candidate(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        values={"ordered_units": 3},
        reason="platform_delay",
        note="",
        user_id=1,
    )

    payload = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in payload["items"] if row["offer_id"] == "offer-a")
    assert product["manual"]["ordered_units"] == 3
    assert product["manual_note"] is None
    with Session(engine) as session:
        manual_audit = session.scalar(
            select(DailyReportAudit)
            .where(
                DailyReportAudit.offer_id == "offer-a",
                DailyReportAudit.action == "manual_candidate",
            )
            .order_by(DailyReportAudit.id.desc())
        )
    assert manual_audit is not None
    assert manual_audit.note is None

    with pytest.raises(DailyReportInputError, match="确认合并必须填写备注"):
        confirm_entry(
            engine,
            business_date=REPORT_DATE,
            offer_id="offer-a",
            source="manual",
            note="",
            user_id=1,
        )


def test_matching_manual_fix_closes_version_difference_without_stock_action() -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )

    next_date = REPORT_DATE + timedelta(days=1)
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 7
        session.add(
            SaleItem(
                order_item_id="sale-a-next",
                order_date=datetime(2026, 7, 25, 1, tzinfo=UTC),
                sales_day=next_date,
                offer_id="offer-a",
                sku="9900000000001",
                quantity=1,
                raw_payload={},
            )
        )
    capture_daily_report(
        engine,
        business_date=next_date,
        slot="morning",
        captured_at=_report_capture_time(next_date, 2),
    )
    with Session(engine) as session, session.begin():
        session.get(SaleItem, "sale-a-next").quantity = 2
    capture_daily_report(
        engine,
        business_date=next_date,
        slot="evening",
        captured_at=_report_capture_time(next_date, 10),
    )

    before = next(
        item
        for item in daily_report_payload(engine, next_date)["pending_actions"]
        if item["offer_id"] == "offer-a"
    )
    assert before["differences"] == ["ordered_units"]

    save_manual_candidate(
        engine,
        business_date=next_date,
        offer_id="offer-a",
        values={"ordered_units": 2, "platform_stock": 7},
        reason="stock_adjustment",
        note="人工核对销量为2且库存公式相符",
        user_id=1,
    )

    payload = daily_report_payload(engine, next_date)
    product = next(
        item for item in payload["items"] if item["offer_id"] == "offer-a"
    )
    assert product["status"] == "confirmed"
    assert product["final"]["ordered_units"] == 2
    assert product["final"]["platform_stock"] == 7
    assert product["differences"] == []
    assert product["stock_check"]["mismatch"] is False
    assert not any(
        item["offer_id"] == "offer-a" for item in payload["pending_actions"]
    )
    assert reminder_payload(engine, next_date)["count"] == 0
    confirmation = next(
        item
        for item in payload["handled_actions"]
        if item["action_type"] == "confirmation"
        and item["offer_id"] == "offer-a"
    )
    assert confirmation["detail"]["source"] == "manual"
    assert confirmation["detail"]["automatic"] is True
    assert confirmation["note"] == "人工核对销量为2且库存公式相符"


def test_confirmation_can_be_reverted_and_reconfirmed_with_full_audit() -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )
    confirm_entry(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        source="latest",
        note="首次采用本周期最新值",
        user_id=1,
    )

    revert_confirmation(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        note="复核后发现选错来源，需要重新确认",
        user_id=1,
    )
    reopened = next(
        item
        for item in daily_report_payload(engine, REPORT_DATE)["items"]
        if item["offer_id"] == "offer-a"
    )
    assert reopened["status"] == "needs_review"
    assert reopened["confirmation_baseline"] is None
    assert reopened["review_issues"] == [
        {"type": "confirmation_reverted", "fields": []}
    ]
    assert reopened["confirmation_revert"]["revert_note"] == (
        "复核后发现选错来源，需要重新确认"
    )
    assert reopened["confirmation_revert"]["previous_confirmation"]["source"] == (
        "latest"
    )
    assert reopened["confirmation_revert"]["previous_confirmation"]["values"][
        "platform_stock"
    ] == 9
    handled_after_revert = daily_report_payload(
        engine,
        REPORT_DATE,
    )["handled_actions"]
    assert handled_after_revert[0]["action_type"] == "confirmation_reverted"
    assert handled_after_revert[0]["note"] == (
        "复核后发现选错来源，需要重新确认"
    )
    original_confirmation = next(
        row
        for row in handled_after_revert
        if row["action_type"] == "confirmation"
    )
    assert original_confirmation["active"] is False
    assert original_confirmation["reversal"]["note"] == (
        "复核后发现选错来源，需要重新确认"
    )

    confirm_entry(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        source="morning",
        note="重新核对后采用早间值",
        user_id=1,
    )
    reconfirmed = next(
        item
        for item in daily_report_payload(engine, REPORT_DATE)["items"]
        if item["offer_id"] == "offer-a"
    )
    assert reconfirmed["status"] == "confirmed"
    assert reconfirmed["confirmation_revert"] is None
    with Session(engine) as session:
        actions = list(
            session.scalars(
                select(DailyReportAudit.action)
                .where(DailyReportAudit.offer_id == "offer-a")
                .order_by(DailyReportAudit.id)
            )
        )
    assert actions[-3:] == ["confirm", "confirmation_reverted", "confirm"]

    with pytest.raises(DailyReportConflictError, match="没有可撤销"):
        revert_confirmation(
            engine,
            business_date=REPORT_DATE,
            offer_id="offer-b",
            note="不存在确认记录",
            user_id=1,
        )


def test_revert_keeps_following_continuity_as_pending_until_reconfirmation() -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )
    confirm_entry(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        source="latest",
        note="确认首日库存9",
        user_id=1,
    )

    next_date = REPORT_DATE + timedelta(days=1)
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 7
        session.add(
            SaleItem(
                order_item_id="sale-a-after-revert",
                order_date=datetime(2026, 7, 25, 1, tzinfo=UTC),
                sales_day=next_date,
                offer_id="offer-a",
                sku="9900000000001",
                quantity=1,
                raw_payload={},
            )
        )
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=next_date,
            slot=slot,
            captured_at=_report_capture_time(next_date, hour),
        )
    before = next(
        item
        for item in daily_report_payload(engine, next_date)["pending_actions"]
        if item["business_date"] == next_date.isoformat()
        and item["offer_id"] == "offer-a"
    )
    assert before["review_issues"][0]["type"] == "stock_continuity"

    revert_confirmation(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        note="首日确认来源需要重新核对",
        user_id=1,
    )
    affected = next(
        item
        for item in daily_report_payload(engine, next_date)["pending_actions"]
        if item["business_date"] == next_date.isoformat()
        and item["offer_id"] == "offer-a"
    )
    assert affected["status"] == "needs_review"
    assert affected["review_issues"] == [
        {
            "type": "confirmation_revert_impact",
            "fields": ["ordered_units", "platform_stock"],
        }
    ]
    assert affected["stock_check"]["mismatch"] is False
    assert affected["stock_check"]["deferred_reason"] == (
        "前一日报日的人工确认已撤销；本日保留待办，待重新确认后立即重算库存连续性"
    )
    assert affected["stock_context"]["source"] == "confirmation_reverted"
    assert affected["stock_context"]["confirmation_revert"]["revert_note"] == (
        "首日确认来源需要重新核对"
    )
    with Session(engine) as session:
        impact_audit = session.scalar(
            select(DailyReportAudit)
            .where(
                DailyReportAudit.business_date == next_date,
                DailyReportAudit.offer_id == "offer-a",
                DailyReportAudit.action
                == "stock_continuity_after_confirmation_revert",
            )
            .order_by(DailyReportAudit.id.desc())
        )
    assert impact_audit is not None
    assert impact_audit.payload["trigger_business_date"] == REPORT_DATE.isoformat()
    assert impact_audit.payload["affected_business_date"] == next_date.isoformat()
    assert impact_audit.payload["current_ordered_units"] == 1
    assert impact_audit.payload["expected_stock_before_revert"] == 8
    assert impact_audit.payload["actual_stock"] == 7

    confirm_entry(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        source="latest",
        note="重新确认首日库存9",
        user_id=1,
    )
    resumed = next(
        item
        for item in daily_report_payload(engine, next_date)["pending_actions"]
        if item["business_date"] == next_date.isoformat()
        and item["offer_id"] == "offer-a"
    )
    assert resumed["review_issues"][0]["type"] == "stock_continuity"
    assert resumed["stock_check"]["expected_stock"] == 8
    assert resumed["stock_check"]["actual_stock"] == 7
    assert resumed["confirmation_trigger"]["confirmation_note"] == (
        "重新确认首日库存9"
    )


def test_revert_impact_todo_resolves_when_reconfirmation_restores_continuity() -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )
    confirm_entry(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        source="latest",
        note="确认首日库存9",
        user_id=1,
    )

    next_date = REPORT_DATE + timedelta(days=1)
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 8
        session.add(
            SaleItem(
                order_item_id="sale-a-revert-impact-resolves",
                order_date=datetime(2026, 7, 25, 1, tzinfo=UTC),
                sales_day=next_date,
                offer_id="offer-a",
                sku="9900000000001",
                quantity=1,
                raw_payload={},
            )
        )
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=next_date,
            slot=slot,
            captured_at=_report_capture_time(next_date, hour),
        )
    assert not any(
        item["business_date"] == next_date.isoformat()
        and item["offer_id"] == "offer-a"
        for item in daily_report_payload(engine, next_date)["pending_actions"]
    )
    assert (
        create_deadline_snapshot(
            engine,
            business_date=next_date,
            snapped_at=datetime(2026, 7, 25, 10, 30, tzinfo=UTC),
        )
        == 0
    )

    revert_confirmation(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        note="撤销后重新核对首日",
        user_id=1,
    )
    pending = daily_report_payload(engine, next_date)["pending_actions"]
    affected = next(
        item
        for item in pending
        if item["business_date"] == next_date.isoformat()
        and item["offer_id"] == "offer-a"
    )
    assert affected["review_issues"][0]["type"] == "confirmation_revert_impact"
    assert reminder_payload(engine, next_date + timedelta(days=1))["count"] == 2

    confirm_entry(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        source="latest",
        note="重新确认首日库存9",
        user_id=1,
    )
    pending = daily_report_payload(engine, next_date)["pending_actions"]
    assert not any(
        item["business_date"] == next_date.isoformat()
        and item["offer_id"] == "offer-a"
        for item in pending
    )
    assert reminder_payload(engine, next_date + timedelta(days=1))["count"] == 0


def test_export_is_blocked_until_every_entry_is_confirmed(tmp_path: Path) -> None:
    engine = _engine()
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 7
        session.get(SaleItem, "sale-a").quantity = 2
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="evening",
        captured_at=_report_capture_time(REPORT_DATE, 10),
    )
    output = tmp_path / "daily.xlsx"
    with pytest.raises(DailyReportConflictError, match="未合并"):
        export_operations_workbook(
            engine,
            business_date=REPORT_DATE,
            destination=output,
        )

    confirm_entry(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        source="evening",
        note="采用晚间库存值",
        user_id=1,
    )
    save_operator_note(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        note="平台临时调仓",
        issue_type="stock_continuity",
        user_id=1,
    )
    export_operations_workbook(
        engine,
        business_date=REPORT_DATE,
        destination=output,
    )
    workbook = load_workbook(output)
    try:
        assert workbook.sheetnames == ["运营日报", "漏爬说明"]
        report_sheet = workbook["运营日报"]
        assert report_sheet["A1"].value == "指标"
        assert report_sheet["A2"].value == "近30天浏览量"
        assert report_sheet["A3"].value == "当天订单数"
        assert report_sheet["C3"].fill.fgColor.rgb == "00FCE4D6"
        assert report_sheet["A4"].value == "平台库存数量（次日实采）"
        assert report_sheet["B4"].value is None
        assert report_sheet["A5"].value == "备注"
        assert report_sheet["B5"].value is None
        assert report_sheet["C5"].value == (
            "（确认：采用晚间库存值） （库存：平台临时调仓）"
        )
        assert "当天访客数" not in {
            report_sheet.cell(row=row, column=1).value
            for row in range(1, report_sheet.max_row + 1)
        }
    finally:
        workbook.close()


def test_delayed_inventory_note_is_exported_in_the_remark_date_cell(
    tmp_path: Path,
) -> None:
    engine = _engine()
    record_daily_report_failure(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2, 6),
        reason="Offers HTTP 403",
    )
    for slot, hour, minute in (("manual", 2, 44), ("evening", 10, 0)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour, minute),
        )

    output = export_operations_workbook(
        engine,
        business_date=REPORT_DATE,
        destination=tmp_path / "delayed-inventory-note.xlsx",
    )
    workbook = load_workbook(output)
    try:
        report_sheet = workbook["运营日报"]
        assert report_sheet["B4"].value is None
        assert "早间库存漏爬已解决" in str(report_sheet["B5"].value)
        assert "北京时间 2026-07-25 10:44:00" in str(report_sheet["B5"].value)
    finally:
        workbook.close()


def test_confirmation_and_stock_difference_notes_are_both_exported(
    tmp_path: Path,
) -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )
    next_date = REPORT_DATE + timedelta(days=1)
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 8
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=next_date,
            slot=slot,
            captured_at=_report_capture_time(next_date, hour),
        )

    confirm_entry(
        engine,
        business_date=next_date,
        offer_id="offer-a",
        source="evening",
        note="采用晚间销量",
        user_id=1,
    )
    dismiss_stock_alert(
        engine,
        business_date=next_date,
        offer_id="offer-a",
        note="平台临时调仓",
        user_id=1,
    )

    item = next(
        row
        for row in daily_report_payload(engine, next_date)["items"]
        if row["offer_id"] == "offer-a"
    )
    assert item["confirmation_baseline"]["confirm_note"] == "采用晚间销量"
    assert item["stock_check"]["dismissed"] is True
    assert item["stock_check"]["note"] == "平台临时调仓"

    output = export_operations_workbook(
        engine,
        business_date=next_date,
        destination=tmp_path / "confirmation-and-stock-note.xlsx",
    )
    workbook = load_workbook(output)
    try:
        sheet = workbook["运营日报"]
        offer_column = next(
            cell.column
            for cell in sheet[1]
            if cell.value and "9900000000001" in str(cell.value)
        )
        date_row = next(
            cell.row
            for cell in sheet["B"]
            if cell.value == next_date.isoformat()
        )
        assert sheet.cell(date_row + 3, offer_column).value == (
            "（确认：采用晚间销量） "
            "（库存差异已确认：平台临时调仓）"
        )
    finally:
        workbook.close()


def test_export_does_not_create_an_empty_workbook(tmp_path: Path) -> None:
    engine = create_engine("sqlite://")
    create_schema(engine)
    with pytest.raises(DailyReportConflictError, match="尚无可导出"):
        export_operations_workbook(
            engine,
            business_date=REPORT_DATE,
            destination=tmp_path / "empty.xlsx",
        )


def test_deadline_treats_missing_evening_capture_as_non_blocking() -> None:
    engine = _engine()
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2, 5),
    )
    assert reminder_payload(engine, REPORT_DATE)["count"] == 0
    assert create_deadline_snapshot(
        engine,
        business_date=REPORT_DATE,
        snapped_at=datetime(2026, 7, 24, 10, 30, tzinfo=UTC),
    ) == 0
    payload = daily_report_payload(engine, REPORT_DATE)
    assert payload["counts"]["missing_capture"] == 0
    assert payload["counts"]["needs_review"] == 0
    assert backfill_stock_continuity_reviews(engine, through=REPORT_DATE) == 2
    assert all(
        item["status"] != "missing_capture"
        for item in daily_report_payload(engine, REPORT_DATE)["items"]
    )
    assert reminder_payload(engine, REPORT_DATE) == {
        "count": 0,
        "dates": [],
    }


def test_failed_capture_reason_is_reported_and_does_not_require_merge(
    tmp_path: Path,
) -> None:
    engine = _engine()
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2, 5),
    )
    record_daily_report_failure(
        engine,
        business_date=REPORT_DATE,
        slot="evening",
        captured_at=datetime(2026, 7, 24, 10, tzinfo=UTC),
        reason="AuthenticationError: Takealot 登录状态失效",
        attempts=[
            {
                "attempt": 1,
                "strategy": "标准接口",
                "status": "failed",
                "reason": "Offers 失败：AuthenticationError HTTP 403",
            },
            {
                "attempt": 2,
                "strategy": "直连备用",
                "status": "failed",
                "reason": "Offers 失败：AuthenticationError HTTP 403",
            },
        ],
    )
    payload = daily_report_payload(engine, REPORT_DATE)
    assert payload["capture_status"]["evening"]["status"] == "failed"
    assert "登录状态失效" in payload["capture_status"]["evening"]["reason"]
    assert payload["capture_status"]["evening"]["attempt_count"] == 2
    assert payload["capture_status"]["evening"]["attempts"][1]["strategy"] == "直连备用"
    assert payload["counts"]["missing_capture"] == 0
    assert payload["counts"]["needs_review"] == 0
    assert all(item["status"] == "ready" for item in payload["items"])

    output = export_operations_workbook(
        engine,
        business_date=REPORT_DATE,
        destination=tmp_path / "missing-capture.xlsx",
    )
    workbook = load_workbook(output)
    try:
        issue_sheet = workbook["漏爬说明"]
        assert issue_sheet["B2"].value == "晚间"
        assert "登录状态失效" in issue_sheet["E2"].value
    finally:
        workbook.close()


def test_null_field_is_missing_data_not_a_conflict() -> None:
    engine = _engine()
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2, 5),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").page_views_30_days = None
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="evening",
        captured_at=_report_capture_time(REPORT_DATE, 10),
    )
    payload = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in payload["items"] if row["offer_id"] == "offer-a")
    assert product["status"] == "ready"
    assert product["differences"] == []
    assert product["missing_fields"] == []
    assert product["missing_capture"] is False
    assert product["current"]["page_views_30_days"] == 50
    assert payload["counts"]["missing_capture"] == 0
    assert not any(
        issue["kind"] == "product" and issue["offer_id"] == "offer-a"
        for issue in payload["capture_issues"]
    )


@pytest.mark.parametrize(
    ("field", "label"),
    (
        ("ordered_units", "当天订单数"),
        ("platform_stock", "平台库存"),
    ),
)
def test_empty_order_or_stock_is_marked_as_missing_capture(
    field: str,
    label: str,
) -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )
    with Session(engine) as session, session.begin():
        resolution = session.scalar(
            select(DailyReportResolution).where(
                DailyReportResolution.business_date == REPORT_DATE,
                DailyReportResolution.offer_id == "offer-a",
            )
        )
        assert resolution is not None
        resolution.selected_source = "manual"
        resolution.final_page_views_30_days = 50
        resolution.final_ordered_units = 0
        resolution.final_platform_stock = 9
        setattr(resolution, f"final_{field}", None)
        resolution.confirm_note = "确认当前可用字段"
        resolution.confirmed_by = 1
        resolution.confirmed_at = _report_capture_time(REPORT_DATE, 11)
        resolution.status = "confirmed"

    payload = daily_report_payload(engine, REPORT_DATE)
    product = next(row for row in payload["items"] if row["offer_id"] == "offer-a")

    assert product["status"] == "missing_capture"
    assert product["missing_capture"] is True
    assert product["missing_fields"] == [field]
    assert label in product["missing_reason"]


def test_payload_includes_recent_dates_for_vertical_comparison() -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").page_views_30_days = 58
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 8
    next_date = REPORT_DATE.replace(day=25)
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=next_date,
            slot=slot,
            captured_at=_report_capture_time(next_date, hour),
        )

    payload = daily_report_payload(engine, next_date)
    history = payload["comparison_history"]
    assert [row["business_date"] for row in history] == [
        REPORT_DATE.isoformat(),
        next_date.isoformat(),
    ]
    first_day = next(
        row for row in history[0]["items"] if row["offer_id"] == "offer-a"
    )
    second_day = next(
        row for row in history[1]["items"] if row["offer_id"] == "offer-a"
    )
    assert first_day["current"]["page_views_30_days"] == 50
    assert second_day["current"]["page_views_30_days"] == 58
    assert second_day["current"]["platform_stock"] == 8
    assert history[0]["inventory_context"]["exception_note"] is None
    assert history[1]["inventory_context"]["exception_note"] is None


def test_payload_reuses_each_historical_date_within_one_request(monkeypatch) -> None:
    engine = _engine()
    for report_date in (REPORT_DATE, REPORT_DATE + timedelta(days=1)):
        for slot, hour in (("morning", 2), ("evening", 10)):
            capture_daily_report(
                engine,
                business_date=report_date,
                slot=slot,
                captured_at=_report_capture_time(report_date, hour),
            )

    original = daily_report_module._comparison_items_for_date
    calls: list[date] = []

    def tracked_comparison_items(session, business_date):
        calls.append(business_date)
        return original(session, business_date)

    monkeypatch.setattr(
        daily_report_module,
        "_comparison_items_for_date",
        tracked_comparison_items,
    )

    payload = daily_report_payload(engine, REPORT_DATE + timedelta(days=1))

    assert len(payload["comparison_history"]) == 2
    assert calls.count(REPORT_DATE) == 1


def test_vertical_comparison_keeps_latest_thirty_data_dates() -> None:
    engine = _engine()
    for offset in range(31):
        report_date = REPORT_DATE + timedelta(days=offset)
        for slot, hour in (("morning", 2), ("evening", 10)):
            capture_daily_report(
                engine,
                business_date=report_date,
                slot=slot,
                captured_at=_report_capture_time(report_date, hour),
            )

    through = REPORT_DATE + timedelta(days=30)
    history = daily_report_payload(engine, through)["comparison_history"]

    assert len(history) == 30
    assert history[0]["business_date"] == (REPORT_DATE + timedelta(days=1)).isoformat()
    assert history[-1]["business_date"] == through.isoformat()


def test_daily_sales_align_with_the_following_morning_inventory() -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )

    next_date = REPORT_DATE + timedelta(days=1)
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 8
        session.add(
            SaleItem(
                order_item_id="sale-a-next",
                order_date=datetime(2026, 7, 25, 1, tzinfo=UTC),
                sales_day=next_date,
                offer_id="offer-a",
                sku="9900000000001",
                quantity=1,
                raw_payload={},
            )
        )
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=next_date,
            slot=slot,
            captured_at=_report_capture_time(next_date, hour),
        )

    payload = daily_report_payload(engine, next_date)
    item = next(row for row in payload["items"] if row["offer_id"] == "offer-a")
    assert item["current"]["ordered_units"] == 1
    assert item["current"]["platform_stock"] == 8
    assert item["stock_context"]["stock"] == 9
    assert item["stock_check"] == {
        "previous_stock": 9,
        "expected_stock": 8,
        "actual_stock": 8,
        "mismatch": False,
        "dismissed": False,
        "note": None,
        "resolution_action": "confirm_difference",
        "deferred_reason": None,
    }
    assert not any(
        pending["business_date"] == next_date.isoformat()
        and pending["offer_id"] == "offer-a"
        for pending in payload["pending_actions"]
    )


def test_first_capture_stock_mismatch_is_persisted_and_pushed_as_pending() -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )

    next_date = REPORT_DATE + timedelta(days=1)
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 8

    capture_daily_report(
        engine,
        business_date=next_date,
        slot="morning",
        captured_at=_report_capture_time(next_date, 2),
    )

    payload = daily_report_payload(engine, next_date)
    item = next(row for row in payload["items"] if row["offer_id"] == "offer-a")
    assert item["status"] == "needs_review"
    assert item["review_issues"] == [
        {
            "type": "stock_continuity",
            "fields": ["ordered_units", "platform_stock"],
        }
    ]
    assert item["stock_check"]["mismatch"] is True
    assert any(
        pending["business_date"] == next_date.isoformat()
        and pending["offer_id"] == "offer-a"
        for pending in payload["pending_actions"]
    )
    with Session(engine) as session:
        resolution = session.scalar(
            select(DailyReportResolution).where(
                DailyReportResolution.business_date == next_date,
                DailyReportResolution.offer_id == "offer-a",
            )
        )
        assert resolution is not None
        assert resolution.status == "needs_review"
    assert reminder_payload(engine, next_date + timedelta(days=1))["count"] == 1


def test_stock_continuity_mismatch_is_a_cross_date_pending_action(
    tmp_path: Path,
) -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )
    next_date = REPORT_DATE.replace(day=25)
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 8
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=next_date,
            slot=slot,
            captured_at=_report_capture_time(next_date, hour),
        )

    payload = daily_report_payload(engine, REPORT_DATE.replace(day=26))
    pending = next(
        item
        for item in payload["pending_actions"]
        if item["business_date"] == next_date.isoformat()
        and item["offer_id"] == "offer-a"
    )
    assert pending["status"] == "needs_review"
    assert pending["differences"] == []
    assert pending["review_issues"] == [
        {
            "type": "stock_continuity",
            "fields": ["ordered_units", "platform_stock"],
        }
    ]
    assert pending["stock_context"] == {
        "business_date": REPORT_DATE.isoformat(),
        "stock": 9,
        "source": "latest_capture",
        "source_label": "前一日报日未确认，暂用最后一次成功拉取库存",
        "selected_source": None,
        "confirmed_by": None,
        "confirmed_at": None,
        "confirm_note": None,
        "capture_label": "晚间采集（实际 07-25 18:00）",
        "continuity_ready": True,
        "version_differences": [],
    }
    assert pending["stock_check"] == {
        "previous_stock": 9,
        "expected_stock": 9,
        "actual_stock": 8,
        "mismatch": True,
        "dismissed": False,
        "note": None,
        "resolution_action": "confirm_difference",
        "deferred_reason": None,
    }
    assert reminder_payload(engine, REPORT_DATE.replace(day=26))["count"] == 1
    with pytest.raises(DailyReportConflictError, match="未合并"):
        export_operations_workbook(
            engine,
            business_date=next_date,
            destination=tmp_path / "blocked.xlsx",
        )

    dismiss_stock_alert(
        engine,
        business_date=next_date,
        offer_id="offer-a",
        note="库存包含非订单调整，人工确认采用晚间值",
        user_id=1,
    )
    with Session(engine) as session, session.begin():
        current_offer = session.get(OfferCurrent, "offer-a")
        assert current_offer is not None
        session.delete(current_offer)
    resolved = daily_report_payload(engine, REPORT_DATE.replace(day=26))
    assert not any(
        item["business_date"] == next_date.isoformat()
        and item["offer_id"] == "offer-a"
        for item in resolved["pending_actions"]
    )
    item = next(
        row
        for row in daily_report_payload(engine, next_date)["items"]
        if row["offer_id"] == "offer-a"
    )
    assert item["status"] == "ready"
    assert item["stock_check"]["mismatch"] is True
    assert item["stock_check"]["dismissed"] is True
    assert item["stock_check"]["note"] == "库存包含非订单调整，人工确认采用晚间值"
    handled = resolved["handled_actions"][0]
    assert handled["action_type"] == "stock_difference"
    assert handled["business_date"] == next_date.isoformat()
    assert handled["offer_id"] == "offer-a"
    assert handled["sku"] == "9900000000001"
    assert handled["title"] == "Product A"
    assert handled["handled_by"] == "Operator"
    assert handled["active"] is True
    assert handled["detail"]["previous_stock"] == 9
    assert handled["detail"]["ordered_units"] == 0
    assert handled["detail"]["expected_stock"] == 9
    assert handled["detail"]["actual_stock"] == 8
    exported = export_operations_workbook(
        engine,
        business_date=next_date,
        destination=tmp_path / "handled-stock-difference.xlsx",
    )
    workbook = load_workbook(exported)
    sheet = workbook["运营日报"]
    offer_column = next(
        cell.column
        for cell in sheet[1]
        if cell.value and "9900000000001" in str(cell.value)
    )
    date_row = next(
        cell.row
        for cell in sheet["B"]
        if cell.value == next_date.isoformat()
    )
    assert sheet.cell(date_row + 2, offer_column).fill.fgColor.rgb == "00FFC7CE"

    reopen_stock_alert(
        engine,
        business_date=next_date,
        offer_id="offer-a",
        note="误点确认，恢复库存差异待办",
        user_id=1,
    )
    reopened = daily_report_payload(engine, REPORT_DATE.replace(day=26))
    assert any(
        row["business_date"] == next_date.isoformat()
        and row["offer_id"] == "offer-a"
        for row in reopened["pending_actions"]
    )
    reopened_item = next(
        row
        for row in daily_report_payload(engine, next_date)["items"]
        if row["offer_id"] == "offer-a"
    )
    assert reopened_item["status"] == "needs_review"
    assert reopened_item["stock_check"]["mismatch"] is True
    assert reopened_item["stock_check"]["dismissed"] is False
    reversal = reopened["handled_actions"][0]
    assert reversal["action_type"] == "stock_alert_reopened"
    assert reversal["note"] == "误点确认，恢复库存差异待办"
    history = next(
        row
        for row in reopened["handled_actions"]
        if row["action_type"] == "stock_difference"
    )
    assert history["active"] is False
    assert history["reversal"]["kind"] == "stock_alert_reopened"
    assert history["reversal"]["handled_by"] == "Operator"
    assert history["reversal"]["note"] == "误点确认，恢复库存差异待办"
    assert reminder_payload(engine, REPORT_DATE.replace(day=26))["count"] == 1


def test_stock_review_only_allows_the_action_matching_the_manual_formula() -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )
    next_date = REPORT_DATE + timedelta(days=1)
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 8
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=next_date,
            slot=slot,
            captured_at=_report_capture_time(next_date, hour),
        )

    save_manual_candidate(
        engine,
        business_date=next_date,
        offer_id="offer-a",
        values={"platform_stock": 9},
        reason="stock_adjustment",
        note="盘点确认库存应为9",
        user_id=1,
    )
    matching = next(
        item
        for item in daily_report_payload(engine, next_date)["pending_actions"]
        if item["offer_id"] == "offer-a"
    )
    assert matching["stock_check"]["resolution_action"] == "eliminate"
    with pytest.raises(DailyReportConflictError, match="只能消除差异"):
        dismiss_stock_alert(
            engine,
            business_date=next_date,
            offer_id="offer-a",
            note="错误尝试确认差异",
            user_id=1,
        )

    eliminate_stock_alert(
        engine,
        business_date=next_date,
        offer_id="offer-a",
        note="采用盘点后的正确库存",
        user_id=1,
    )
    eliminated_payload = daily_report_payload(engine, next_date)
    eliminated = next(
        item
        for item in eliminated_payload["items"]
        if item["offer_id"] == "offer-a"
    )
    assert eliminated["status"] == "confirmed"
    assert eliminated["final"]["platform_stock"] == 9
    assert eliminated["stock_check"]["mismatch"] is False
    eliminated_audit = next(
        item
        for item in eliminated_payload["handled_actions"]
        if item["action_type"] == "stock_eliminated"
    )
    assert eliminated_audit["active"] is True
    assert eliminated_audit["detail"]["expected_stock"] == 9
    assert eliminated_audit["detail"]["actual_stock"] == 9

    revert_confirmation(
        engine,
        business_date=next_date,
        offer_id="offer-a",
        note="演练撤销消除差异",
        user_id=1,
    )
    save_manual_candidate(
        engine,
        business_date=next_date,
        offer_id="offer-a",
        values={"platform_stock": 7},
        reason="stock_adjustment",
        note="再次盘点后库存为7",
        user_id=1,
    )
    still_mismatched = next(
        item
        for item in daily_report_payload(engine, next_date)["pending_actions"]
        if item["offer_id"] == "offer-a"
    )
    assert (
        still_mismatched["stock_check"]["resolution_action"]
        == "confirm_difference"
    )
    with pytest.raises(DailyReportConflictError, match="只能确认库存差异"):
        eliminate_stock_alert(
            engine,
            business_date=next_date,
            offer_id="offer-a",
            note="错误尝试消除差异",
            user_id=1,
        )

    dismiss_stock_alert(
        engine,
        business_date=next_date,
        offer_id="offer-a",
        note="确认盘点后仍存在非订单库存差异",
        user_id=1,
    )
    confirmed_difference = next(
        item
        for item in daily_report_payload(engine, next_date)["items"]
        if item["offer_id"] == "offer-a"
    )
    assert confirmed_difference["status"] == "confirmed"
    assert confirmed_difference["final"]["platform_stock"] == 7
    assert confirmed_difference["stock_check"]["mismatch"] is True
    assert confirmed_difference["stock_check"]["dismissed"] is True


def test_continuity_waits_until_previous_day_version_difference_is_confirmed() -> None:
    engine = _engine()
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="morning",
        captured_at=_report_capture_time(REPORT_DATE, 2),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 10
        session.get(SaleItem, "sale-a").quantity = 2
    capture_daily_report(
        engine,
        business_date=REPORT_DATE,
        slot="evening",
        captured_at=_report_capture_time(REPORT_DATE, 10),
    )

    next_date = REPORT_DATE + timedelta(days=1)
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 7
        session.add(
            SaleItem(
                order_item_id="sale-a-next",
                order_date=datetime(2026, 7, 25, 1, tzinfo=UTC),
                sales_day=next_date,
                offer_id="offer-a",
                sku="9900000000001",
                quantity=1,
                raw_payload={},
            )
        )
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=next_date,
            slot=slot,
            captured_at=_report_capture_time(next_date, hour),
        )

    before = daily_report_payload(engine, next_date)
    first_day = next(
        item
        for item in before["pending_actions"]
        if item["business_date"] == REPORT_DATE.isoformat()
        and item["offer_id"] == "offer-a"
    )
    next_day = next(
        item for item in before["items"] if item["offer_id"] == "offer-a"
    )
    assert [issue["type"] for issue in first_day["review_issues"]] == [
        "capture_difference"
    ]
    assert next_day["review_issues"] == []
    assert next_day["stock_check"]["mismatch"] is False
    assert next_day["stock_check"]["deferred_reason"] == (
        "前一日报日仍有同周期版本差异，确认正确版本后再计算库存连续性"
    )

    confirm_entry(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        source="evening",
        note="确认前一日报日晚间库存10正确",
        user_id=1,
    )

    after = daily_report_payload(engine, next_date)
    next_pending = next(
        item
        for item in after["pending_actions"]
        if item["business_date"] == next_date.isoformat()
        and item["offer_id"] == "offer-a"
    )
    assert [issue["type"] for issue in next_pending["review_issues"]] == [
        "stock_continuity"
    ]
    assert next_pending["stock_check"]["previous_stock"] == 9
    assert next_pending["stock_check"]["expected_stock"] == 8
    assert next_pending["stock_check"]["actual_stock"] == 7


def test_confirmed_version_is_the_baseline_for_a_later_capture_difference() -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )
    next_date = REPORT_DATE + timedelta(days=1)
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 8
        session.add(
            SaleItem(
                order_item_id="sale-a-next",
                order_date=datetime(2026, 7, 25, 1, tzinfo=UTC),
                sales_day=next_date,
                offer_id="offer-a",
                sku="9900000000001",
                quantity=1,
                raw_payload={},
            )
        )
    capture_daily_report(
        engine,
        business_date=next_date,
        slot="morning",
        captured_at=_report_capture_time(next_date, 2),
    )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 7
        session.get(SaleItem, "sale-a-next").quantity = 2
    capture_daily_report(
        engine,
        business_date=next_date,
        slot="manual",
        captured_at=_report_capture_time(next_date, 6),
    )

    version_pending = next(
        item
        for item in daily_report_payload(engine, next_date)["pending_actions"]
        if item["business_date"] == next_date.isoformat()
        and item["offer_id"] == "offer-a"
    )
    assert [issue["type"] for issue in version_pending["review_issues"]] == [
        "capture_difference"
    ]
    assert version_pending["stock_check"]["mismatch"] is False

    confirm_entry(
        engine,
        business_date=next_date,
        offer_id="offer-a",
        source="latest",
        note="确认手动刷新库存7为正确值",
        user_id=1,
    )
    continuity_pending = next(
        item
        for item in daily_report_payload(engine, next_date)["pending_actions"]
        if item["business_date"] == next_date.isoformat()
        and item["offer_id"] == "offer-a"
    )
    assert [issue["type"] for issue in continuity_pending["review_issues"]] == [
        "stock_continuity"
    ]
    assert continuity_pending["current"]["platform_stock"] == 8
    assert continuity_pending["confirmation_baseline"]["values"]["platform_stock"] == 8

    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 6
        session.get(SaleItem, "sale-a-next").quantity = 3
    late_capture_time = datetime.now(UTC) + timedelta(minutes=1)
    capture_daily_report(
        engine,
        business_date=next_date,
        slot="evening",
        captured_at=late_capture_time,
    )

    reopened = next(
        item
        for item in daily_report_payload(engine, next_date)["pending_actions"]
        if item["business_date"] == next_date.isoformat()
        and item["offer_id"] == "offer-a"
    )
    assert [issue["type"] for issue in reopened["review_issues"]] == [
        "capture_difference"
    ]
    assert reopened["stock_check"]["mismatch"] is False
    assert [
        (version["kind"], version["values"]["ordered_units"])
        for version in reopened["review_versions"]
    ] == [
        ("confirmed", 2),
        ("capture", 3),
    ]

    confirm_entry(
        engine,
        business_date=next_date,
        offer_id="offer-a",
        source="latest",
        note="确认晚间库存6为新的正确值",
        user_id=1,
    )
    final_pending = next(
        item
        for item in daily_report_payload(engine, next_date)["pending_actions"]
        if item["business_date"] == next_date.isoformat()
        and item["offer_id"] == "offer-a"
    )
    assert [issue["type"] for issue in final_pending["review_issues"]] == [
        "stock_continuity"
    ]
    assert final_pending["current"]["platform_stock"] == 8
    assert final_pending["stock_check"]["expected_stock"] == 6
    assert final_pending["stock_check"]["actual_stock"] == 8


def test_manual_confirmation_reopens_following_stock_conflict_and_keeps_context() -> None:
    engine = _engine()
    second_date = REPORT_DATE + timedelta(days=1)
    third_date = REPORT_DATE + timedelta(days=2)
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 8
        session.add(
            SaleItem(
                order_item_id="sale-a-day-2",
                order_date=datetime(2026, 7, 25, 1, tzinfo=UTC),
                sales_day=second_date,
                offer_id="offer-a",
                sku="9900000000001",
                quantity=1,
                raw_payload={},
            )
        )
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=second_date,
            slot=slot,
            captured_at=_report_capture_time(second_date, hour),
        )
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 7
        session.add(
            SaleItem(
                order_item_id="sale-a-day-3",
                order_date=datetime(2026, 7, 26, 1, tzinfo=UTC),
                sales_day=third_date,
                offer_id="offer-a",
                sku="9900000000001",
                quantity=1,
                raw_payload={},
            )
        )
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=third_date,
            slot=slot,
            captured_at=_report_capture_time(third_date, hour),
        )
    assert (
        create_deadline_snapshot(
            engine,
            business_date=second_date,
            snapped_at=datetime(2026, 7, 25, 10, 30, tzinfo=UTC),
        )
        == 0
    )

    save_manual_candidate(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        values={"platform_stock": 10},
        reason="stock_adjustment",
        note="盘点后确认首日库存为10",
        user_id=1,
    )
    confirm_entry(
        engine,
        business_date=REPORT_DATE,
        offer_id="offer-a",
        source="manual",
        note="采用人工盘点库存10",
        user_id=1,
    )

    second_pending = next(
        item
        for item in daily_report_payload(engine, third_date)["pending_actions"]
        if item["business_date"] == second_date.isoformat()
        and item["offer_id"] == "offer-a"
    )
    assert second_pending["stock_check"]["mismatch"] is True
    assert reminder_payload(engine, second_date)["count"] == 1
    assert second_pending["confirmation_trigger"] == {
        "kind": "previous_confirmation",
        "message": (
            "2026-07-24 人工确认合并后，触发 2026-07-25 库存连续性冲突"
        ),
        "trigger_business_date": "2026-07-24",
        "affected_business_date": "2026-07-25",
        "confirmation_source": "manual",
        "confirmation_source_label": "人工修改值",
        "confirmed_by": "Operator",
        "confirmed_at": second_pending["confirmation_trigger"]["confirmed_at"],
        "confirmation_note": "采用人工盘点库存10",
        "previous_stock_before_confirmation": 9,
        "confirmed_previous_stock": 10,
        "current_ordered_units": 1,
        "expected_stock_before_confirmation": 8,
        "comparison_before_state": "matched",
        "expected_stock_after_confirmation": 9,
        "actual_stock": 8,
        "affected_previous_status": "ready",
        "affected_previous_final": None,
        "affected_previous_confirmed_by": None,
        "affected_previous_confirmed_at": None,
        "affected_previous_confirm_note": None,
        "affected_current_values": {
            "page_views_30_days": 50,
            "ordered_units": 1,
            "platform_stock": 8,
        },
    }

    save_manual_candidate(
        engine,
        business_date=second_date,
        offer_id="offer-a",
        values={"platform_stock": 9},
        reason="stock_adjustment",
        note="按前一日确认值修正第二日库存",
        user_id=1,
    )
    confirm_entry(
        engine,
        business_date=second_date,
        offer_id="offer-a",
        source="manual",
        note="采用修正后的第二日库存9",
        user_id=1,
    )

    pending = daily_report_payload(engine, third_date)["pending_actions"]
    assert not any(
        item["business_date"] == second_date.isoformat()
        and item["offer_id"] == "offer-a"
        for item in pending
    )
    third_pending = next(
        item
        for item in pending
        if item["business_date"] == third_date.isoformat()
        and item["offer_id"] == "offer-a"
    )
    assert third_pending["stock_check"] == {
        "previous_stock": 9,
        "expected_stock": 8,
        "actual_stock": 7,
        "mismatch": True,
        "dismissed": False,
        "note": None,
        "resolution_action": "confirm_difference",
        "deferred_reason": None,
    }
    assert third_pending["confirmation_trigger"]["trigger_business_date"] == (
        second_date.isoformat()
    )
    assert third_pending["confirmation_trigger"]["affected_business_date"] == (
        third_date.isoformat()
    )
    assert third_pending["confirmation_trigger"]["previous_stock_before_confirmation"] == 8
    assert third_pending["confirmation_trigger"]["confirmed_previous_stock"] == 9
    assert reminder_payload(engine, third_date + timedelta(days=1))["count"] == 1


def test_backfill_promotes_existing_stock_mismatch_to_review() -> None:
    engine = _engine()
    for slot, hour in (("morning", 2), ("evening", 10)):
        capture_daily_report(
            engine,
            business_date=REPORT_DATE,
            slot=slot,
            captured_at=_report_capture_time(REPORT_DATE, hour),
        )
    next_date = REPORT_DATE.replace(day=25)
    with Session(engine) as session, session.begin():
        session.get(OfferCurrent, "offer-a").takealot_available_stock = 8
    capture_daily_report(
        engine,
        business_date=next_date,
        slot="morning",
        captured_at=_report_capture_time(next_date, 2),
    )
    with Session(engine) as session, session.begin():
        resolution = session.scalar(
            select(DailyReportResolution).where(
                DailyReportResolution.business_date == next_date,
                DailyReportResolution.offer_id == "offer-a",
            )
        )
        assert resolution is not None
        resolution.status = "ready"
    capture_daily_report(
        engine,
        business_date=next_date,
        slot="evening",
        captured_at=_report_capture_time(next_date, 10),
    )
    with Session(engine) as session, session.begin():
        resolution = session.scalar(
            select(DailyReportResolution).where(
                DailyReportResolution.business_date == next_date,
                DailyReportResolution.offer_id == "offer-a",
            )
        )
        assert resolution is not None
        resolution.status = "ready"

    assert backfill_stock_continuity_reviews(engine, through=next_date) == 1
    payload = daily_report_payload(engine, next_date)
    item = next(row for row in payload["items"] if row["offer_id"] == "offer-a")
    assert item["status"] == "needs_review"


def test_future_capture_slots_are_pending_not_missing() -> None:
    engine = _engine()
    future_business_date = date.today()

    payload = daily_report_payload(engine, future_business_date)

    assert payload["capture_status"]["morning"]["status"] == "pending"
    assert payload["capture_status"]["evening"]["status"] == "pending"
    assert payload["capture_status"]["pre_close"]["status"] == "pending"
    assert payload["capture_issues"] == []
    assert payload["counts"]["missing_capture"] == 0
