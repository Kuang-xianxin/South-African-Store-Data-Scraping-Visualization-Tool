from __future__ import annotations

from datetime import UTC, date, datetime
import importlib.util
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from takealot_ops.nft102 import build_update_payload, read_product_columns
from takealot_ops.storage.models import Base, CollectionRun, OfferSnapshot, SaleItem


def _load_writer_module():
    path = Path(__file__).parents[2] / "scripts" / "write_nft102_workbook.py"
    spec = importlib.util.spec_from_file_location("write_nft102_workbook", path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _template(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "NFT102"
    worksheet.cell(1, 1, "访客")
    worksheet.cell(1, 2, "日期")
    worksheet.cell(1, 3, "商品A\n9902253460734")
    worksheet.cell(1, 4, "无SKU商品")
    worksheet.cell(1, 5, "旧列\n9902253460734")
    worksheet.cell(2, 1, "访客总数")
    worksheet.cell(2, 2, date(2026, 7, 20))
    worksheet.cell(3, 1, "当天访客数")
    worksheet.cell(4, 1, "当天订单数")
    worksheet.cell(4, 2, "=SUM(C4:E4)")
    worksheet.cell(4, 3, 5)
    worksheet.cell(4, 3).fill = PatternFill(
        fill_type="solid", fgColor="FFFBE5D6"
    )
    worksheet.cell(5, 1, "平台库存数量")
    worksheet["E7"].number_format = "0"
    other = workbook.create_sheet("其他店铺")
    other["A1"] = "不得改变"
    workbook.save(path)


def test_read_product_columns_extracts_13_digit_skus(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    _template(template)

    columns, max_column = read_product_columns(template)

    assert max_column == 5
    assert [column.sku for column in columns] == ["9902253460734", None, "9902253460734"]


def test_payload_uses_today_traffic_and_yesterday_sales_without_faking_fields(
    tmp_path: Path,
) -> None:
    template = tmp_path / "template.xlsx"
    _template(template)
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    report_date = date(2026, 7, 21)
    now = datetime(2026, 7, 21, 8, tzinfo=UTC)

    with Session(engine) as session:
        session.add(
            CollectionRun(
                run_id="offer-run",
                run_type="offers",
                scope_date=report_date,
                started_at=now,
                finished_at=now,
                status="success",
                counts={"records": 1},
                error=None,
            )
        )
        session.add(
            OfferSnapshot(
                snapshot_date=report_date,
                offer_id="123",
                sku="9902253460734",
                captured_at=now,
                page_views_30_days=88,
                total_stock=7,
                takealot_available_stock=7,
            )
        )
        session.add(
            SaleItem(
                order_item_id="order-item",
                order_date=now,
                sales_day=date(2026, 7, 20),
                sku="9902253460734",
                quantity=3,
                raw_payload={},
            )
        )
        session.commit()

        payload = build_update_payload(
            session, template, report_date, sales_data_complete=True
        )

    first, no_sku, active_duplicate = payload["columns"]
    assert first["active"] is False
    assert first["traffic_value"] is None
    assert no_sku["active"] is False
    assert active_duplicate["traffic_value"] == 88
    assert active_duplicate["order_value"] == 3
    assert active_duplicate["platform_stock_value"] == 7
    assert payload["sales_date"] == "2026-07-20"
    assert payload["field_policy"]["当天访客数"].endswith("留空")
    assert "quantity_available" in payload["field_policy"]["平台库存数量"]


def test_writer_changes_only_nft102_xml_and_appends_four_rows(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    _template(template)
    writer = _load_writer_module()
    payload = {
        "sheet_name": "NFT102",
        "report_date": "2026-07-21",
        "max_column_letter": "E",
        "columns": [
            {
                "column_letter": "C",
                "traffic_value": 10,
                "order_value": 0,
                "platform_stock_value": 4,
            },
            {
                "column_letter": "D",
                "traffic_value": None,
                "order_value": 2,
                "platform_stock_value": None,
            },
            {
                "column_letter": "E",
                "traffic_value": 20,
                "order_value": 3,
                "platform_stock_value": 6,
            },
        ],
        "summary": {"ordered_units_mapped": 5},
    }

    writer.patch_workbook(template, output, payload)

    with zipfile.ZipFile(template) as source_zip, zipfile.ZipFile(output) as output_zip:
        changed = [
            name
            for name in source_zip.namelist()
            if source_zip.read(name) != output_zip.read(name)
        ]
    assert changed == ["xl/worksheets/sheet1.xml"]

    workbook = load_workbook(output, data_only=False)
    try:
        worksheet = workbook["NFT102"]
        assert worksheet["B6"].value.date() == date(2026, 7, 21)
        assert worksheet["C6"].value == 10
        assert worksheet["C7"].value is None
        assert worksheet["C8"].value == 0
        assert worksheet["C8"].fill.fgColor.rgb != "FFFBE5D6"
        assert worksheet["D8"].value == 2
        assert worksheet["D8"].fill.fgColor.rgb == "FFFBE5D6"
        assert worksheet["C9"].value == 4
        assert worksheet["E9"].value == 6
        assert worksheet["B8"].value == "=SUM(C8:E8)"
        assert worksheet.max_row == 11
        assert worksheet["E11"].number_format == "0"
        assert workbook["其他店铺"]["A1"].value == "不得改变"
    finally:
        workbook.close()
