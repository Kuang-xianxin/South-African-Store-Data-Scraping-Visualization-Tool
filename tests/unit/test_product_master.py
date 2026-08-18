from __future__ import annotations

import re
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterator, Sequence
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook
from sqlalchemy import Engine, func, select
from sqlalchemy.orm import Session

from takealot_ops.product_master import (
    ProductMasterConflictError,
    ProductMasterImportService,
    ProductMasterInputError,
    enrich_product_master_records,
    load_company_inventory_for_plid,
    load_product_master_links,
    parse_product_master_workbook,
)
from takealot_ops.storage.migrations import create_engine_for_database_url, create_schema
from takealot_ops.storage.models import (
    CompanyProduct,
    CompanyProductCost,
    ErpStore,
    LogisticsProviderSnapshot,
    OfferCurrent,
    PlatformSkuMapping,
    ProductMasterImportBatch,
    ProductMasterImportRow,
)


@pytest.fixture
def engine(tmp_path: Path) -> Iterator[Engine]:
    database_engine = create_engine_for_database_url(
        f"sqlite:///{(tmp_path / 'product-master.db').as_posix()}"
    )
    create_schema(database_engine)
    yield database_engine
    database_engine.dispose()


def _write_workbook(
    path: Path,
    rows: Sequence[Sequence[object]],
    *,
    headers: Sequence[str] = ("店铺", "平台SKU", "自编sku", "中文", "成本（RMB）"),
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "主数据"
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))
    workbook.save(path)
    workbook.close()
    return path


def _seed_offer(
    engine: Engine,
    *,
    store_code: str,
    offer_id: str,
    platform_sku: str,
) -> None:
    with Session(engine) as session, session.begin():
        session.add(
            OfferCurrent(
                store_code=store_code,
                offer_id=offer_id,
                sku=platform_sku,
                productline_id=f"plid-{offer_id}",
                title=f"Product {offer_id}",
                captured_at=datetime(2026, 8, 17, 3, 0),
            )
        )


def test_parser_splits_multi_sku_cells_and_quantizes_rmb_cost(tmp_path: Path) -> None:
    source = _write_workbook(
        tmp_path / "source.xlsx",
        [
            ["C1", "9900000000001\n9900000000002", "INT-Blue", "蓝色产品", 12.34567],
            ["C2", 9900000000003, "INT-RED", "红色产品", "RMB 8.5"],
        ],
    )

    parsed = parse_product_master_workbook(
        source,
        effective_date=date(2026, 8, 15),
    )

    assert parsed.sheet_name == "主数据"
    assert parsed.source_row_count == 2
    assert [row.platform_sku for row in parsed.rows] == [
        "9900000000001",
        "9900000000002",
        "9900000000003",
    ]
    assert [row.cost_rmb for row in parsed.rows] == [
        Decimal("12.3457"),
        Decimal("12.3457"),
        Decimal("8.5000"),
    ]
    assert parsed.warnings == ()


def test_parser_accepts_xlsx_without_worksheet_dimension_metadata(
    tmp_path: Path,
) -> None:
    source = _write_workbook(
        tmp_path / "with-dimension.xlsx",
        [["C1", "9900000000001", "INT-1", "产品", 12.5]],
    )
    without_dimension = tmp_path / "without-dimension.xlsx"
    with ZipFile(source) as input_archive, ZipFile(
        without_dimension,
        "w",
        compression=ZIP_DEFLATED,
    ) as output_archive:
        for member in input_archive.infolist():
            content = input_archive.read(member.filename)
            if member.filename == "xl/worksheets/sheet1.xml":
                content, replacements = re.subn(
                    rb'<dimension ref="[^"]+"\s*/>',
                    b"",
                    content,
                    count=1,
                )
                assert replacements == 1
            output_archive.writestr(member, content)

    parsed = parse_product_master_workbook(
        without_dimension,
        effective_date=date(2026, 8, 15),
    )

    assert parsed.source_row_count == 1
    assert parsed.rows[0].platform_sku == "9900000000001"
    assert parsed.rows[0].cost_rmb == Decimal("12.5000")


def test_parser_rejects_conflicting_costs_for_one_company_sku(tmp_path: Path) -> None:
    source = _write_workbook(
        tmp_path / "conflict.xlsx",
        [
            ["C1", "9900000000001", "INT-1", "同一产品", 10],
            ["C2", "9900000000002", "INT-1", "同一产品", 11],
        ],
    )

    with pytest.raises(ProductMasterInputError, match="同一批次出现多个 RMB 成本"):
        parse_product_master_workbook(
            source,
            effective_date=date(2026, 8, 15),
        )


def test_bare_monthly_purchase_expense_is_not_treated_as_unit_cost(tmp_path: Path) -> None:
    source = _write_workbook(
        tmp_path / "monthly-expense.xlsx",
        [["C1", "9900000000001", "INT-1", "产品", 520.33]],
        headers=("店铺", "平台SKU", "自编sku", "中文", "采购费用"),
    )

    parsed = parse_product_master_workbook(
        source,
        effective_date=date(2026, 8, 1),
    )

    assert parsed.rows[0].cost_rmb is None
    assert any("只导入产品与 SKU 映射" in warning for warning in parsed.warnings)


def test_formatted_blank_tail_does_not_count_as_source_rows(tmp_path: Path) -> None:
    source = tmp_path / "formatted-tail.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "主数据"
    worksheet.append(["店铺", "平台SKU", "自编sku", "中文", "成本（RMB）"])
    worksheet.append(["C1", "9900000000001", "INT-1", "产品", 10])
    worksheet.cell(row=1_048_576, column=1).number_format = "@"
    workbook.save(source)
    workbook.close()

    parsed = parse_product_master_workbook(
        source,
        effective_date=date(2026, 8, 15),
    )

    assert parsed.source_row_count == 1
    assert len(parsed.rows) == 1


def test_import_is_audited_idempotent_and_matches_offer_current(
    engine: Engine,
    tmp_path: Path,
) -> None:
    _seed_offer(
        engine,
        store_code="current",
        offer_id="offer-1",
        platform_sku="9900000000001",
    )
    _seed_offer(
        engine,
        store_code="store-02",
        offer_id="offer-2",
        platform_sku="9900000000002",
    )
    source = _write_workbook(
        tmp_path / "source.xlsx",
        [
            ["C1", "9900000000001\n9900000000002", "INT-1", "产品一", 12.5],
            ["C3", "9900000000003", "INT-2", "产品二", 25],
        ],
    )
    service = ProductMasterImportService(engine)

    preview = service.preview(source, effective_date=date(2026, 8, 15))
    assert preview.mode == "preview"
    assert preview.product_count == 2
    assert preview.mapping_count == 3
    assert preview.cost_count == 2
    assert preview.matched_offer_count == 2
    assert preview.unmatched_offer_count == 1
    assert preview.mapping_actions == {"created": 3}

    imported = service.import_workbook(
        source,
        effective_date=date(2026, 8, 15),
        imported_by="finance-user",
    )
    repeated = service.import_workbook(
        source,
        effective_date=date(2026, 8, 15),
        imported_by="finance-user",
    )

    assert imported.mode == "imported"
    assert repeated.mode == "already_imported"
    assert repeated.batch_id == imported.batch_id
    with Session(engine) as session:
        assert session.scalar(select(func.count()).select_from(CompanyProduct)) == 2
        assert session.scalar(select(func.count()).select_from(PlatformSkuMapping)) == 3
        assert session.scalar(select(func.count()).select_from(CompanyProductCost)) == 2
        assert session.scalar(select(func.count()).select_from(ProductMasterImportBatch)) == 1
        assert session.scalar(select(func.count()).select_from(ProductMasterImportRow)) == 3
        mapped = session.scalar(
            select(PlatformSkuMapping).where(
                PlatformSkuMapping.platform_sku == "9900000000001"
            )
        )
        batch = session.get(ProductMasterImportBatch, imported.batch_id)
        links = load_product_master_links(
            session,
            platform_skus=["9900000000001", "9900000000002"],
            as_of_date=date(2026, 8, 15),
        )

    assert mapped is not None
    assert mapped.resolved_store_code == "current"
    assert mapped.resolved_offer_id == "offer-1"
    assert batch is not None
    assert batch.imported_by == "finance-user"
    assert batch.source_file_name == "source.xlsx"
    assert batch.summary is not None
    assert batch.summary["mode"] == "imported"
    assert links["9900000000001"].company_sku == "INT-1"
    assert links["9900000000002"].company_sku == "INT-1"
    assert links["9900000000001"].cost_rmb == Decimal("12.5000")
    assert links["9900000000001"].cost_effective_date == date(2026, 8, 15)


def test_unchanged_cost_does_not_duplicate_history(
    engine: Engine,
    tmp_path: Path,
) -> None:
    first = _write_workbook(
        tmp_path / "first.xlsx",
        [["C1", "9900000000001", "INT-1", "产品一", 10]],
    )
    second = _write_workbook(
        tmp_path / "second.xlsx",
        [["C1", "9900000000001", "INT-1", "产品一", 10]],
    )
    service = ProductMasterImportService(engine)

    service.import_workbook(
        first,
        effective_date=date(2026, 8, 1),
        imported_by="finance-user",
    )
    report = service.import_workbook(
        second,
        effective_date=date(2026, 8, 15),
        imported_by="finance-user",
    )

    assert report.cost_actions == {"unchanged": 1}
    with Session(engine) as session:
        assert session.scalar(select(func.count()).select_from(CompanyProductCost)) == 1
        assert session.scalar(select(func.count()).select_from(ProductMasterImportBatch)) == 2


def test_reassignment_requires_explicit_confirmation_and_keeps_row_audit(
    engine: Engine,
    tmp_path: Path,
) -> None:
    first = _write_workbook(
        tmp_path / "first.xlsx",
        [["C1", "9900000000001", "INT-A", "产品 A", 10]],
    )
    second = _write_workbook(
        tmp_path / "second.xlsx",
        [["C1", "9900000000001", "INT-B", "产品 B", 20]],
    )
    service = ProductMasterImportService(engine)
    service.import_workbook(
        first,
        effective_date=date(2026, 8, 1),
        imported_by="operations-user",
    )

    preview = service.preview(second, effective_date=date(2026, 8, 15))
    assert preview.mapping_actions == {"reassigned": 1}
    assert preview.reassignments == (
        {
            "platform_sku": "9900000000001",
            "from_company_sku": "INT-A",
            "to_company_sku": "INT-B",
        },
    )
    with pytest.raises(ProductMasterConflictError, match="默认拒绝静默覆盖"):
        service.import_workbook(
            second,
            effective_date=date(2026, 8, 15),
            imported_by="operations-user",
        )

    imported = service.import_workbook(
        second,
        effective_date=date(2026, 8, 15),
        imported_by="operations-user",
        allow_reassignments=True,
    )

    with Session(engine) as session:
        mapping, product = session.execute(
            select(PlatformSkuMapping, CompanyProduct)
            .join(CompanyProduct, CompanyProduct.id == PlatformSkuMapping.company_product_id)
            .where(PlatformSkuMapping.platform_sku == "9900000000001")
        ).one()
        audit = session.scalar(
            select(ProductMasterImportRow).where(
                ProductMasterImportRow.batch_id == imported.batch_id
            )
        )

    assert mapping.company_product_id == product.id
    assert product.company_sku == "INT-B"
    assert audit is not None
    assert audit.previous_company_sku == "INT-A"
    assert audit.mapping_action == "reassigned"


def test_record_enrichment_exposes_company_sku_and_marks_unmapped(
    engine: Engine,
    tmp_path: Path,
) -> None:
    source = _write_workbook(
        tmp_path / "enrich.xlsx",
        [["C1", "9900000000001", "INT-1", "产品一", 12.5]],
    )
    ProductMasterImportService(engine).import_workbook(
        source,
        effective_date=date(2026, 8, 15),
        imported_by="finance-user",
    )

    with Session(engine) as session:
        records = enrich_product_master_records(
            session,
            [
                {"sku": "9900000000001", "title": "mapped"},
                {"sku": "9900000000999", "title": "unmapped"},
            ],
        )

    assert records[0]["company_sku"] == "INT-1"
    assert records[0]["company_product_name"] == "产品一"
    assert records[0]["cost_rmb"] == 12.5
    assert records[1]["company_sku"] is None
    assert records[1]["company_product_name"] is None


def test_company_inventory_uses_authorized_platform_offers_and_shared_w8_once(
    engine: Engine,
    tmp_path: Path,
) -> None:
    now = datetime(2026, 8, 17, 6, 0, tzinfo=UTC)
    with Session(engine) as session, session.begin():
        session.add_all(
            [
                ErpStore(
                    code="store-01",
                    display_name="店铺一",
                    active=True,
                    data_connected=True,
                    created_at=now,
                    updated_at=now,
                ),
                ErpStore(
                    code="store-02",
                    display_name="店铺二",
                    active=True,
                    data_connected=True,
                    created_at=now,
                    updated_at=now,
                ),
                ErpStore(
                    code="store-03",
                    display_name="未授权店铺",
                    active=True,
                    data_connected=True,
                    created_at=now,
                    updated_at=now,
                ),
            ]
        )
    offer_rows = [
        ("store-01", "offer-1", "9900000000001", "100", 5, 2, 1),
        ("store-02", "offer-2", "9900000000002", "other-plid", 7, 3, 2),
        ("store-03", "offer-3", "9900000000002", "other-plid", 99, 99, 99),
    ]
    for store_code, offer_id, sku, plid, available, on_way, receiving in offer_rows:
        with Session(engine) as session, session.begin():
            session.add(
                OfferCurrent(
                    store_code=store_code,
                    offer_id=offer_id,
                    sku=sku,
                    productline_id=plid,
                    title=offer_id,
                    status="buyable",
                    takealot_available_stock=available,
                    takealot_stock_on_way=on_way,
                    takealot_stock_in_receiving=receiving,
                    captured_at=now,
                )
            )
    source = _write_workbook(
        tmp_path / "inventory.xlsx",
        [["C1", "9900000000001\n9900000000002", "INT-1", "产品一", 12.5]],
    )
    ProductMasterImportService(engine).import_workbook(
        source,
        effective_date=date(2026, 8, 15),
        imported_by="finance-user",
    )
    with Session(engine) as session, session.begin():
        session.add(
            LogisticsProviderSnapshot(
                store_code="store-02",
                provider="w8",
                fetched_at=now,
                payload={
                    "connected": True,
                    "warehouse": {"code": "CRZA", "name": "南非仓"},
                    "inventory_detail_available": True,
                    "inventory_items": [
                        {
                            "company_sku": "int-1",
                            "stock_total": 20,
                            "usable_stock": 15,
                            "locked_stock": 5,
                            "outbound_allocated": 3,
                            "transit_stock": 4,
                            "defective_stock": 1,
                        },
                        {
                            "company_sku": "INT-1-EXTRA",
                            "stock_total": 999,
                        },
                    ],
                },
            )
        )

    payload = load_company_inventory_for_plid(
        engine,
        plid="100",
        store_codes={"store-01", "store-02"},
    )

    assert payload["w8_shared_once"] is True
    assert payload["stage_totals_are_additive"] is False
    assert payload["company_sku_count"] == 1
    item = payload["items"][0]
    assert item["company_sku"] == "INT-1"
    assert item["overseas_warehouse"]["record_count"] == 1
    assert item["overseas_warehouse"]["stages"]["stock_total"] == {
        "value": 20,
        "coverage": 1,
    }
    assert item["platform_warehouse"]["stages"]["available"] == {
        "value": 12,
        "coverage": 2,
    }
    assert {row["store_code"] for row in item["platform_warehouse"]["offers"]} == {
        "store-01",
        "store-02",
    }
