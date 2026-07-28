"""Versioned operations daily report with human reconciliation and audit history."""

from __future__ import annotations

from collections.abc import Mapping
from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta
from pathlib import Path
from typing import Any
from uuid import uuid4
from zoneinfo import ZoneInfo

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from sqlalchemy import Engine, func, select
from sqlalchemy.orm import Session

from takealot_ops.storage.models import (
    DailyInventorySnapshot,
    DailyReportAudit,
    DailyReportDeadlineSnapshot,
    DailyReportObservation,
    DailyReportResolution,
    DailyReportRun,
    ErpUser,
    OfferCurrent,
    SaleItem,
)


SCHEDULED_SLOTS = frozenset({"morning", "evening"})
CAPTURE_SLOTS = frozenset({"morning", "evening", "manual"})
SOURCES = frozenset({"morning", "evening", "latest", "manual"})
MANUAL_REASONS = frozenset({"platform_delay", "stock_adjustment", "other"})
NOTE_ISSUE_TYPES = frozenset({"general", "capture_difference", "stock_continuity"})
NOTE_AUDIT_ACTIONS = frozenset(
    {"operator_note", "operator_note_updated", "operator_note_deleted"}
)
HANDLED_ACTIONS = frozenset(
    {
        "confirm",
        "bulk_confirm",
        "dismiss_stock_alert",
        "eliminate_stock_alert",
    }
)
HANDLED_REVERSALS = frozenset(
    {"confirmation_reverted", "stock_alert_reopened"}
)
HANDLED_SUPPORT_ACTIONS = frozenset({"manual_candidate"}) | NOTE_AUDIT_ACTIONS
HANDLED_AUDIT_ACTIONS = (
    HANDLED_ACTIONS | HANDLED_REVERSALS | HANDLED_SUPPORT_ACTIONS
)
OPEN_STATUSES = frozenset({"needs_review"})
EXPORTABLE_STATUSES = frozenset({"ready", "confirmed", "missing_capture"})


@dataclass(frozen=True)
class ReportCaptureResult:
    run_id: str
    business_date: date
    slot: str
    product_count: int
    reopened_count: int


@dataclass(frozen=True)
class InventorySnapshotBackfillResult:
    snapshots_created: int
    observations_updated: int
    observations_missing_snapshot: int


class DailyReportInputError(ValueError):
    """Safe validation failure for an operator action."""


class DailyReportConflictError(RuntimeError):
    """Safe state conflict for an operator action."""


def record_daily_report_failure(
    engine: Engine,
    *,
    business_date: date,
    slot: str,
    captured_at: datetime,
    reason: str,
    attempts: list[dict[str, object]] | None = None,
) -> str:
    """Persist a failed scheduled capture so operators can see why data is missing."""
    if slot not in CAPTURE_SLOTS:
        raise DailyReportInputError("采集时段只能是 morning、evening 或 manual")
    now = _naive_utc(captured_at)
    run_id = str(uuid4())
    clean_reason = reason.strip() or "采集任务失败，但没有返回具体错误"
    attempt_rows = attempts or []
    with Session(engine) as session, session.begin():
        session.add(
            DailyReportRun(
                run_id=run_id,
                business_date=business_date,
                slot=slot,
                captured_at=now,
                status="failed",
                counts={
                    "products": 0,
                    "missing_reason": clean_reason,
                    "final_reason": clean_reason,
                    "attempts": attempt_rows,
                    "attempt_count": len(attempt_rows),
                },
                created_at=now,
            )
        )
        if slot == "evening":
            rows = list(
                session.scalars(
                    select(DailyReportResolution).where(
                        DailyReportResolution.business_date == business_date,
                        DailyReportResolution.status == "awaiting_evening",
                    )
                )
            )
            for row in rows:
                row.status = "missing_capture"
                row.updated_at = now
        _audit(
            session,
            business_date,
            None,
            "capture_failed",
            {
                "slot": slot,
                "reason": clean_reason,
                "attempt_count": len(attempt_rows),
            },
            clean_reason,
            None,
            now,
        )
    return run_id


def operations_business_date(captured_at: datetime) -> date:
    """Return the report day for the Beijing 10:00-to-10:00 comparison cycle."""
    if captured_at.tzinfo is None:
        raise ValueError("captured_at 必须包含时区")
    china_time = captured_at.astimezone(ZoneInfo("Asia/Shanghai"))
    cycle_start = china_time.date()
    if china_time.time() < time(10, 0):
        cycle_start -= timedelta(days=1)
    return cycle_start - timedelta(days=1)


def capture_daily_report(
    engine: Engine,
    *,
    business_date: date,
    slot: str,
    captured_at: datetime,
    capture_details: Mapping[str, object] | None = None,
) -> ReportCaptureResult:
    """Freeze one sales version and attach the next morning's 10:05 stock."""
    if slot not in CAPTURE_SLOTS:
        raise DailyReportInputError("采集时段只能是 morning、evening 或 manual")
    now = _naive_utc(captured_at)
    run_id = str(uuid4())
    reopened = 0
    with Session(engine) as session, session.begin():
        offers = list(
            session.scalars(select(OfferCurrent).order_by(OfferCurrent.offer_id))
        )
        sales: dict[str, int] = {}
        sale_rows = session.execute(
            select(
                SaleItem.offer_id,
                func.coalesce(func.sum(SaleItem.quantity), 0),
            )
            .where(
                SaleItem.sales_day == business_date,
                SaleItem.offer_id.is_not(None),
            )
            .group_by(SaleItem.offer_id)
        ).all()
        for sale_offer_id, quantity in sale_rows:
            if sale_offer_id is not None:
                sales[sale_offer_id] = int(quantity)
        counts: dict[str, object] = {"products": len(offers)}
        if capture_details:
            counts.update(capture_details)
        run = DailyReportRun(
            run_id=run_id,
            business_date=business_date,
            slot=slot,
            captured_at=now,
            status="success",
            counts=dict(counts),
            created_at=now,
        )
        session.add(run)
        session.flush()
        if slot == "morning":
            inventory_date, created_inventory_snapshots = (
                _freeze_morning_inventory_snapshots(
                    session,
                    run=run,
                    offers=offers,
                )
            )
            counts["captured_inventory_date"] = inventory_date.isoformat()
            counts["captured_inventory_snapshots"] = created_inventory_snapshots
        reported_inventory_date = business_date + timedelta(days=1)
        inventory_snapshots = _daily_inventory_snapshot_map(
            session,
            reported_inventory_date,
        )
        counts["reported_inventory_date"] = reported_inventory_date.isoformat()
        counts["reported_inventory_snapshots"] = len(inventory_snapshots)
        counts["reported_inventory_missing"] = sum(
            1
            for offer in offers
            if (
                offer.offer_id not in inventory_snapshots
                or inventory_snapshots[offer.offer_id].platform_stock is None
            )
        )
        if counts["reported_inventory_missing"]:
            counts["reported_inventory_reason"] = (
                f"{reported_inventory_date.isoformat()} 10:05期末库存快照缺失；"
                f"{business_date.isoformat()}整日销量仍保留，"
                "不得用其他时点库存代替"
            )
        run.counts = dict(counts)
        previous = _previous_values(session, business_date)
        captured_by_offer = _all_observations(session, business_date)
        for offer in offers:
            inventory_snapshot = inventory_snapshots.get(offer.offer_id)
            platform_stock = (
                inventory_snapshot.platform_stock
                if inventory_snapshot is not None
                else None
            )
            stock_source = (
                "next_morning_1005"
                if inventory_snapshot is not None
                and inventory_snapshot.platform_stock is not None
                else (
                    "next_morning_1005_value_missing"
                    if inventory_snapshot is not None
                    else "next_morning_1005_snapshot_missing"
                )
            )
            observation = DailyReportObservation(
                run_id=run_id,
                offer_id=offer.offer_id,
                sku=offer.sku,
                title=offer.title,
                page_views_30_days=offer.page_views_30_days,
                ordered_units=int(sales.get(offer.offer_id, 0) or 0),
                platform_stock=platform_stock,
                stock_source=stock_source,
            )
            session.add(observation)
            capture_rows = captured_by_offer.setdefault(offer.offer_id, [])
            capture_rows.append((run, observation))
            resolution = session.scalar(
                select(DailyReportResolution).where(
                    DailyReportResolution.business_date == business_date,
                    DailyReportResolution.offer_id == offer.offer_id,
                )
            )
            if resolution is None:
                resolution = DailyReportResolution(
                    business_date=business_date,
                    offer_id=offer.offer_id,
                    status="awaiting_evening",
                    stock_alert_dismissed=False,
                    created_at=now,
                    updated_at=now,
                )
                session.add(resolution)
            previous_status = resolution.status
            resolution.status = _status_after_capture(
                session,
                resolution=resolution,
                slot=slot,
                incoming=_value_dict(observation),
                previous_stock=previous.get(offer.offer_id),
                capture_rows=capture_rows,
            )
            if _version_differences(capture_rows, resolution):
                _defer_following_stock_continuity(
                    session,
                    resolution=resolution,
                    deferred_at=now,
                )
            if previous_status == "confirmed" and resolution.status != "confirmed":
                reopened += 1
                session.add(
                    DailyReportAudit(
                        business_date=business_date,
                        offer_id=offer.offer_id,
                        action="system_reopened",
                        payload={"slot": slot, "run_id": run_id},
                        note="确认后平台数据发生变化，已重新进入待确认。",
                        user_id=None,
                        created_at=now,
                    )
                )
            resolution.updated_at = now
    return ReportCaptureResult(
        run_id=run_id,
        business_date=business_date,
        slot=slot,
        product_count=len(offers),
        reopened_count=reopened,
    )


def daily_report_payload(engine: Engine, business_date: date) -> dict[str, Any]:
    """Build one date's comparison table and all persistent prior reminders."""
    with Session(engine) as session:
        runs = list(
            session.scalars(
                select(DailyReportRun)
                .where(DailyReportRun.business_date == business_date)
                .order_by(DailyReportRun.captured_at)
            )
        )
        observations = _latest_observations(session, business_date)
        all_observations = _all_observations(session, business_date)
        successful_runs = [run for run in runs if run.status == "success"]
        capture_status = _capture_status(runs, business_date)
        confirmation_triggers = _confirmation_trigger_map(session, business_date)
        operator_notes = _operator_note_history_map(session, business_date)
        resolutions = {
            row.offer_id: row
            for row in session.scalars(
                select(DailyReportResolution).where(
                    DailyReportResolution.business_date == business_date
                )
            )
        }
        confirmation_baselines = _confirmation_baseline_map(
            session,
            list(resolutions.values()),
        )
        confirmation_reverts = _confirmation_revert_map(
            session,
            business_date,
        )
        previous_contexts = _previous_stock_contexts(session, business_date)
        offer_ids = sorted(all_observations)
        items = [
            _item_payload(
                offer_id,
                observations["morning"].get(offer_id),
                observations["evening"].get(offer_id),
                all_observations.get(offer_id, []),
                successful_runs,
                resolutions.get(offer_id),
                previous_contexts.get(offer_id),
                capture_status,
                confirmation_triggers.get(offer_id),
                operator_notes.get(offer_id, []),
                confirmation_baselines.get(offer_id),
                confirmation_reverts.get(offer_id),
            )
            for offer_id in offer_ids
        ]
        reminders = _reminders(session, before=business_date)
        deadline = session.get(DailyReportDeadlineSnapshot, business_date)
        comparison_history = _comparison_history(
            session,
            through=business_date,
            current_items=items,
        )
        pending_actions = _pending_actions(
            session,
            through=business_date,
            current_items=items,
        )
        handled_actions = _handled_actions(session, through=business_date)
    counts = {
        "products": len(items),
        "with_sales": sum(
            1
            for item in items
            if int(item["current"]["ordered_units"] or 0) > 0
        ),
        "awaiting_evening": sum(item["status"] == "awaiting_evening" for item in items),
        "ready": sum(item["status"] == "ready" for item in items),
        "needs_review": sum(item["status"] == "needs_review" for item in items),
        "missing_capture": sum(item["status"] == "missing_capture" for item in items),
        "confirmed": sum(item["status"] == "confirmed" for item in items),
        "stock_alerts": sum(item["stock_check"]["mismatch"] for item in items),
    }
    return {
        "business_date": business_date.isoformat(),
        "runs": [
            {
                "run_id": run.run_id,
                "slot": run.slot,
                "captured_at": run.captured_at.isoformat(),
                "status": run.status,
                "counts": run.counts or {},
            }
            for run in runs
        ],
        "capture_status": capture_status,
        "capture_issues": _capture_issues(items, capture_status, runs),
        "comparison_history": comparison_history,
        "pending_actions": pending_actions,
        "handled_actions": handled_actions,
        "counts": counts,
        "items": items,
        "prior_reminders": reminders,
        "deadline_snapshot": (
            {
                "snapped_at": deadline.snapped_at.isoformat(),
                "unresolved_count": deadline.unresolved_count,
                "resolved_at": (
                    deadline.resolved_at.isoformat()
                    if deadline.resolved_at is not None
                    else None
                ),
            }
            if deadline is not None
            else None
        ),
    }


def reminder_payload(
    engine: Engine,
    current_business_date: date | None = None,
) -> dict[str, Any]:
    """Return unresolved past dates for the global start-of-work reminder."""
    today = current_business_date or datetime.now(
        ZoneInfo("Africa/Johannesburg")
    ).date()
    with Session(engine) as session:
        rows = _reminders(session, before=today)
        active_deadlines = list(
            session.scalars(
                select(DailyReportDeadlineSnapshot.business_date).where(
                    DailyReportDeadlineSnapshot.resolved_at.is_(None),
                    DailyReportDeadlineSnapshot.business_date >= today,
                )
            )
        )
        for deadline_date in active_deadlines:
            rows.extend(_reminders_for_date(session, deadline_date))
    rows = list({str(row["business_date"]): row for row in rows}.values())
    rows.sort(key=lambda row: str(row["business_date"]))
    return {
        "count": sum(int(row["unresolved_count"]) for row in rows),
        "dates": rows,
    }


def save_manual_candidate(
    engine: Engine,
    *,
    business_date: date,
    offer_id: str,
    values: Mapping[str, int | None],
    reason: str,
    note: str,
    user_id: int,
) -> None:
    if reason not in MANUAL_REASONS:
        raise DailyReportInputError("人工修改原因无效")
    clean_note = _required_note(note, "人工修改必须填写备注")
    supplied = {key: values.get(key) for key in _VALUE_KEYS if key in values}
    if not supplied:
        raise DailyReportInputError("至少填写一个人工修改值")
    for key, value in supplied.items():
        if value is not None and int(value) < 0:
            raise DailyReportInputError(f"{key} 不能小于 0")
    now = _utc_now()
    with Session(engine) as session, session.begin():
        resolution = _resolution_or_error(session, business_date, offer_id)
        capture_rows = _all_observations(session, business_date).get(
            offer_id,
            [],
        )
        version_differences_before_edit = _version_differences(
            capture_rows,
            resolution,
        )
        previous_effective_stock = _effective_stock_before_confirmation(
            session,
            resolution,
        )
        before = _manual_values(resolution)
        if "page_views_30_days" in supplied:
            resolution.manual_page_views_30_days = supplied["page_views_30_days"]
        if "ordered_units" in supplied:
            resolution.manual_ordered_units = supplied["ordered_units"]
        if "platform_stock" in supplied:
            resolution.manual_platform_stock = supplied["platform_stock"]
        resolution.manual_reason = reason
        resolution.manual_note = clean_note
        resolution.manual_by = user_id
        resolution.manual_at = now
        resolution.stock_alert_dismissed = False
        resolution.stock_alert_note = None
        resolution.stock_alert_dismissed_by = None
        resolution.stock_alert_dismissed_at = None
        if resolution.status == "confirmed":
            resolution.status = "needs_review"
        elif _latest_observation_any_slot(session, business_date, offer_id) is None:
            resolution.status = "awaiting_evening"
        else:
            resolution.status = "needs_review"
        resolution.updated_at = now
        _audit(
            session,
            business_date,
            offer_id,
            "manual_candidate",
            {"before": before, "after": _manual_values(resolution), "reason": reason},
            clean_note,
            user_id,
            now,
        )
        manual_values = _source_values(session, resolution, "manual")
        previous_stock = _previous_values(session, business_date).get(offer_id)
        can_auto_confirm = bool(
            version_differences_before_edit
            and previous_stock is not None
            and manual_values is not None
            and manual_values.get("ordered_units") is not None
            and manual_values.get("platform_stock") is not None
            and not _stock_continuity_mismatch(previous_stock, manual_values)
        )
        if not can_auto_confirm or manual_values is None:
            return
        _apply_final(
            resolution,
            manual_values,
            "manual",
            clean_note,
            user_id,
            now,
        )
        _audit(
            session,
            business_date,
            offer_id,
            "confirm",
            {
                "source": "manual",
                "values": manual_values,
                "automatic": True,
                "reason": "manual_version_fix_matches_stock_continuity",
            },
            clean_note,
            user_id,
            now,
        )
        _propagate_confirmation_stock_conflict(
            session,
            resolution=resolution,
            previous_effective_stock=previous_effective_stock,
            confirmed_values=manual_values,
            source="manual",
            note=clean_note,
            user_id=user_id,
            confirmed_at=now,
            resolved_version_difference=True,
        )
        _resolve_deadline_if_complete(session, business_date, now)


def save_operator_note(
    engine: Engine,
    *,
    business_date: date,
    offer_id: str,
    note: str,
    user_id: int,
    issue_type: str = "general",
) -> None:
    if issue_type not in NOTE_ISSUE_TYPES:
        raise DailyReportInputError("备注关联问题无效")
    clean_note = _required_note(note, "备注不能为空")
    now = _utc_now()
    with Session(engine) as session, session.begin():
        resolution = _resolution_or_error(session, business_date, offer_id)
        resolution.operator_note = clean_note
        resolution.updated_at = now
        _audit(
            session,
            business_date,
            offer_id,
            "operator_note",
            {"note": clean_note, "issue_type": issue_type},
            clean_note,
            user_id,
            now,
        )


def update_operator_note(
    engine: Engine,
    *,
    business_date: date,
    offer_id: str,
    note_id: int,
    note: str,
    user_id: int,
    issue_type: str = "general",
) -> None:
    if issue_type not in NOTE_ISSUE_TYPES:
        raise DailyReportInputError("备注关联问题无效")
    clean_note = _required_note(note, "备注不能为空")
    now = _utc_now()
    with Session(engine) as session, session.begin():
        resolution = _resolution_or_error(session, business_date, offer_id)
        existing_notes = _operator_note_history_map(session, business_date).get(
            offer_id, []
        )
        existing = next((row for row in existing_notes if row["id"] == note_id), None)
        if existing is None:
            raise DailyReportConflictError("该备注不存在或已经删除")
        _audit(
            session,
            business_date,
            offer_id,
            "operator_note_updated",
            {
                "note_id": note_id,
                "before_note": existing["note"],
                "before_issue_type": existing["issue_type"],
                "note": clean_note,
                "issue_type": issue_type,
            },
            clean_note,
            user_id,
            now,
        )
        existing["note"] = clean_note
        resolution.operator_note = existing_notes[-1]["note"]
        resolution.updated_at = now


def delete_operator_note(
    engine: Engine,
    *,
    business_date: date,
    offer_id: str,
    note_id: int,
    note: str,
    user_id: int,
) -> None:
    now = _utc_now()
    with Session(engine) as session, session.begin():
        resolution = _resolution_or_error(session, business_date, offer_id)
        existing_notes = _operator_note_history_map(session, business_date).get(
            offer_id, []
        )
        existing = next(
            (row for row in existing_notes if row["id"] == note_id),
            None,
        )
        if existing is None:
            raise DailyReportConflictError("该备注不存在或已经删除")
        clean_note = str(note or "").strip()
        if existing["issue_type"] != "general":
            clean_note = _required_note(
                clean_note,
                "删除问题备注必须填写删除原因",
            )
        elif len(clean_note) > 2000:
            raise DailyReportInputError("备注不能超过 2000 个字符")
        _audit(
            session,
            business_date,
            offer_id,
            "operator_note_deleted",
            {
                "note_id": note_id,
                "deleted_note": existing["note"],
                "issue_type": existing["issue_type"],
            },
            clean_note or None,
            user_id,
            now,
        )
        remaining = [row for row in existing_notes if row["id"] != note_id]
        resolution.operator_note = remaining[-1]["note"] if remaining else None
        resolution.updated_at = now


def confirm_entry(
    engine: Engine,
    *,
    business_date: date,
    offer_id: str,
    source: str,
    note: str,
    user_id: int,
) -> None:
    if source not in SOURCES:
        raise DailyReportInputError("最终值来源只能选择早间、晚间、最新拉取或人工值")
    clean_note = _required_note(note, "确认合并必须填写备注")
    now = _utc_now()
    with Session(engine) as session, session.begin():
        resolution = _resolution_or_error(session, business_date, offer_id)
        capture_rows = _all_observations(session, business_date).get(
            offer_id,
            [],
        )
        resolved_version_differences = _version_differences(
            capture_rows,
            resolution,
        )
        reconfirming_reverted_entry = bool(
            _confirmation_revert_map(session, business_date).get(offer_id)
        ) and not _has_confirmation_baseline(resolution)
        previous_effective_stock = _effective_stock_before_confirmation(
            session,
            resolution,
        )
        values = _source_values(session, resolution, source)
        if values is None:
            raise DailyReportConflictError("所选来源没有可用数据")
        resolution.stock_alert_dismissed = False
        resolution.stock_alert_note = None
        resolution.stock_alert_dismissed_by = None
        resolution.stock_alert_dismissed_at = None
        _apply_final(resolution, values, source, clean_note, user_id, now)
        previous_stock = _previous_values(session, business_date).get(offer_id)
        if _stock_continuity_mismatch(previous_stock, values):
            resolution.status = "needs_review"
            _audit(
                session,
                business_date,
                offer_id,
                "stock_continuity_after_confirmation",
                {
                    "source": source,
                    "values": values,
                    "previous_stock": previous_stock,
                    "ordered_units": values.get("ordered_units"),
                    "actual_stock": values.get("platform_stock"),
                },
                "人工确认正确版本后仍存在前后日报日库存连续性问题。",
                user_id,
                now,
            )
        _audit(
            session,
            business_date,
            offer_id,
            "confirm",
            {"source": source, "values": values},
            clean_note,
            user_id,
            now,
        )
        _propagate_confirmation_stock_conflict(
            session,
            resolution=resolution,
            previous_effective_stock=previous_effective_stock,
            confirmed_values=values,
            source=source,
            note=clean_note,
            user_id=user_id,
            confirmed_at=now,
            resolved_version_difference=(
                bool(resolved_version_differences)
                or reconfirming_reverted_entry
            ),
        )
        _resolve_deadline_if_complete(session, business_date, now)


def revert_confirmation(
    engine: Engine,
    *,
    business_date: date,
    offer_id: str,
    note: str,
    user_id: int,
) -> None:
    """Reopen a confirmed entry without deleting the original audit trail."""
    clean_note = _required_note(note, "撤销确认必须填写原因")
    now = _utc_now()
    with Session(engine) as session, session.begin():
        resolution = _resolution_or_error(session, business_date, offer_id)
        if not _has_confirmation_baseline(resolution):
            raise DailyReportConflictError("该商品没有可撤销的人工确认")

        previous_confirmation = {
            "values": _final_values(resolution),
            "source": resolution.selected_source,
            "source_label": _confirmation_source_label(
                str(resolution.selected_source)
            ),
            "confirmed_by": _user_display_name(session, resolution.confirmed_by),
            "confirmed_at": (
                resolution.confirmed_at.isoformat()
                if resolution.confirmed_at is not None
                else None
            ),
            "confirm_note": resolution.confirm_note,
        }
        previous_stock_alert = {
            "dismissed": resolution.stock_alert_dismissed,
            "note": resolution.stock_alert_note,
            "dismissed_by": _user_display_name(
                session,
                resolution.stock_alert_dismissed_by,
            ),
            "dismissed_at": (
                resolution.stock_alert_dismissed_at.isoformat()
                if resolution.stock_alert_dismissed_at is not None
                else None
            ),
        }
        reverted_by = _user_display_name(session, user_id)

        resolution.selected_source = None
        resolution.final_page_views_30_days = None
        resolution.final_ordered_units = None
        resolution.final_platform_stock = None
        resolution.confirm_note = None
        resolution.confirmed_by = None
        resolution.confirmed_at = None
        resolution.stock_alert_dismissed = False
        resolution.stock_alert_note = None
        resolution.stock_alert_dismissed_by = None
        resolution.stock_alert_dismissed_at = None
        resolution.status = "needs_review"
        resolution.updated_at = now
        _audit(
            session,
            business_date,
            offer_id,
            "confirmation_reverted",
            {
                "previous_confirmation": previous_confirmation,
                "previous_stock_alert": previous_stock_alert,
                "reverted_by": reverted_by,
                "reverted_at": now.isoformat(),
                "revert_note": clean_note,
            },
            clean_note,
            user_id,
            now,
        )
        _queue_following_revert_impact(
            session,
            resolution=resolution,
            previous_confirmation=previous_confirmation,
            reverted_by=reverted_by,
            revert_note=clean_note,
            reverted_at=now,
            user_id=user_id,
        )
        _refresh_deadline_snapshot(session, business_date, now)


def confirm_ready_entries(
    engine: Engine,
    *,
    business_date: date,
    note: str,
    user_id: int,
) -> int:
    clean_note = _required_note(note, "批量确认必须填写备注")
    now = _utc_now()
    confirmed = 0
    with Session(engine) as session, session.begin():
        rows = list(
            session.scalars(
                select(DailyReportResolution).where(
                    DailyReportResolution.business_date == business_date,
                    DailyReportResolution.status == "ready",
                )
            )
        )
        for resolution in rows:
            previous_effective_stock = _effective_stock_before_confirmation(
                session,
                resolution,
            )
            values = _source_values(session, resolution, "latest")
            if values is None:
                continue
            _apply_final(resolution, values, "latest", clean_note, user_id, now)
            _audit(
                session,
                business_date,
                resolution.offer_id,
                "bulk_confirm",
                {"source": "latest", "values": values},
                clean_note,
                user_id,
                now,
            )
            _propagate_confirmation_stock_conflict(
                session,
                resolution=resolution,
                previous_effective_stock=previous_effective_stock,
                confirmed_values=values,
                source="latest",
                note=clean_note,
                user_id=user_id,
                confirmed_at=now,
            )
            confirmed += 1
        _resolve_deadline_if_complete(session, business_date, now)
    return confirmed


def dismiss_stock_alert(
    engine: Engine,
    *,
    business_date: date,
    offer_id: str,
    note: str,
    user_id: int,
) -> None:
    clean_note = _required_note(note, "确认库存差异必须填写原因")
    now = _utc_now()
    with Session(engine) as session, session.begin():
        resolution = _resolution_or_error(session, business_date, offer_id)
        if resolution.stock_alert_dismissed:
            raise DailyReportConflictError("该库存差异已经确认，无需重复处理")
        capture_rows = _all_observations(session, business_date).get(
            offer_id,
            [],
        )
        differences = _version_differences(capture_rows, resolution)
        if differences:
            raise DailyReportConflictError(
                "当前仍有同周期版本差异，请先确认合并正确版本"
            )
        pending_manual = _manual_candidate_is_pending(resolution)
        if pending_manual:
            values = _source_values(session, resolution, "manual")
        else:
            values = (
                _final_values(resolution)
                if _has_confirmation_baseline(resolution)
                else _coalesced_capture_values(
                    [_value_dict(observation) for _, observation in capture_rows]
                )
            )
        if values is None:
            raise DailyReportConflictError("当前没有可用于核对的库存数据")
        previous_stock = _previous_values(session, business_date).get(offer_id)
        if not _stock_continuity_mismatch(previous_stock, values):
            if pending_manual:
                raise DailyReportConflictError(
                    "修改后的库存公式已经相符，只能消除差异"
                )
            raise DailyReportConflictError("当前库存连续性已经相符，无需确认差异")
        previous_status = resolution.status
        previous_effective_stock = _effective_stock_before_confirmation(
            session,
            resolution,
        )
        if pending_manual:
            _apply_final(
                resolution,
                values,
                "manual",
                clean_note,
                user_id,
                now,
            )
        resolution.stock_alert_dismissed = True
        resolution.stock_alert_note = clean_note
        resolution.stock_alert_dismissed_by = user_id
        resolution.stock_alert_dismissed_at = now
        latest = _latest_observation_any_slot(session, business_date, offer_id)
        if latest is not None:
            resolution.status = _status_after_capture(
                session,
                resolution=resolution,
                slot="manual",
                incoming=_value_dict(latest),
                previous_stock=_previous_values(session, business_date).get(offer_id),
            )
        resolution.updated_at = now
        _audit(
            session,
            business_date,
            offer_id,
            "dismiss_stock_alert",
            {
                "previous_status": previous_status,
                "source": "manual" if pending_manual else None,
                "previous_stock": previous_stock,
                "ordered_units": values.get("ordered_units"),
                "expected_stock": (
                    previous_stock - int(values.get("ordered_units") or 0)
                    if previous_stock is not None
                    else None
                ),
                "actual_stock": values.get("platform_stock"),
                "values": values,
            },
            clean_note,
            user_id,
            now,
        )
        if pending_manual:
            _propagate_confirmation_stock_conflict(
                session,
                resolution=resolution,
                previous_effective_stock=previous_effective_stock,
                confirmed_values=values,
                source="manual",
                note=clean_note,
                user_id=user_id,
                confirmed_at=now,
            )
        _refresh_deadline_snapshot(session, business_date, now)


def eliminate_stock_alert(
    engine: Engine,
    *,
    business_date: date,
    offer_id: str,
    note: str,
    user_id: int,
) -> None:
    """Apply a matching manual candidate and close the continuity todo."""
    clean_note = _required_note(note, "消除库存差异必须填写备注")
    now = _utc_now()
    with Session(engine) as session, session.begin():
        resolution = _resolution_or_error(session, business_date, offer_id)
        capture_rows = _all_observations(session, business_date).get(
            offer_id,
            [],
        )
        if _version_differences(capture_rows, resolution):
            raise DailyReportConflictError(
                "当前仍有同周期版本差异，请先确认合并正确版本"
            )
        if not _manual_candidate_is_pending(resolution):
            raise DailyReportConflictError(
                "请先人工修改库存，使库存连续性公式相符"
            )
        values = _source_values(session, resolution, "manual")
        if values is None:
            raise DailyReportConflictError("当前没有可用于核对的人工修改值")
        previous_stock = _previous_values(session, business_date).get(offer_id)
        if previous_stock is None or values.get("platform_stock") is None:
            raise DailyReportConflictError("库存连续性数据不完整，暂时无法消除差异")
        if _stock_continuity_mismatch(previous_stock, values):
            raise DailyReportConflictError(
                "修改后的库存公式仍不相符，只能确认库存差异"
            )
        previous_effective_stock = _effective_stock_before_confirmation(
            session,
            resolution,
        )
        resolution.stock_alert_dismissed = False
        resolution.stock_alert_note = None
        resolution.stock_alert_dismissed_by = None
        resolution.stock_alert_dismissed_at = None
        _apply_final(
            resolution,
            values,
            "manual",
            clean_note,
            user_id,
            now,
        )
        expected_stock = previous_stock - int(values.get("ordered_units") or 0)
        _audit(
            session,
            business_date,
            offer_id,
            "eliminate_stock_alert",
            {
                "source": "manual",
                "previous_stock": previous_stock,
                "ordered_units": values.get("ordered_units"),
                "expected_stock": expected_stock,
                "actual_stock": values.get("platform_stock"),
                "values": values,
            },
            clean_note,
            user_id,
            now,
        )
        _propagate_confirmation_stock_conflict(
            session,
            resolution=resolution,
            previous_effective_stock=previous_effective_stock,
            confirmed_values=values,
            source="manual",
            note=clean_note,
            user_id=user_id,
            confirmed_at=now,
        )
        _resolve_deadline_if_complete(session, business_date, now)


def reopen_stock_alert(
    engine: Engine,
    *,
    business_date: date,
    offer_id: str,
    note: str,
    user_id: int,
) -> None:
    """Undo one stock-difference acknowledgement without deleting its audit."""
    clean_note = _required_note(note, "撤销库存差异确认必须填写原因")
    now = _utc_now()
    with Session(engine) as session, session.begin():
        resolution = _resolution_or_error(session, business_date, offer_id)
        if not resolution.stock_alert_dismissed:
            raise DailyReportConflictError("该商品没有可撤销的库存差异确认")
        previous_handling = {
            "note": resolution.stock_alert_note,
            "handled_by": _user_display_name(
                session,
                resolution.stock_alert_dismissed_by,
            ),
            "handled_at": (
                resolution.stock_alert_dismissed_at.isoformat()
                if resolution.stock_alert_dismissed_at is not None
                else None
            ),
        }
        resolution.stock_alert_dismissed = False
        resolution.stock_alert_note = None
        resolution.stock_alert_dismissed_by = None
        resolution.stock_alert_dismissed_at = None
        capture_rows = _all_observations(session, business_date).get(
            offer_id,
            [],
        )
        latest = _latest_observation_any_slot(session, business_date, offer_id)
        if latest is not None:
            resolution.status = _status_after_capture(
                session,
                resolution=resolution,
                slot="manual",
                incoming=_value_dict(latest),
                previous_stock=_previous_values(session, business_date).get(offer_id),
                capture_rows=capture_rows,
            )
        resolution.updated_at = now
        _audit(
            session,
            business_date,
            offer_id,
            "stock_alert_reopened",
            {"previous_handling": previous_handling},
            clean_note,
            user_id,
            now,
        )
        _refresh_deadline_snapshot(session, business_date, now)


def backfill_stock_continuity_reviews(
    engine: Engine,
    *,
    through: date | None = None,
) -> int:
    """Reconcile persisted review states with version-first continuity rules."""
    now = _utc_now()
    updated = 0
    with Session(engine) as session, session.begin():
        statement = select(DailyReportResolution)
        if through is not None:
            statement = statement.where(
                DailyReportResolution.business_date <= through
            )
        rows = list(
            session.scalars(
                statement.order_by(
                    DailyReportResolution.business_date,
                    DailyReportResolution.offer_id,
                )
            )
        )
        observations_by_date: dict[
            date,
            dict[str, list[tuple[DailyReportRun, DailyReportObservation]]],
        ] = {}
        previous_by_date: dict[date, dict[str, int | None]] = {}
        report_dates: set[date] = set()
        for resolution in rows:
            report_date = resolution.business_date
            report_dates.add(report_date)
            if report_date not in observations_by_date:
                observations_by_date[report_date] = _all_observations(
                    session,
                    report_date,
                )
            observations = observations_by_date[report_date]
            latest_rows = observations.get(resolution.offer_id, [])
            if not latest_rows:
                continue
            if report_date not in previous_by_date:
                previous_by_date[report_date] = _previous_values(
                    session,
                    report_date,
                )
            previous = previous_by_date[report_date]
            latest = latest_rows[-1][1]
            new_status = _status_after_capture(
                session,
                resolution=resolution,
                slot="manual",
                incoming=_value_dict(latest),
                previous_stock=previous.get(resolution.offer_id),
                capture_rows=latest_rows,
            )
            if new_status == resolution.status:
                continue
            old_status = resolution.status
            resolution.status = new_status
            resolution.updated_at = now
            _audit(
                session,
                report_date,
                resolution.offer_id,
                "review_state_reconciled",
                {
                    "before": old_status,
                    "after": new_status,
                    "rule": "sales_version_difference_before_stock_continuity",
                },
                "按“先确认当日销量版本、再核对前后日报日库存”规则重算待办状态。",
                None,
                now,
            )
            updated += 1
        for report_date in report_dates:
            _refresh_deadline_snapshot(session, report_date, now)
    return updated


def backfill_daily_inventory_snapshots(
    engine: Engine,
    *,
    through: date | None = None,
) -> InventorySnapshotBackfillResult:
    """Build actual-date snapshots and attach next-morning stock to report rows."""
    snapshots_created = 0
    observations_updated = 0
    observations_missing_snapshot = 0
    with Session(engine) as session, session.begin():
        existing_keys = {
            (inventory_date, offer_id)
            for inventory_date, offer_id in session.execute(
                select(
                    DailyInventorySnapshot.inventory_date,
                    DailyInventorySnapshot.offer_id,
                )
            )
        }
        morning_rows = session.execute(
            select(DailyReportRun, DailyReportObservation)
            .join(
                DailyReportObservation,
                DailyReportObservation.run_id == DailyReportRun.run_id,
            )
            .where(
                DailyReportRun.status == "success",
                DailyReportRun.slot == "morning",
            )
            .order_by(
                DailyReportRun.captured_at,
                DailyReportRun.run_id,
                DailyReportObservation.offer_id,
            )
        ).all()
        for run, observation in morning_rows:
            inventory_date = _beijing_date(run.captured_at)
            key = (inventory_date, observation.offer_id)
            if key in existing_keys:
                continue
            session.add(
                DailyInventorySnapshot(
                    inventory_date=inventory_date,
                    offer_id=observation.offer_id,
                    run_id=run.run_id,
                    captured_at=run.captured_at,
                    platform_stock=observation.platform_stock,
                    stock_source=observation.stock_source,
                )
            )
            existing_keys.add(key)
            snapshots_created += 1
        session.flush()

        snapshots = {
            (row.inventory_date, row.offer_id): row
            for row in session.scalars(select(DailyInventorySnapshot))
        }
        statement = (
            select(DailyReportRun, DailyReportObservation)
            .join(
                DailyReportObservation,
                DailyReportObservation.run_id == DailyReportRun.run_id,
            )
            .where(DailyReportRun.status == "success")
            .order_by(
                DailyReportRun.business_date,
                DailyReportRun.captured_at,
                DailyReportObservation.offer_id,
            )
        )
        if through is not None:
            statement = statement.where(DailyReportRun.business_date <= through)
        run_stats: dict[str, dict[str, Any]] = {}
        for run, observation in session.execute(statement):
            reported_inventory_date = run.business_date + timedelta(days=1)
            snapshot = snapshots.get(
                (reported_inventory_date, observation.offer_id)
            )
            stats = run_stats.setdefault(
                run.run_id,
                {
                    "run": run,
                    "reported_inventory_date": reported_inventory_date,
                    "snapshot_count": 0,
                    "missing_count": 0,
                },
            )
            if snapshot is not None:
                stats["snapshot_count"] += 1
            if snapshot is None or snapshot.platform_stock is None:
                stats["missing_count"] += 1
            expected_stock = (
                snapshot.platform_stock if snapshot is not None else None
            )
            expected_source = (
                "next_morning_1005"
                if snapshot is not None and snapshot.platform_stock is not None
                else (
                    "next_morning_1005_value_missing"
                    if snapshot is not None
                    else "next_morning_1005_snapshot_missing"
                )
            )
            if snapshot is None or snapshot.platform_stock is None:
                observations_missing_snapshot += 1
            if (
                observation.platform_stock == expected_stock
                and observation.stock_source == expected_source
            ):
                continue
            observation.platform_stock = expected_stock
            observation.stock_source = expected_source
            observations_updated += 1
        for stats in run_stats.values():
            run = stats["run"]
            reported_inventory_date = stats["reported_inventory_date"]
            counts = dict(run.counts or {})
            counts["reported_inventory_date"] = reported_inventory_date.isoformat()
            counts["reported_inventory_snapshots"] = stats["snapshot_count"]
            counts["reported_inventory_missing"] = stats["missing_count"]
            counts["reported_inventory_reason"] = (
                (
                    f"{reported_inventory_date.isoformat()} "
                    "10:05期末库存快照缺失；"
                    f"{run.business_date.isoformat()}整日销量仍保留，"
                    "不得用其他时点库存代替"
                )
                if stats["missing_count"]
                else None
            )
            run.counts = counts
    return InventorySnapshotBackfillResult(
        snapshots_created=snapshots_created,
        observations_updated=observations_updated,
        observations_missing_snapshot=observations_missing_snapshot,
    )


def create_deadline_snapshot(
    engine: Engine,
    *,
    business_date: date,
    snapped_at: datetime,
) -> int:
    now = _naive_utc(snapped_at)
    with Session(engine) as session, session.begin():
        awaiting = list(
            session.scalars(
                select(DailyReportResolution).where(
                    DailyReportResolution.business_date == business_date,
                    DailyReportResolution.status == "awaiting_evening",
                )
            )
        )
        for row in awaiting:
            row.status = "missing_capture"
            row.updated_at = now
        unresolved = list(
            session.scalars(
                select(DailyReportResolution).where(
                    DailyReportResolution.business_date == business_date,
                    DailyReportResolution.status.in_(OPEN_STATUSES),
                )
            )
        )
        details = [
            {"offer_id": row.offer_id, "status": row.status}
            for row in unresolved
        ]
        snapshot = session.get(DailyReportDeadlineSnapshot, business_date)
        if snapshot is None:
            snapshot = DailyReportDeadlineSnapshot(
                business_date=business_date,
                snapped_at=now,
                unresolved_count=len(unresolved),
                details=details,
                resolved_at=now if not unresolved else None,
            )
            session.add(snapshot)
        else:
            snapshot.snapped_at = now
            snapshot.unresolved_count = len(unresolved)
            snapshot.details = details
            snapshot.resolved_at = now if not unresolved else None
        _audit(
            session,
            business_date,
            None,
            "deadline_snapshot",
            {"unresolved_count": len(unresolved), "details": details},
            None,
            None,
            now,
        )
    return len(unresolved)


def export_operations_workbook(
    engine: Engine,
    *,
    business_date: date,
    destination: Path,
) -> Path:
    """Export confirmed history in the reference workbook's product-column layout."""
    with Session(engine) as session:
        unresolved = list(
            session.scalars(
                select(DailyReportResolution).where(
                    DailyReportResolution.business_date <= business_date,
                    DailyReportResolution.status.in_(OPEN_STATUSES),
                )
            )
        )
        if unresolved:
            locations = ", ".join(
                f"{row.business_date.isoformat()} / {row.offer_id}"
                for row in unresolved[:8]
            )
            suffix = " 等" if len(unresolved) > 8 else ""
            raise DailyReportConflictError(
                f"仍有 {len(unresolved)} 个数据未合并：{locations}{suffix}"
            )
        dates = list(
            session.scalars(
                select(DailyReportResolution.business_date)
                .where(
                    DailyReportResolution.business_date <= business_date,
                    DailyReportResolution.status.in_(EXPORTABLE_STATUSES),
                )
                .distinct()
                .order_by(DailyReportResolution.business_date)
            )
        )
        if not dates:
            raise DailyReportConflictError("尚无可导出的运营日报数据")
        resolutions = list(
            session.scalars(
                select(DailyReportResolution)
                .where(
                    DailyReportResolution.business_date.in_(dates),
                    DailyReportResolution.status.in_(EXPORTABLE_STATUSES),
                )
                .order_by(
                    DailyReportResolution.business_date,
                    DailyReportResolution.offer_id,
                )
            )
        )
        identities = _identity_map(session, resolutions)
        values_by_key = {
            (row.business_date, row.offer_id): _export_values(session, row)
            for row in resolutions
        }
        notes_by_key = {
            (report_date, offer_id): notes
            for report_date in dates
            for offer_id, notes in _operator_note_history_map(
                session,
                report_date,
            ).items()
        }
        previous = _all_previous_stock(resolutions, values_by_key)
        missing_rows = _export_missing_rows(session, resolutions, identities)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "运营日报"
    offer_ids = sorted({row.offer_id for row in resolutions})
    by_key = {(row.business_date, row.offer_id): row for row in resolutions}
    sheet.cell(1, 1, "指标")
    sheet.cell(1, 2, "日期")
    for column, offer_id in enumerate(offer_ids, start=3):
        identity = identities.get(offer_id, {})
        title = str(identity.get("title") or offer_id)
        sku = str(identity.get("sku") or offer_id)
        sheet.cell(1, column, f"{title}\n{sku}")
    header_fill = PatternFill("solid", fgColor="FFFF00")
    date_fill = PatternFill("solid", fgColor="D9E5F5")
    sales_fill = PatternFill("solid", fgColor="FCE4D6")
    alert_fill = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    row_number = 2
    for report_date in dates:
        labels = (
            "近30天浏览量",
            "当天订单数",
            "平台库存数量（次日10:05期末）",
            "备注",
        )
        for offset, label in enumerate(labels):
            sheet.cell(row_number + offset, 1, label)
            if offset == 0:
                sheet.cell(row_number + offset, 2, report_date.isoformat())
                for column in range(1, len(offer_ids) + 3):
                    sheet.cell(row_number + offset, column).fill = date_fill
        for column, offer_id in enumerate(offer_ids, start=3):
            resolution = by_key.get((report_date, offer_id))
            if resolution is None:
                continue
            values = values_by_key[(report_date, offer_id)]
            sheet.cell(row_number, column, values["page_views_30_days"])
            orders = values["ordered_units"]
            stock = values["platform_stock"]
            sheet.cell(row_number + 1, column, orders)
            sheet.cell(row_number + 2, column, stock)
            sheet.cell(
                row_number + 3,
                column,
                _operator_note_cell_text(
                    notes_by_key.get((report_date, offer_id), []),
                    confirmation_note=(
                        resolution.confirm_note
                        if _has_confirmation_baseline(resolution)
                        else None
                    ),
                    stock_confirmation_note=(
                        resolution.stock_alert_note
                        if resolution.stock_alert_dismissed
                        else None
                    ),
                ),
            )
            if int(orders or 0) > 0:
                sheet.cell(row_number + 1, column).fill = sales_fill
            previous_stock = previous.get((report_date, offer_id))
            if (
                previous_stock is not None
                and stock is not None
                and previous_stock - int(orders or 0) != stock
            ):
                sheet.cell(row_number + 2, column).fill = alert_fill
        sheet.row_dimensions[row_number + 3].height = 36
        row_number += 4
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(name="宋体", size=11, bold=False)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = Font(name="宋体", size=11)
            cell.border = border
            if cell.value is not None:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
    sheet.freeze_panes = "C2"
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 15
    for column in range(3, len(offer_ids) + 3):
        sheet.column_dimensions[get_column_letter(column)].width = 22
    sheet.row_dimensions[1].height = 54
    note_sheet = workbook.create_sheet("漏爬说明")
    note_sheet.append(("日期", "时段", "商品", "SKU / Offer ID", "漏爬原因", "自动采用"))
    for row in missing_rows:
        note_sheet.append(row)
    for cell in note_sheet[1]:
        cell.fill = header_fill
        cell.font = Font(name="宋体", size=11)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in note_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=11)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    note_sheet.freeze_panes = "A2"
    for column, width in enumerate((13, 10, 36, 20, 62, 16), start=1):
        note_sheet.column_dimensions[get_column_letter(column)].width = width
    workbook.save(destination)
    workbook.close()
    return destination


def unresolved_locations(engine: Engine, through: date) -> list[dict[str, Any]]:
    with Session(engine) as session:
        rows = list(
            session.scalars(
                select(DailyReportResolution)
                .where(
                    DailyReportResolution.business_date <= through,
                    DailyReportResolution.status.in_(OPEN_STATUSES),
                )
                .order_by(
                    DailyReportResolution.business_date,
                    DailyReportResolution.offer_id,
                )
            )
        )
        identities = _identity_map(session, rows)
    return [
        {
            "business_date": row.business_date.isoformat(),
            "offer_id": row.offer_id,
            "sku": identities.get(row.offer_id, {}).get("sku"),
            "title": identities.get(row.offer_id, {}).get("title"),
            "status": row.status,
        }
        for row in rows
    ]


_VALUE_KEYS = ("page_views_30_days", "ordered_units", "platform_stock")
_MISSING_CAPTURE_KEYS = ("ordered_units", "platform_stock")
_RECONCILIATION_KEYS = ("ordered_units",)


def _platform_stock(offer: OfferCurrent) -> tuple[int | None, str | None]:
    if offer.takealot_available_stock is not None:
        return int(offer.takealot_available_stock), "takealot_available_stock"
    if offer.total_stock is not None:
        return int(offer.total_stock), "total_stock_fallback"
    return None, None


def _freeze_morning_inventory_snapshots(
    session: Session,
    *,
    run: DailyReportRun,
    offers: list[OfferCurrent],
) -> tuple[date, int]:
    """Freeze the first successful 10:05 stock by its actual Beijing date."""
    inventory_date = _beijing_date(run.captured_at)
    existing_offer_ids = set(
        session.scalars(
            select(DailyInventorySnapshot.offer_id).where(
                DailyInventorySnapshot.inventory_date == inventory_date
            )
        )
    )
    created = 0
    for offer in offers:
        if offer.offer_id in existing_offer_ids:
            continue
        platform_stock, stock_source = _platform_stock(offer)
        session.add(
            DailyInventorySnapshot(
                inventory_date=inventory_date,
                offer_id=offer.offer_id,
                run_id=run.run_id,
                captured_at=run.captured_at,
                platform_stock=platform_stock,
                stock_source=stock_source,
            )
        )
        created += 1
    session.flush()
    return inventory_date, created


def _daily_inventory_snapshot_map(
    session: Session,
    inventory_date: date,
) -> dict[str, DailyInventorySnapshot]:
    return {
        row.offer_id: row
        for row in session.scalars(
            select(DailyInventorySnapshot)
            .where(DailyInventorySnapshot.inventory_date == inventory_date)
            .order_by(DailyInventorySnapshot.offer_id)
        )
    }


def _status_after_capture(
    session: Session,
    *,
    resolution: DailyReportResolution,
    slot: str,
    incoming: dict[str, int | None],
    previous_stock: int | None,
    capture_rows: (
        list[tuple[DailyReportRun, DailyReportObservation]] | None
    ) = None,
) -> str:
    del slot
    del incoming
    captured = (
        capture_rows
        if capture_rows is not None
        else _all_observations(session, resolution.business_date).get(
            resolution.offer_id,
            [],
        )
    )
    if _version_differences(captured, resolution):
        return "needs_review"
    if _has_confirmation_baseline(resolution):
        selected = _final_values(resolution)
        if any(selected.get(key) is None for key in _MISSING_CAPTURE_KEYS):
            return "missing_capture"
        if (
            not resolution.stock_alert_dismissed
            and _stock_continuity_mismatch(previous_stock, selected)
        ):
            return "needs_review"
        return "confirmed"
    candidates = [_value_dict(observation) for _, observation in captured]
    selected = _coalesced_capture_values(candidates[: len(captured)])
    if any(selected.get(key) is None for key in _MISSING_CAPTURE_KEYS):
        return "missing_capture"
    if len(captured) < 2:
        return "awaiting_evening"
    if (
        not resolution.stock_alert_dismissed
        and _stock_continuity_mismatch(previous_stock, selected)
    ):
        return "needs_review"
    return "ready"


def _stock_continuity_mismatch(
    previous_stock: int | None,
    values: Mapping[str, int | None],
) -> bool:
    current_stock = values.get("platform_stock")
    if previous_stock is None or current_stock is None:
        return False
    orders = int(values.get("ordered_units") or 0)
    return previous_stock - orders != current_stock


def _latest_observations(
    session: Session, business_date: date
) -> dict[str, dict[str, DailyReportObservation]]:
    result: dict[str, dict[str, DailyReportObservation]] = {
        "morning": {},
        "evening": {},
    }
    for slot in SCHEDULED_SLOTS:
        run = session.scalar(
            select(DailyReportRun)
            .where(
                DailyReportRun.business_date == business_date,
                DailyReportRun.slot == slot,
                DailyReportRun.status == "success",
            )
            .order_by(DailyReportRun.captured_at.desc())
        )
        if run is None:
            continue
        result[slot] = {
            row.offer_id: row
            for row in session.scalars(
                select(DailyReportObservation).where(
                    DailyReportObservation.run_id == run.run_id
                )
            )
        }
    return result


def _all_observations(
    session: Session,
    business_date: date,
) -> dict[str, list[tuple[DailyReportRun, DailyReportObservation]]]:
    """Return every successful immutable capture, grouped by product."""
    result: dict[str, list[tuple[DailyReportRun, DailyReportObservation]]] = {}
    rows = session.execute(
        select(DailyReportRun, DailyReportObservation)
        .join(
            DailyReportObservation,
            DailyReportObservation.run_id == DailyReportRun.run_id,
        )
        .where(
            DailyReportRun.business_date == business_date,
            DailyReportRun.status == "success",
        )
        .order_by(DailyReportRun.captured_at, DailyReportRun.run_id)
    ).all()
    for run, observation in rows:
        result.setdefault(observation.offer_id, []).append((run, observation))
    return result


def _latest_observation(
    session: Session,
    business_date: date,
    slot: str,
    offer_id: str,
) -> DailyReportObservation | None:
    return session.scalar(
        select(DailyReportObservation)
        .join(DailyReportRun, DailyReportRun.run_id == DailyReportObservation.run_id)
        .where(
            DailyReportRun.business_date == business_date,
            DailyReportRun.slot == slot,
            DailyReportRun.status == "success",
            DailyReportObservation.offer_id == offer_id,
        )
        .order_by(DailyReportRun.captured_at.desc())
    )


def _latest_observation_any_slot(
    session: Session,
    business_date: date,
    offer_id: str,
) -> DailyReportObservation | None:
    return session.scalar(
        select(DailyReportObservation)
        .join(DailyReportRun, DailyReportRun.run_id == DailyReportObservation.run_id)
        .where(
            DailyReportRun.business_date == business_date,
            DailyReportRun.status == "success",
            DailyReportObservation.offer_id == offer_id,
        )
        .order_by(DailyReportRun.captured_at.desc(), DailyReportRun.run_id.desc())
    )


def _value_dict(
    row: DailyReportObservation,
) -> dict[str, int | None]:
    return {
        "page_views_30_days": row.page_views_30_days,
        "ordered_units": row.ordered_units,
        "platform_stock": row.platform_stock,
    }


def _coalesced_capture_values(
    candidates: list[dict[str, int | None]],
) -> dict[str, int | None]:
    """Use the newest non-null value per field while preserving missing-data logs."""
    return {
        key: next(
            (
                candidate[key]
                for candidate in reversed(candidates)
                if candidate.get(key) is not None
            ),
            None,
        )
        for key in _VALUE_KEYS
    }


def _capture_run_label(run: DailyReportRun) -> str:
    labels = {
        "morning": "10:05定时",
        "evening": "18:00定时",
        "manual": "手动刷新",
    }
    captured_at = run.captured_at.replace(tzinfo=UTC).astimezone(
        ZoneInfo("Asia/Shanghai")
    )
    return f"{labels.get(run.slot, run.slot)}（{captured_at:%m-%d %H:%M}）"


def _manual_values(
    row: DailyReportResolution,
) -> dict[str, int | None]:
    return {
        "page_views_30_days": row.manual_page_views_30_days,
        "ordered_units": row.manual_ordered_units,
        "platform_stock": row.manual_platform_stock,
    }


def _final_values(
    row: DailyReportResolution,
) -> dict[str, int | None]:
    return {
        "page_views_30_days": row.final_page_views_30_days,
        "ordered_units": row.final_ordered_units,
        "platform_stock": row.final_platform_stock,
    }


def _has_confirmation_baseline(row: DailyReportResolution | None) -> bool:
    return bool(
        row is not None
        and row.selected_source is not None
        and row.confirmed_at is not None
    )


def _capture_rows_after_confirmation(
    capture_rows: list[tuple[DailyReportRun, DailyReportObservation]],
    resolution: DailyReportResolution | None,
) -> list[tuple[DailyReportRun, DailyReportObservation]]:
    if not _has_confirmation_baseline(resolution):
        return capture_rows
    assert resolution is not None
    assert resolution.confirmed_at is not None
    return [
        pair
        for pair in capture_rows
        if pair[0].captured_at > resolution.confirmed_at
    ]


def _manual_candidate_is_pending(
    resolution: DailyReportResolution | None,
) -> bool:
    if resolution is None or resolution.manual_at is None:
        return False
    if not any(value is not None for value in _manual_values(resolution).values()):
        return False
    if not _has_confirmation_baseline(resolution):
        return True
    assert resolution.confirmed_at is not None
    return resolution.manual_at > resolution.confirmed_at


def _version_candidate_values(
    capture_rows: list[tuple[DailyReportRun, DailyReportObservation]],
    resolution: DailyReportResolution | None,
) -> list[dict[str, int | None]]:
    candidates: list[dict[str, int | None]] = []
    if _has_confirmation_baseline(resolution):
        assert resolution is not None
        candidates.append(_final_values(resolution))
    candidates.extend(
        _value_dict(observation)
        for _, observation in _capture_rows_after_confirmation(
            capture_rows,
            resolution,
        )
    )
    if _manual_candidate_is_pending(resolution):
        assert resolution is not None
        candidates.append(_manual_values(resolution))
    return candidates


def _version_differences(
    capture_rows: list[tuple[DailyReportRun, DailyReportObservation]],
    resolution: DailyReportResolution | None,
) -> list[str]:
    candidates = _version_candidate_values(capture_rows, resolution)
    return [
        key
        for key in _RECONCILIATION_KEYS
        if _has_non_null_difference(_non_null_values(candidates, key))
    ]


def _source_values(
    session: Session,
    resolution: DailyReportResolution,
    source: str,
) -> dict[str, int | None] | None:
    if source == "latest":
        latest = _latest_observation_any_slot(
            session,
            resolution.business_date,
            resolution.offer_id,
        )
        return _value_dict(latest) if latest is not None else None
    if source == "manual":
        manual = _manual_values(resolution)
        if not any(value is not None for value in manual.values()):
            return None
        base_observation = _latest_observation_any_slot(
            session,
            resolution.business_date,
            resolution.offer_id,
        )
        base = _value_dict(base_observation) if base_observation is not None else {}
        return {
            key: manual[key] if manual[key] is not None else base.get(key)
            for key in _VALUE_KEYS
        }
    observation = _latest_observation(
        session,
        resolution.business_date,
        source,
        resolution.offer_id,
    )
    return _value_dict(observation) if observation is not None else None


def _effective_stock_before_confirmation(
    session: Session,
    resolution: DailyReportResolution,
) -> int | None:
    if _has_confirmation_baseline(resolution):
        return resolution.final_platform_stock
    latest = _latest_observation_any_slot(
        session,
        resolution.business_date,
        resolution.offer_id,
    )
    return latest.platform_stock if latest is not None else None


def _resolution_current_values(
    session: Session,
    resolution: DailyReportResolution,
) -> dict[str, int | None]:
    if _has_confirmation_baseline(resolution):
        return _final_values(resolution)
    captured = _all_observations(session, resolution.business_date).get(
        resolution.offer_id,
        [],
    )
    if captured:
        return _coalesced_capture_values(
            [_value_dict(observation) for _, observation in captured]
        )
    manual = _manual_values(resolution)
    return (
        manual
        if any(value is not None for value in manual.values())
        else {key: None for key in _VALUE_KEYS}
    )


def _propagate_confirmation_stock_conflict(
    session: Session,
    *,
    resolution: DailyReportResolution,
    previous_effective_stock: int | None,
    confirmed_values: Mapping[str, int | None],
    source: str,
    note: str,
    user_id: int,
    confirmed_at: datetime,
    resolved_version_difference: bool = False,
) -> None:
    """Recheck the following report day after a correct value is confirmed."""
    confirmed_stock = confirmed_values.get("platform_stock")
    if confirmed_stock is None:
        return
    if (
        previous_effective_stock == confirmed_stock
        and not resolved_version_difference
    ):
        return
    next_date = session.scalar(
        select(func.min(DailyReportResolution.business_date)).where(
            DailyReportResolution.business_date > resolution.business_date
        )
    )
    if next_date is None:
        return
    following = session.scalar(
        select(DailyReportResolution).where(
            DailyReportResolution.business_date == next_date,
            DailyReportResolution.offer_id == resolution.offer_id,
        )
    )
    if following is None:
        return
    following_capture_rows = _all_observations(session, next_date).get(
        following.offer_id,
        [],
    )
    if _version_differences(following_capture_rows, following):
        following.status = "needs_review"
        following.updated_at = confirmed_at
        return
    current_values = _resolution_current_values(session, following)
    ordered_units = current_values.get("ordered_units")
    actual_stock = current_values.get("platform_stock")
    if ordered_units is None or actual_stock is None:
        return
    orders = int(ordered_units)
    expected_before = (
        previous_effective_stock - orders
        if previous_effective_stock is not None
        else None
    )
    expected_after = int(confirmed_stock) - orders
    mismatch_before = (
        expected_before is not None and expected_before != actual_stock
    )
    mismatch_after = expected_after != actual_stock
    if not mismatch_after:
        if following.status == "needs_review":
            latest = _latest_observation_any_slot(
                session,
                next_date,
                following.offer_id,
            )
            if latest is not None:
                following.status = _status_after_capture(
                    session,
                    resolution=following,
                    slot="manual",
                    incoming=_value_dict(latest),
                    previous_stock=int(confirmed_stock),
                    capture_rows=following_capture_rows,
                )
                following.updated_at = confirmed_at
                _resolve_deadline_if_complete(
                    session,
                    next_date,
                    confirmed_at,
                )
        return

    following_previous_status = following.status
    following_previous_final = (
        _final_values(following)
        if _has_confirmation_baseline(following)
        else None
    )
    following_previous_confirmed_by = _user_display_name(
        session,
        following.confirmed_by,
    )
    following.status = "needs_review"
    following.stock_alert_dismissed = False
    following.stock_alert_note = None
    following.stock_alert_dismissed_by = None
    following.stock_alert_dismissed_at = None
    following.updated_at = confirmed_at
    confirmer = _user_display_name(session, user_id)
    trigger_message = (
        f"{resolution.business_date.isoformat()} 人工确认合并后，"
        f"触发 {next_date.isoformat()} 库存连续性冲突"
    )
    trigger = {
        "kind": "previous_confirmation",
        "message": trigger_message,
        "trigger_business_date": resolution.business_date.isoformat(),
        "affected_business_date": next_date.isoformat(),
        "confirmation_source": source,
        "confirmation_source_label": _confirmation_source_label(source),
        "confirmed_by": confirmer,
        "confirmed_at": confirmed_at.isoformat(),
        "confirmation_note": note,
        "previous_stock_before_confirmation": previous_effective_stock,
        "confirmed_previous_stock": int(confirmed_stock),
        "current_ordered_units": orders,
        "expected_stock_before_confirmation": expected_before,
        "comparison_before_state": (
            "unavailable"
            if expected_before is None
            else ("mismatch" if mismatch_before else "matched")
        ),
        "expected_stock_after_confirmation": expected_after,
        "actual_stock": int(actual_stock),
        "affected_previous_status": following_previous_status,
        "affected_previous_final": following_previous_final,
        "affected_previous_confirmed_by": following_previous_confirmed_by,
        "affected_previous_confirmed_at": (
            following.confirmed_at.isoformat()
            if following.confirmed_at is not None
            else None
        ),
        "affected_previous_confirm_note": following.confirm_note,
        "affected_current_values": current_values,
    }
    _audit(
        session,
        next_date,
        following.offer_id,
        "stock_conflict_after_confirmation",
        trigger,
        trigger_message,
        user_id,
        confirmed_at,
    )
    deadline = session.get(DailyReportDeadlineSnapshot, next_date)
    if deadline is not None:
        deadline.resolved_at = None
        deadline.unresolved_count = int(
            session.scalar(
                select(func.count(DailyReportResolution.id)).where(
                    DailyReportResolution.business_date == next_date,
                    DailyReportResolution.status.in_(OPEN_STATUSES),
                )
            )
            or 0
        )
        details = list(deadline.details or [])
        details.append(
            {
                "offer_id": following.offer_id,
                "status": "needs_review",
                "reason": "stock_conflict_after_confirmation",
            }
        )
        deadline.details = details


def _queue_following_revert_impact(
    session: Session,
    *,
    resolution: DailyReportResolution,
    previous_confirmation: Mapping[str, Any],
    reverted_by: str | None,
    revert_note: str,
    reverted_at: datetime,
    user_id: int,
) -> None:
    """Keep the following day pending until the reverted value is reconfirmed."""
    next_date = session.scalar(
        select(func.min(DailyReportResolution.business_date)).where(
            DailyReportResolution.business_date > resolution.business_date
        )
    )
    if next_date is None:
        return
    following = session.scalar(
        select(DailyReportResolution).where(
            DailyReportResolution.business_date == next_date,
            DailyReportResolution.offer_id == resolution.offer_id,
        )
    )
    if following is None:
        return
    latest = _latest_observation_any_slot(
        session,
        next_date,
        following.offer_id,
    )
    if latest is None:
        return
    previous_status = following.status
    current_values = _resolution_current_values(session, following)
    previous_stock = previous_confirmation.get("values", {}).get(
        "platform_stock"
    )
    ordered_units = current_values.get("ordered_units")
    actual_stock = current_values.get("platform_stock")
    expected_stock = (
        int(previous_stock) - int(ordered_units)
        if previous_stock is not None and ordered_units is not None
        else None
    )
    following.status = "needs_review"
    following.updated_at = reverted_at
    message = (
        f"{resolution.business_date.isoformat()} 人工确认撤销后，"
        f"{next_date.isoformat()} 库存连续性等待重新核对"
    )
    _audit(
        session,
        next_date,
        following.offer_id,
        "stock_continuity_after_confirmation_revert",
        {
            "kind": "previous_confirmation_reverted",
            "message": message,
            "trigger_business_date": resolution.business_date.isoformat(),
            "affected_business_date": next_date.isoformat(),
            "previous_confirmation": dict(previous_confirmation),
            "reverted_by": reverted_by,
            "reverted_at": reverted_at.isoformat(),
            "revert_note": revert_note,
            "current_ordered_units": ordered_units,
            "expected_stock_before_revert": expected_stock,
            "actual_stock": actual_stock,
            "affected_previous_status": previous_status,
            "affected_current_values": current_values,
        },
        message,
        user_id,
        reverted_at,
    )
    deadline = session.get(DailyReportDeadlineSnapshot, next_date)
    if deadline is not None:
        deadline.resolved_at = None
        unresolved = list(
            session.scalars(
                select(DailyReportResolution).where(
                    DailyReportResolution.business_date == next_date,
                    DailyReportResolution.status.in_(OPEN_STATUSES),
                )
            )
        )
        deadline.unresolved_count = len(unresolved)
        deadline.details = [
            {
                "offer_id": row.offer_id,
                "status": row.status,
                **(
                    {"reason": "confirmation_revert_impact"}
                    if row.offer_id == following.offer_id
                    else {}
                ),
            }
            for row in unresolved
        ]


def _defer_following_stock_continuity(
    session: Session,
    *,
    resolution: DailyReportResolution,
    deferred_at: datetime,
) -> None:
    """Pause a following-day continuity task while this day's versions differ."""
    next_date = session.scalar(
        select(func.min(DailyReportResolution.business_date)).where(
            DailyReportResolution.business_date > resolution.business_date
        )
    )
    if next_date is None:
        return
    following = session.scalar(
        select(DailyReportResolution).where(
            DailyReportResolution.business_date == next_date,
            DailyReportResolution.offer_id == resolution.offer_id,
        )
    )
    if following is None:
        return
    latest = _latest_observation_any_slot(
        session,
        next_date,
        following.offer_id,
    )
    if latest is None:
        return
    previous_status = following.status
    following.status = _status_after_capture(
        session,
        resolution=following,
        slot="manual",
        incoming=_value_dict(latest),
        previous_stock=None,
    )
    following.updated_at = deferred_at
    if previous_status == "needs_review" and following.status != "needs_review":
        _audit(
            session,
            next_date,
            following.offer_id,
            "stock_continuity_deferred",
            {
                "blocking_business_date": resolution.business_date.isoformat(),
                "previous_status": previous_status,
                "new_status": following.status,
            },
            "前一日报日正确值未定，库存连续性待办暂停，待正确值确认后重算。",
            None,
            deferred_at,
        )
        deadline = session.get(DailyReportDeadlineSnapshot, next_date)
        if deadline is not None:
            unresolved = list(
                session.scalars(
                    select(DailyReportResolution).where(
                        DailyReportResolution.business_date == next_date,
                        DailyReportResolution.status.in_(OPEN_STATUSES),
                    )
                )
            )
            deadline.unresolved_count = len(unresolved)
            deadline.details = [
                {"offer_id": row.offer_id, "status": row.status}
                for row in unresolved
            ]
            deadline.resolved_at = deferred_at if not unresolved else None


def _confirmation_source_label(source: str) -> str:
    return {
        "morning": "10:05早间值",
        "evening": "18:00晚间值",
        "latest": "本周期最新拉取值",
        "manual": "人工修改值",
    }.get(source, source)


def _user_display_name(session: Session, user_id: int | None) -> str | None:
    if user_id is None:
        return None
    user = session.get(ErpUser, user_id)
    return user.display_name if user is not None else f"用户 {user_id}"


def _apply_final(
    resolution: DailyReportResolution,
    values: Mapping[str, int | None],
    source: str,
    note: str,
    user_id: int,
    now: datetime,
) -> None:
    resolution.selected_source = source
    resolution.final_page_views_30_days = values["page_views_30_days"]
    resolution.final_ordered_units = values["ordered_units"]
    resolution.final_platform_stock = values["platform_stock"]
    resolution.confirm_note = note
    resolution.confirmed_by = user_id
    resolution.confirmed_at = now
    resolution.status = "confirmed"
    resolution.updated_at = now


def _item_payload(
    offer_id: str,
    morning: DailyReportObservation | None,
    evening: DailyReportObservation | None,
    capture_rows: list[tuple[DailyReportRun, DailyReportObservation]],
    successful_runs: list[DailyReportRun],
    resolution: DailyReportResolution | None,
    previous_context: dict[str, Any] | None,
    capture_status: dict[str, dict[str, Any]],
    confirmation_trigger: dict[str, Any] | None,
    operator_notes: list[dict[str, Any]],
    confirmation_baseline: dict[str, Any] | None,
    confirmation_revert: dict[str, Any] | None,
) -> dict[str, Any]:
    report_date = (
        resolution.business_date
        if resolution is not None
        else capture_rows[0][0].business_date
    )
    previous_stock = (
        previous_context.get("stock") if previous_context is not None else None
    )
    identity = capture_rows[-1][1] if capture_rows else evening or morning
    morning_values = _value_dict(morning) if morning is not None else None
    evening_values = _value_dict(evening) if evening is not None else None
    capture_values = [_value_dict(observation) for _, observation in capture_rows]
    manual_values = _manual_values(resolution) if resolution is not None else None
    final_values = _final_values(resolution) if resolution is not None else None
    has_confirmation_baseline = _has_confirmation_baseline(resolution)
    if has_confirmation_baseline:
        current = final_values or {}
    elif capture_values:
        current = _coalesced_capture_values(capture_values)
    elif manual_values is not None and any(value is not None for value in manual_values.values()):
        current = {
            key: manual_values[key]
            for key in _VALUE_KEYS
        }
    else:
        current = {key: None for key in _VALUE_KEYS}
    differences = _version_differences(capture_rows, resolution)
    review_capture_rows = _capture_rows_after_confirmation(
        capture_rows,
        resolution,
    )
    review_versions: list[dict[str, Any]] = []
    if confirmation_baseline is not None:
        review_versions.append(
            {
                "kind": "confirmed",
                "run_id": None,
                "slot": None,
                "label": "上次人工确认值",
                "captured_at": confirmation_baseline["confirmed_at"],
                "values": confirmation_baseline["values"],
                "source_label": confirmation_baseline["source_label"],
                "user_name": confirmation_baseline["confirmed_by"],
                "note": confirmation_baseline["confirm_note"],
            }
        )
    review_versions.extend(
        {
            "kind": "capture",
            "run_id": run.run_id,
            "slot": run.slot,
            "label": _capture_run_label(run),
            "captured_at": run.captured_at.isoformat(),
            "values": _value_dict(observation),
            "source_label": None,
            "user_name": None,
            "note": None,
        }
        for run, observation in review_capture_rows
    )
    if _manual_candidate_is_pending(resolution):
        assert resolution is not None
        review_versions.append(
            {
                "kind": "manual",
                "run_id": None,
                "slot": None,
                "label": "人工候选值",
                "captured_at": (
                    resolution.manual_at.isoformat()
                    if resolution.manual_at is not None
                    else None
                ),
                "values": _manual_values(resolution),
                "source_label": "人工修改值",
                "user_name": None,
                "note": resolution.manual_note,
            }
        )
    missing_slots = [
        slot
        for slot, values in (("morning", morning_values), ("evening", evening_values))
        if values is None and capture_status[slot]["status"] != "pending"
    ]
    observed_run_ids = {run.run_id for run, _ in capture_rows}
    missing_runs = [
        run for run in successful_runs if run.run_id not in observed_run_ids
    ]
    missing_fields = [
        key
        for key in _MISSING_CAPTURE_KEYS
        if current.get(key) is None
    ]
    missing_capture = bool(missing_fields)
    inventory_snapshot_missing = bool(
        capture_rows
        and all(
            observation.platform_stock is None
            and str(observation.stock_source or "").startswith(
                "next_morning_1005"
            )
            for _, observation in capture_rows
        )
    )
    stored_status = resolution.status if resolution is not None else "awaiting_evening"
    missing_reasons = (
        [
            _missing_slot_reason(slot, capture_status[slot])
            for slot in missing_slots
        ]
        if missing_capture
        else []
    )
    if missing_capture and missing_runs:
        missing_reasons.append(
            "以下成功采集批次没有返回该商品："
            + "、".join(_capture_run_label(run) for run in missing_runs)
            + "；本次记为漏爬，不计入数据冲突"
        )
    if missing_capture and inventory_snapshot_missing:
        inventory_date = report_date + timedelta(days=1)
        missing_reasons.append(
            f"{inventory_date.isoformat()} 10:05期末库存快照缺失；"
            f"{report_date.isoformat()}整日销量仍保留，"
            "其他时点库存未用于代替"
        )
    generic_missing_fields = [
        field
        for field in missing_fields
        if field != "platform_stock" or not inventory_snapshot_missing
    ]
    if generic_missing_fields:
        missing_reasons.append(
            "当前采用值的"
            + "、".join(_field_label(field) for field in generic_missing_fields)
            + "字段为空"
        )
    orders = int(current.get("ordered_units") or 0)
    current_stock = current.get("platform_stock")
    previous_ready = bool(
        previous_context is not None
        and previous_context.get("continuity_ready", True)
    )
    previous_confirmation_reverted = bool(
        previous_context is not None
        and previous_context.get("source") == "confirmation_reverted"
    )
    expected_stock = (
        previous_stock - orders
        if previous_stock is not None and previous_ready and not differences
        else None
    )
    mismatch = (
        expected_stock is not None
        and current_stock is not None
        and expected_stock != current_stock
    )
    pending_manual = _manual_candidate_is_pending(resolution)
    manual_candidate = (
        {
            key: (
                manual_values[key]
                if manual_values is not None and manual_values[key] is not None
                else current.get(key)
            )
            for key in _VALUE_KEYS
        }
        if pending_manual
        else None
    )
    manual_candidate_mismatch = (
        _stock_continuity_mismatch(previous_stock, manual_candidate)
        if manual_candidate is not None
        and previous_ready
        and not differences
        else None
    )
    stock_resolution_action = (
        "eliminate"
        if mismatch
        and manual_candidate is not None
        and manual_candidate.get("platform_stock") is not None
        and manual_candidate_mismatch is False
        else "confirm_difference"
    )
    dismissed = (
        resolution.stock_alert_dismissed if resolution is not None else False
    )
    if differences:
        status = "needs_review"
    elif mismatch and not dismissed:
        status = "needs_review"
    elif confirmation_revert is not None and not has_confirmation_baseline:
        status = "needs_review"
    elif previous_confirmation_reverted:
        status = "needs_review"
    elif missing_capture:
        status = "missing_capture"
    elif has_confirmation_baseline:
        status = "confirmed"
    elif stored_status == "missing_capture":
        status = (
            "awaiting_evening"
            if capture_status["evening"]["status"] == "pending"
            else "ready"
        )
    elif stored_status == "needs_review":
        status = "ready"
    else:
        status = stored_status
    review_issues = []
    if differences:
        review_issues.append(
            {
                "type": "capture_difference",
                "fields": differences,
            }
        )
    elif mismatch and not dismissed:
        review_issues.append(
            {
                "type": "stock_continuity",
                "fields": ["ordered_units", "platform_stock"],
            }
        )
    elif confirmation_revert is not None and not has_confirmation_baseline:
        review_issues.append(
            {
                "type": "confirmation_reverted",
                "fields": [],
            }
        )
    elif previous_confirmation_reverted:
        review_issues.append(
            {
                "type": "confirmation_revert_impact",
                "fields": ["ordered_units", "platform_stock"],
            }
        )
    return {
        "offer_id": offer_id,
        "sku": identity.sku if identity is not None else None,
        "title": identity.title if identity is not None else offer_id,
        "status": status,
        "morning": morning_values,
        "evening": evening_values,
        "capture_versions": [
            {
                "run_id": run.run_id,
                "slot": run.slot,
                "label": _capture_run_label(run),
                "captured_at": run.captured_at.isoformat(),
                "values": values,
            }
            for (run, _), values in zip(capture_rows, capture_values, strict=True)
        ],
        "manual": manual_values,
        "manual_reason": resolution.manual_reason if resolution is not None else None,
        "manual_note": resolution.manual_note if resolution is not None else None,
        "manual_at": (
            resolution.manual_at.isoformat()
            if resolution is not None and resolution.manual_at is not None
            else None
        ),
        "final": final_values if has_confirmation_baseline else None,
        "confirmation_baseline": confirmation_baseline,
        "confirmation_revert": (
            confirmation_revert if not has_confirmation_baseline else None
        ),
        "review_versions": review_versions,
        "selected_source": resolution.selected_source if resolution is not None else None,
        "confirm_note": resolution.confirm_note if resolution is not None else None,
        "operator_note": resolution.operator_note if resolution is not None else None,
        "operator_notes": operator_notes,
        "confirmation_trigger": confirmation_trigger,
        "differences": differences,
        "review_issues": review_issues,
        "missing_capture": missing_capture,
        "missing_slots": missing_slots,
        "missing_run_ids": [run.run_id for run in missing_runs],
        "missing_fields": missing_fields,
        "missing_reason": "；".join(missing_reasons) or None,
        "current": current,
        "stock_context": previous_context,
        "stock_check": {
            "previous_stock": previous_stock,
            "expected_stock": expected_stock,
            "actual_stock": current_stock,
            "mismatch": mismatch,
            "dismissed": dismissed,
            "note": resolution.stock_alert_note if resolution is not None else None,
            "resolution_action": stock_resolution_action,
            "deferred_reason": (
                "当前日报日仍有同周期版本差异，确认正确版本后再计算库存连续性"
                if differences
                else (
                    (
                        "前一日报日的人工确认已撤销；本日保留待办，待重新确认后立即重算库存连续性"
                        if previous_confirmation_reverted
                        else "前一日报日仍有同周期版本差异，确认正确版本后再计算库存连续性"
                    )
                    if previous_context is not None and not previous_ready
                    else None
                )
            ),
        },
    }


def _comparison_history(
    session: Session,
    *,
    through: date,
    current_items: list[dict[str, Any]],
    limit: int = 30,
) -> list[dict[str, Any]]:
    recent_dates = list(
        session.scalars(
            select(DailyReportResolution.business_date)
            .where(DailyReportResolution.business_date <= through)
            .distinct()
            .order_by(DailyReportResolution.business_date.desc())
            .limit(limit)
        )
    )
    recent_dates.reverse()
    history: list[dict[str, Any]] = []
    for report_date in recent_dates:
        if report_date == through:
            items = current_items
        else:
            items = _comparison_items_for_date(session, report_date)
        history.append(
            {
                "business_date": report_date.isoformat(),
                "items": [
                    {
                        "offer_id": item["offer_id"],
                        "sku": item["sku"],
                        "title": item["title"],
                        "status": item["status"],
                        "missing_capture": item["missing_capture"],
                        "missing_reason": item["missing_reason"],
                        "current": item["current"],
                        "stock_check": item["stock_check"],
                        "operator_note": item["operator_note"],
                        "operator_notes": item["operator_notes"],
                        "confirmation_baseline": item["confirmation_baseline"],
                        "confirmation_revert": item["confirmation_revert"],
                    }
                    for item in items
                ],
            }
        )
    return history


def _pending_actions(
    session: Session,
    *,
    through: date,
    current_items: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows = list(
        session.execute(
            select(
                DailyReportResolution.business_date,
                DailyReportResolution.offer_id,
            )
            .where(
                DailyReportResolution.business_date <= through,
                DailyReportResolution.status.in_(OPEN_STATUSES),
            )
            .order_by(
                DailyReportResolution.business_date,
                DailyReportResolution.offer_id,
            )
        ).all()
    )
    if not rows:
        return []
    requested_by_date: dict[date, set[str]] = {}
    for report_date, offer_id in rows:
        requested_by_date.setdefault(report_date, set()).add(offer_id)
    result: list[dict[str, Any]] = []
    for report_date, offer_ids in requested_by_date.items():
        items = (
            current_items
            if report_date == through
            else _comparison_items_for_date(session, report_date)
        )
        for item in items:
            if item["offer_id"] not in offer_ids or item["status"] != "needs_review":
                continue
            result.append({"business_date": report_date.isoformat(), **item})
    return result


def _handled_actions(
    session: Session,
    *,
    through: date,
    limit: int = 100,
) -> list[dict[str, Any]]:
    """Return append-only todo completions with their current undo state."""
    audits = list(
        session.scalars(
            select(DailyReportAudit)
            .where(
                DailyReportAudit.business_date <= through,
                DailyReportAudit.offer_id.is_not(None),
                DailyReportAudit.action.in_(HANDLED_AUDIT_ACTIONS),
            )
            .order_by(DailyReportAudit.created_at, DailyReportAudit.id)
        )
    )
    if not audits:
        return []
    offer_ids = sorted(
        {str(row.offer_id) for row in audits if row.offer_id is not None}
    )
    user_ids = {row.user_id for row in audits if row.user_id is not None}
    users = {
        user.id: user.display_name
        for user in session.scalars(
            select(ErpUser).where(ErpUser.id.in_(user_ids))
        )
    }
    identities = {
        row.offer_id: row
        for row in session.scalars(
            select(OfferCurrent).where(OfferCurrent.offer_id.in_(offer_ids))
        )
    }
    snapshot_identities: dict[
        tuple[date, str],
        DailyReportObservation,
    ] = {}
    identity_rows = session.execute(
        select(DailyReportRun.business_date, DailyReportObservation)
        .join(
            DailyReportObservation,
            DailyReportObservation.run_id == DailyReportRun.run_id,
        )
        .where(
            DailyReportRun.business_date <= through,
            DailyReportObservation.offer_id.in_(offer_ids),
        )
        .order_by(
            DailyReportRun.captured_at,
            DailyReportObservation.id,
        )
    )
    for report_date, observation in identity_rows:
        snapshot_identities[(report_date, observation.offer_id)] = observation
    resolutions = {
        (row.business_date, row.offer_id): row
        for row in session.scalars(
            select(DailyReportResolution).where(
                DailyReportResolution.business_date <= through,
                DailyReportResolution.offer_id.in_(offer_ids),
            )
        )
    }
    result: list[dict[str, Any]] = []
    active: dict[tuple[date, str, str], dict[str, Any]] = {}

    def identity_for(
        report_date: date,
        offer_id: str,
    ) -> tuple[str | None, str]:
        snapshot_identity = snapshot_identities.get((report_date, offer_id))
        current_identity = identities.get(offer_id)
        sku = (
            snapshot_identity.sku
            if snapshot_identity is not None and snapshot_identity.sku
            else current_identity.sku
            if current_identity is not None
            else None
        )
        title = (
            snapshot_identity.title
            if snapshot_identity is not None and snapshot_identity.title
            else current_identity.title
            if current_identity is not None and current_identity.title
            else offer_id
        )
        return sku, title

    def audit_user_name(audit: DailyReportAudit) -> str:
        return (
            users.get(audit.user_id, "系统")
            if audit.user_id is not None
            else "系统"
        )

    def audit_values(
        audit: DailyReportAudit,
        payload: dict[str, Any],
    ) -> dict[str, int | None]:
        values = payload.get("values")
        if not isinstance(values, dict):
            values = payload.get("after")
        if not isinstance(values, dict):
            resolution = resolutions.get(
                (audit.business_date, str(audit.offer_id))
            )
            values = (
                _export_values(session, resolution)
                if resolution is not None
                else {}
            )
        return {
            key_name: values.get(key_name)
            for key_name in _VALUE_KEYS
        }

    def support_entry(
        audit: DailyReportAudit,
        payload: dict[str, Any],
    ) -> dict[str, Any]:
        assert audit.offer_id is not None
        offer_id = audit.offer_id
        sku, title = identity_for(audit.business_date, offer_id)
        current = audit_values(audit, payload)
        before_values = payload.get("before")
        after_values = payload.get("after")
        return {
            "id": audit.id,
            "action_type": audit.action,
            "business_date": audit.business_date.isoformat(),
            "offer_id": offer_id,
            "sku": sku,
            "title": title,
            "handled_by": audit_user_name(audit),
            "handled_at": audit.created_at.isoformat(),
            "note": audit.note,
            "active": False,
            "reversal": None,
            "current": current,
            "detail": {
                "source": None,
                "source_label": None,
                "previous_stock": None,
                "ordered_units": current.get("ordered_units"),
                "expected_stock": None,
                "actual_stock": current.get("platform_stock"),
                "reason": payload.get("reason"),
                "issue_type": payload.get("issue_type"),
                "before_note": payload.get("before_note"),
                "after_note": payload.get("note"),
                "deleted_note": payload.get("deleted_note"),
                "before_values": (
                    {
                        key_name: before_values.get(key_name)
                        for key_name in _VALUE_KEYS
                    }
                    if isinstance(before_values, dict)
                    else None
                ),
                "after_values": (
                    {
                        key_name: after_values.get(key_name)
                        for key_name in _VALUE_KEYS
                    }
                    if isinstance(after_values, dict)
                    else None
                ),
            },
        }

    for audit in audits:
        if audit.offer_id is None:
            continue
        offer_id = audit.offer_id
        payload = audit.payload if isinstance(audit.payload, dict) else {}
        if audit.action in HANDLED_SUPPORT_ACTIONS:
            result.append(support_entry(audit, payload))
            continue
        action_type = (
            "stock_difference"
            if audit.action == "dismiss_stock_alert"
            else "stock_eliminated"
            if audit.action == "eliminate_stock_alert"
            else "confirmation"
            if audit.action in {"confirm", "bulk_confirm"}
            else None
        )
        if action_type is not None:
            state_type = (
                "confirmation"
                if action_type in {"confirmation", "stock_eliminated"}
                else action_type
            )
            key = (audit.business_date, offer_id, state_type)
            previous = active.get(key)
            if previous is not None and previous["active"]:
                previous["active"] = False
                previous["reversal"] = {
                    "kind": "superseded",
                    "handled_by": audit_user_name(audit),
                    "handled_at": audit.created_at.isoformat(),
                    "note": "已被后续处理记录替代",
                }
            current = audit_values(audit, payload)
            sku, title = identity_for(audit.business_date, offer_id)
            entry = {
                "id": audit.id,
                "action_type": action_type,
                "business_date": audit.business_date.isoformat(),
                "offer_id": offer_id,
                "sku": sku,
                "title": title,
                "handled_by": audit_user_name(audit),
                "handled_at": audit.created_at.isoformat(),
                "note": audit.note,
                "active": True,
                "reversal": None,
                "current": current,
                "detail": {
                    "source": payload.get("source"),
                    "source_label": (
                        _confirmation_source_label(str(payload.get("source")))
                        if action_type in {"confirmation", "stock_eliminated"}
                        and payload.get("source") is not None
                        else None
                    ),
                    "previous_stock": payload.get("previous_stock"),
                    "ordered_units": payload.get(
                        "ordered_units",
                        current.get("ordered_units"),
                    ),
                    "expected_stock": payload.get("expected_stock"),
                    "actual_stock": payload.get(
                        "actual_stock",
                        current.get("platform_stock"),
                    ),
                    "automatic": bool(payload.get("automatic", False)),
                    "reason": None,
                    "issue_type": None,
                    "before_note": None,
                    "after_note": None,
                    "deleted_note": None,
                    "before_values": None,
                    "after_values": None,
                },
            }
            result.append(entry)
            active[key] = entry
            continue
        reversal_type = (
            "confirmation"
            if audit.action == "confirmation_reverted"
            else "stock_difference"
        )
        key = (audit.business_date, offer_id, reversal_type)
        target = active.get(key)
        if target is not None and target["active"]:
            target["active"] = False
            target["reversal"] = {
                "kind": audit.action,
                "handled_by": audit_user_name(audit),
                "handled_at": audit.created_at.isoformat(),
                "note": audit.note,
            }
        result.append(support_entry(audit, payload))
    for key, entry in active.items():
        report_date, offer_id, state_type = key
        resolution = resolutions.get((report_date, offer_id))
        currently_active = bool(
            resolution is not None
            and (
                (
                    state_type == "confirmation"
                    and _has_confirmation_baseline(resolution)
                )
                or (
                    state_type == "stock_difference"
                    and resolution.stock_alert_dismissed
                )
            )
        )
        if entry["active"] and not currently_active:
            entry["active"] = False
            entry["reversal"] = {
                "kind": "state_changed",
                "handled_by": "系统",
                "handled_at": (
                    resolution.updated_at.isoformat()
                    if resolution is not None
                    else entry["handled_at"]
                ),
                "note": "后续数据或操作已经替代此处理状态",
            }
    return list(reversed(result))[:limit]


def _comparison_items_for_date(
    session: Session,
    business_date: date,
) -> list[dict[str, Any]]:
    runs = list(
        session.scalars(
            select(DailyReportRun)
            .where(DailyReportRun.business_date == business_date)
            .order_by(DailyReportRun.captured_at)
        )
    )
    capture_status = _capture_status(runs, business_date)
    observations = _latest_observations(session, business_date)
    all_observations = _all_observations(session, business_date)
    successful_runs = [run for run in runs if run.status == "success"]
    confirmation_triggers = _confirmation_trigger_map(session, business_date)
    operator_notes = _operator_note_history_map(session, business_date)
    resolutions = {
        row.offer_id: row
        for row in session.scalars(
            select(DailyReportResolution).where(
                DailyReportResolution.business_date == business_date
            )
        )
    }
    confirmation_baselines = _confirmation_baseline_map(
        session,
        list(resolutions.values()),
    )
    confirmation_reverts = _confirmation_revert_map(
        session,
        business_date,
    )
    previous_contexts = _previous_stock_contexts(session, business_date)
    offer_ids = sorted(all_observations)
    return [
        _item_payload(
            offer_id,
            observations["morning"].get(offer_id),
            observations["evening"].get(offer_id),
            all_observations.get(offer_id, []),
            successful_runs,
            resolutions.get(offer_id),
            previous_contexts.get(offer_id),
            capture_status,
            confirmation_triggers.get(offer_id),
            operator_notes.get(offer_id, []),
            confirmation_baselines.get(offer_id),
            confirmation_reverts.get(offer_id),
        )
        for offer_id in offer_ids
    ]


def _confirmation_trigger_map(
    session: Session,
    business_date: date,
) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    rows = session.scalars(
        select(DailyReportAudit)
        .where(
            DailyReportAudit.business_date == business_date,
            DailyReportAudit.action == "stock_conflict_after_confirmation",
        )
        .order_by(DailyReportAudit.created_at, DailyReportAudit.id)
    )
    for row in rows:
        if row.offer_id is not None and isinstance(row.payload, dict):
            result[row.offer_id] = row.payload
    return result


def _operator_note_history_map(
    session: Session,
    business_date: date,
) -> dict[str, list[dict[str, Any]]]:
    rows = list(
        session.scalars(
            select(DailyReportAudit)
            .where(
                DailyReportAudit.business_date == business_date,
                DailyReportAudit.action.in_(NOTE_AUDIT_ACTIONS),
            )
            .order_by(DailyReportAudit.created_at, DailyReportAudit.id)
        )
    )
    user_ids = {row.user_id for row in rows if row.user_id is not None}
    users = {
        user.id: user.display_name
        for user in session.scalars(
            select(ErpUser).where(ErpUser.id.in_(user_ids))
        )
    }
    result: dict[str, dict[int, dict[str, Any]]] = {}
    for row in rows:
        if row.offer_id is None:
            continue
        payload = row.payload if isinstance(row.payload, dict) else {}
        if row.action == "operator_note_updated":
            note_id = int(payload.get("note_id") or 0)
            existing = result.get(row.offer_id, {}).get(note_id)
            if existing is not None:
                issue_type = str(payload.get("issue_type") or existing["issue_type"])
                if issue_type not in NOTE_ISSUE_TYPES:
                    issue_type = "general"
                existing.update(
                    {
                        "issue_type": issue_type,
                        "note": row.note or str(payload.get("note") or ""),
                        "updated_by": (
                            users.get(row.user_id, "系统")
                            if row.user_id is not None
                            else "系统"
                        ),
                        "updated_at": row.created_at.isoformat(),
                    }
                )
            continue
        if row.action == "operator_note_deleted":
            note_id = int(payload.get("note_id") or 0)
            result.get(row.offer_id, {}).pop(note_id, None)
            continue
        issue_type = str(payload.get("issue_type") or "general")
        if issue_type not in NOTE_ISSUE_TYPES:
            issue_type = "general"
        result.setdefault(row.offer_id, {})[row.id] = {
            "id": row.id,
            "issue_type": issue_type,
            "note": row.note or str(payload.get("note") or ""),
            "user_id": row.user_id,
            "user_name": (
                users.get(row.user_id, "系统")
                if row.user_id is not None
                else "系统"
            ),
            "created_at": row.created_at.isoformat(),
            "updated_by": None,
            "updated_at": None,
        }
    return {
        offer_id: list(notes.values())
        for offer_id, notes in result.items()
        if notes
    }


def _operator_note_cell_text(
    notes: list[dict[str, Any]],
    *,
    confirmation_note: str | None = None,
    stock_confirmation_note: str | None = None,
) -> str | None:
    labels = {
        "general": "通用",
        "capture_difference": "版本",
        "stock_continuity": "库存",
    }
    rendered = (
        [f"（确认：{confirmation_note.strip()}）"]
        if confirmation_note is not None and confirmation_note.strip()
        else []
    )
    if (
        stock_confirmation_note is not None
        and stock_confirmation_note.strip()
    ):
        rendered.append(
            f"（库存差异已确认：{stock_confirmation_note.strip()}）"
        )
    rendered.extend(
        f"（{labels.get(str(note.get('issue_type')), '通用')}：{note.get('note', '')}）"
        for note in notes
        if str(note.get("note") or "").strip()
    )
    return " ".join(rendered) or None


def _confirmation_baseline_map(
    session: Session,
    resolutions: list[DailyReportResolution],
) -> dict[str, dict[str, Any]]:
    confirmed = [
        row for row in resolutions if _has_confirmation_baseline(row)
    ]
    user_ids = {
        row.confirmed_by
        for row in confirmed
        if row.confirmed_by is not None
    }
    users = {
        user.id: user.display_name
        for user in session.scalars(
            select(ErpUser).where(ErpUser.id.in_(user_ids))
        )
    }
    return {
        row.offer_id: {
            "values": _final_values(row),
            "source": row.selected_source,
            "source_label": _confirmation_source_label(
                str(row.selected_source)
            ),
            "confirmed_by": (
                users.get(row.confirmed_by, "系统")
                if row.confirmed_by is not None
                else "系统"
            ),
            "confirmed_at": (
                row.confirmed_at.isoformat()
                if row.confirmed_at is not None
                else None
            ),
            "confirm_note": row.confirm_note,
        }
        for row in confirmed
    }


def _confirmation_revert_map(
    session: Session,
    business_date: date,
) -> dict[str, dict[str, Any]]:
    """Return the latest append-only confirmation rollback for each offer."""
    result: dict[str, dict[str, Any]] = {}
    rows = session.scalars(
        select(DailyReportAudit)
        .where(
            DailyReportAudit.business_date == business_date,
            DailyReportAudit.action == "confirmation_reverted",
        )
        .order_by(DailyReportAudit.created_at, DailyReportAudit.id)
    )
    for row in rows:
        if row.offer_id is not None and isinstance(row.payload, dict):
            result[row.offer_id] = row.payload
    return result


def _previous_stock_contexts(
    session: Session,
    business_date: date,
) -> dict[str, dict[str, Any]]:
    previous_date = session.scalar(
        select(func.max(DailyReportResolution.business_date)).where(
            DailyReportResolution.business_date < business_date
        )
    )
    if previous_date is None:
        return {}
    rows = list(
        session.scalars(
            select(DailyReportResolution).where(
                DailyReportResolution.business_date == previous_date
            )
        )
    )
    captures = _all_observations(session, previous_date)
    confirmation_reverts = _confirmation_revert_map(session, previous_date)
    user_ids = {
        row.confirmed_by
        for row in rows
        if _has_confirmation_baseline(row) and row.confirmed_by is not None
    }
    users = {
        user.id: user.display_name
        for user in session.scalars(
            select(ErpUser).where(ErpUser.id.in_(user_ids))
        )
    }
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        latest_pair = captures.get(row.offer_id, [])
        latest_run, latest_observation = (
            latest_pair[-1] if latest_pair else (None, None)
        )
        version_differences = _version_differences(latest_pair, row)
        confirmation_revert = confirmation_reverts.get(row.offer_id)
        if confirmation_revert is not None and not _has_confirmation_baseline(row):
            result[row.offer_id] = {
                "business_date": previous_date.isoformat(),
                "stock": None,
                "source": "confirmation_reverted",
                "source_label": "前一日报日的人工确认已撤销，待重新确认后再计算库存连续性",
                "selected_source": None,
                "confirmed_by": None,
                "confirmed_at": None,
                "confirm_note": None,
                "capture_label": (
                    _capture_run_label(latest_run)
                    if latest_run is not None
                    else None
                ),
                "continuity_ready": False,
                "version_differences": version_differences,
                "confirmation_revert": confirmation_revert,
            }
            continue
        if version_differences:
            result[row.offer_id] = {
                "business_date": previous_date.isoformat(),
                "stock": None,
                "source": "version_difference",
                "source_label": "前一日报日仍有同周期版本差异，暂不计算库存连续性",
                "selected_source": row.selected_source,
                "confirmed_by": (
                    users.get(row.confirmed_by, "系统")
                    if row.confirmed_by is not None
                    else None
                ),
                "confirmed_at": (
                    row.confirmed_at.isoformat()
                    if row.confirmed_at is not None
                    else None
                ),
                "confirm_note": row.confirm_note,
                "capture_label": (
                    _capture_run_label(latest_run)
                    if latest_run is not None
                    else None
                ),
                "continuity_ready": False,
                "version_differences": version_differences,
            }
            continue
        if _has_confirmation_baseline(row):
            result[row.offer_id] = {
                "business_date": previous_date.isoformat(),
                "stock": row.final_platform_stock,
                "source": "confirmed",
                "source_label": "前一日报日人工确认库存",
                "selected_source": row.selected_source,
                "confirmed_by": (
                    users.get(row.confirmed_by, "系统")
                    if row.confirmed_by is not None
                    else "系统"
                ),
                "confirmed_at": (
                    row.confirmed_at.isoformat()
                    if row.confirmed_at is not None
                    else None
                ),
                "confirm_note": row.confirm_note,
                "capture_label": (
                    _capture_run_label(latest_run)
                    if latest_run is not None
                    else None
                ),
                "continuity_ready": True,
                "version_differences": [],
            }
            continue
        result[row.offer_id] = {
            "business_date": previous_date.isoformat(),
            "stock": (
                latest_observation.platform_stock
                if latest_observation is not None
                else None
            ),
            "source": "latest_capture",
            "source_label": "前一日报日未确认，暂用最后一次成功拉取库存",
            "selected_source": None,
            "confirmed_by": None,
            "confirmed_at": None,
            "confirm_note": None,
            "capture_label": (
                _capture_run_label(latest_run) if latest_run is not None else None
            ),
            "continuity_ready": True,
            "version_differences": [],
        }
    return result


def _previous_values(
    session: Session, business_date: date
) -> dict[str, int | None]:
    return {
        offer_id: context.get("stock")
        for offer_id, context in _previous_stock_contexts(
            session,
            business_date,
        ).items()
    }


def _reminders(
    session: Session, *, before: date | None
) -> list[dict[str, Any]]:
    statement = (
        select(
            DailyReportResolution.business_date,
            func.count(DailyReportResolution.id),
        )
        .where(DailyReportResolution.status.in_(OPEN_STATUSES))
        .group_by(DailyReportResolution.business_date)
        .order_by(DailyReportResolution.business_date)
    )
    if before is not None:
        statement = statement.where(DailyReportResolution.business_date < before)
    return [
        {
            "business_date": business_date.isoformat(),
            "unresolved_count": int(count),
        }
        for business_date, count in session.execute(statement).all()
    ]


def _reminders_for_date(
    session: Session, business_date: date
) -> list[dict[str, Any]]:
    count = int(
        session.scalar(
            select(func.count(DailyReportResolution.id)).where(
                DailyReportResolution.business_date == business_date,
                DailyReportResolution.status.in_(OPEN_STATUSES),
            )
        )
        or 0
    )
    return (
        [{"business_date": business_date.isoformat(), "unresolved_count": count}]
        if count
        else []
    )


def _resolution_or_error(
    session: Session, business_date: date, offer_id: str
) -> DailyReportResolution:
    row = session.scalar(
        select(DailyReportResolution).where(
            DailyReportResolution.business_date == business_date,
            DailyReportResolution.offer_id == offer_id,
        )
    )
    if row is None:
        raise DailyReportConflictError("该商品当天还没有日报采集数据")
    return row


def _resolve_deadline_if_complete(
    session: Session, business_date: date, now: datetime
) -> None:
    remaining = int(
        session.scalar(
            select(func.count(DailyReportResolution.id)).where(
                DailyReportResolution.business_date == business_date,
                DailyReportResolution.status.in_(OPEN_STATUSES),
            )
        )
        or 0
    )
    snapshot = session.get(DailyReportDeadlineSnapshot, business_date)
    if snapshot is not None and remaining == 0:
        snapshot.resolved_at = now


def _refresh_deadline_snapshot(
    session: Session,
    business_date: date,
    now: datetime,
) -> None:
    snapshot = session.get(DailyReportDeadlineSnapshot, business_date)
    if snapshot is None:
        return
    unresolved = list(
        session.scalars(
            select(DailyReportResolution).where(
                DailyReportResolution.business_date == business_date,
                DailyReportResolution.status.in_(OPEN_STATUSES),
            )
        )
    )
    snapshot.unresolved_count = len(unresolved)
    snapshot.details = [
        {"offer_id": row.offer_id, "status": row.status}
        for row in unresolved
    ]
    snapshot.resolved_at = now if not unresolved else None


def _audit(
    session: Session,
    business_date: date,
    offer_id: str | None,
    action: str,
    payload: dict[str, Any] | None,
    note: str | None,
    user_id: int | None,
    now: datetime,
) -> None:
    session.add(
        DailyReportAudit(
            business_date=business_date,
            offer_id=offer_id,
            action=action,
            payload=payload,
            note=note,
            user_id=user_id,
            created_at=now,
        )
    )


def _required_note(note: str, message: str) -> str:
    clean = note.strip()
    if not clean:
        raise DailyReportInputError(message)
    if len(clean) > 2000:
        raise DailyReportInputError("备注不能超过 2000 个字符")
    return clean


def _identity_map(
    session: Session,
    rows: list[DailyReportResolution],
) -> dict[str, dict[str, str | None]]:
    offer_ids = sorted({row.offer_id for row in rows})
    if not offer_ids:
        return {}
    current = {
        row.offer_id: {"sku": row.sku, "title": row.title}
        for row in session.scalars(
            select(OfferCurrent).where(OfferCurrent.offer_id.in_(offer_ids))
        )
    }
    missing = set(offer_ids) - set(current)
    if missing:
        observations = session.scalars(
            select(DailyReportObservation)
            .where(DailyReportObservation.offer_id.in_(missing))
            .order_by(DailyReportObservation.id.desc())
        )
        for row in observations:
            current.setdefault(
                row.offer_id,
                {"sku": row.sku, "title": row.title},
            )
    return current


def _all_previous_stock(
    rows: list[DailyReportResolution],
    values_by_key: dict[tuple[date, str], dict[str, int | None]],
) -> dict[tuple[date, str], int | None]:
    by_offer: dict[str, list[DailyReportResolution]] = {}
    for row in rows:
        by_offer.setdefault(row.offer_id, []).append(row)
    result: dict[tuple[date, str], int | None] = {}
    for offer_id, values in by_offer.items():
        previous_stock: int | None = None
        for row in sorted(values, key=lambda item: item.business_date):
            result[(row.business_date, offer_id)] = previous_stock
            previous_stock = values_by_key[(row.business_date, offer_id)][
                "platform_stock"
            ]
    return result


def _capture_status(
    runs: list[DailyReportRun],
    business_date: date,
) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for slot in ("morning", "evening"):
        slot_runs = [run for run in runs if run.slot == slot]
        successful = [run for run in slot_runs if run.status == "success"]
        selected = (
            max(successful, key=lambda run: run.captured_at)
            if successful
            else max(slot_runs, key=lambda run: run.captured_at, default=None)
        )
        if selected is None:
            scheduled_at = datetime.combine(
                business_date + timedelta(days=1),
                time(10, 5) if slot == "morning" else time(18, 0),
                tzinfo=ZoneInfo("Asia/Shanghai"),
            )
            now = datetime.now(ZoneInfo("Asia/Shanghai"))
            if now < scheduled_at:
                result[slot] = {
                    "status": "pending",
                    "captured_at": None,
                    "product_count": 0,
                    "reason": f"计划于北京时间{scheduled_at:%m月%d日 %H:%M}执行",
                    "attempts": [],
                    "attempt_count": 0,
                    "recovered": False,
                    "capture_method": None,
                }
                continue
            result[slot] = {
                "status": "missing",
                "captured_at": None,
                "product_count": 0,
                "reason": (
                    f"{'10:05早间' if slot == 'morning' else '18:00晚间'}"
                    "未生成采集记录，可能是定时任务未执行、服务未启动或日志缺失"
                ),
                "attempts": [],
                "attempt_count": 0,
                "recovered": False,
                "capture_method": None,
            }
            continue
        counts = selected.counts or {}
        attempts = counts.get("attempts")
        attempt_rows = attempts if isinstance(attempts, list) else []
        recovered = bool(counts.get("recovered"))
        result[slot] = {
            "status": selected.status,
            "captured_at": selected.captured_at.isoformat(),
            "product_count": int(counts.get("products") or 0),
            "reason": (
                str(counts.get("missing_reason") or "")
                if selected.status != "success"
                else (
                    f"第 {len(attempt_rows)} 次尝试恢复成功，"
                    f"采用{counts.get('capture_method') or '标准接口'}"
                    if recovered
                    else None
                )
            ),
            "attempts": attempt_rows,
            "attempt_count": len(attempt_rows),
            "recovered": recovered,
            "capture_method": counts.get("capture_method"),
        }
    return result


def _capture_issues(
    items: list[dict[str, Any]],
    capture_status: dict[str, dict[str, Any]],
    runs: list[DailyReportRun],
) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for slot in ("morning", "evening"):
        state = capture_status[slot]
        if state["status"] not in {"success", "pending"}:
            issues.append(
                {
                    "kind": "slot",
                    "slot": slot,
                    "offer_id": None,
                    "sku": None,
                    "title": None,
                    "reason": state["reason"],
                }
            )
    for run in runs:
        if run.slot != "manual" or run.status == "success":
            continue
        counts = run.counts or {}
        issues.append(
            {
                "kind": "slot",
                "slot": "manual",
                "offer_id": None,
                "sku": None,
                "title": None,
                "reason": str(
                    counts.get("missing_reason")
                    or counts.get("final_reason")
                    or "手动刷新失败，但没有返回具体错误"
                ),
            }
        )
    for item in items:
        product_missing_slots = [
            slot
            for slot in item["missing_slots"]
            if capture_status[slot]["status"] == "success"
        ]
        if item["missing_capture"] and (
            product_missing_slots
            or item["missing_run_ids"]
            or item["missing_fields"]
        ):
            issues.append(
                {
                    "kind": "product",
                    "slot": (
                        product_missing_slots[0]
                        if product_missing_slots
                        else None
                    ),
                    "offer_id": item["offer_id"],
                    "sku": item["sku"],
                    "title": item["title"],
                    "reason": item["missing_reason"],
                }
            )
    return issues


def _missing_slot_reason(slot: str, state: dict[str, Any]) -> str:
    label = "10:05早间" if slot == "morning" else "18:00晚间"
    if state["status"] == "success":
        return (
            f"{label}整次采集成功，但该商品未出现在接口返回结果中；"
            "可能已下架、暂时不可见或为另一时段新增商品，已自动采用可用时段的数据"
        )
    return f"{label}漏爬：{state['reason']}；已自动采用另一时段的可用数据"


def _field_label(key: str) -> str:
    return {
        "page_views_30_days": "近30天浏览量",
        "ordered_units": "当天订单数",
        "platform_stock": "平台库存",
    }.get(key, key)


def _has_non_null_difference(values: list[int]) -> bool:
    return len(values) > 1 and any(value != values[0] for value in values[1:])


def _non_null_values(
    candidates: list[dict[str, int | None]],
    key: str,
) -> list[int]:
    values: list[int] = []
    for candidate in candidates:
        value = candidate.get(key)
        if value is not None:
            values.append(value)
    return values


def _export_values(
    session: Session,
    row: DailyReportResolution,
) -> dict[str, int | None]:
    if row.status == "confirmed":
        return _final_values(row)
    captured = _all_observations(session, row.business_date).get(row.offer_id, [])
    return _coalesced_capture_values(
        [_value_dict(observation) for _, observation in captured]
    )


def _export_missing_rows(
    session: Session,
    rows: list[DailyReportResolution],
    identities: dict[str, dict[str, str | None]],
) -> list[tuple[str, str, str, str, str, str]]:
    del identities
    result: list[tuple[str, str, str, str, str, str]] = []
    dates = sorted({row.business_date for row in rows})
    for report_date in dates:
        runs = list(
            session.scalars(
                select(DailyReportRun).where(
                    DailyReportRun.business_date == report_date
                )
            )
        )
        status = _capture_status(runs, report_date)
        items = _comparison_items_for_date(session, report_date)
        for issue in _capture_issues(items, status, runs):
            slot = issue["slot"]
            slot_label = {
                "morning": "早间",
                "evening": "晚间",
                "manual": "手动刷新",
            }.get(slot, "字段")
            result.append(
                (
                    report_date.isoformat(),
                    slot_label,
                    str(issue["title"] or issue["offer_id"] or "整次采集"),
                    str(issue["sku"] or issue["offer_id"] or "—"),
                    str(issue["reason"]),
                    "采用本周期其他成功版本的可用值",
                )
            )
    return result


def _naive_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value
    return value.astimezone(UTC).replace(tzinfo=None)


def _beijing_date(value: datetime) -> date:
    aware = value.replace(tzinfo=UTC) if value.tzinfo is None else value
    return aware.astimezone(ZoneInfo("Asia/Shanghai")).date()


def _utc_now() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)
