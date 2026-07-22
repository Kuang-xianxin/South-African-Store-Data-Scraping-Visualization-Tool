"""Safe upload and generation helpers for the local NFT102 dashboard page."""

from __future__ import annotations

import hashlib
import re
import subprocess
from collections.abc import Callable, Sequence
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import is_zipfile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.datetime import from_excel  # type: ignore[import-untyped]

from takealot_ops.nft102 import SHEET_NAME, SKU_PATTERN


MAX_UPLOAD_BYTES = 100 * 1024 * 1024
DAILY_BLOCK_LABEL = "访客总数"
SAFE_FILENAME = re.compile(r"[^\w.()（） -]+", re.UNICODE)


@dataclass(frozen=True)
class Nft102UploadInspection:
    """Validated metadata shown before an operator starts generation."""

    filename: str
    size_bytes: int
    sha256: str
    latest_report_date: date
    suggested_report_date: date
    product_columns: int


@dataclass(frozen=True)
class Nft102GenerationResult:
    """Paths created by one successful NFT102 generation run."""

    baseline_path: Path
    workbook_path: Path
    audit_json_path: Path
    audit_text_path: Path
    report_date: date


Runner = Callable[..., subprocess.CompletedProcess[str]]


def inspect_nft102_upload(filename: str, content: bytes) -> Nft102UploadInspection:
    """Validate an uploaded workbook without saving or mutating it."""
    safe_name = _safe_xlsx_filename(filename)
    if not content:
        raise ValueError("上传文件为空。")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError("上传文件超过100兆字节，请确认选择了正确的电子表格。")
    stream = BytesIO(content)
    if not is_zipfile(stream):
        raise ValueError("文件不是有效的电子表格工作簿。")
    stream.seek(0)

    try:
        workbook = load_workbook(stream, read_only=True, data_only=False)
    except (OSError, ValueError, KeyError) as exc:
        raise ValueError("无法读取工作簿，请确认文件没有损坏。") from exc
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"工作簿中缺少 {SHEET_NAME} 工作表。")
        worksheet = workbook[SHEET_NAME]
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        product_columns = sum(
            bool(SKU_PATTERN.search(str(value))) for value in first_row[2:] if value is not None
        )
        report_dates: list[date] = []
        for label, raw_date in worksheet.iter_rows(
            min_col=1, max_col=2, values_only=True
        ):
            if str(label).strip() != DAILY_BLOCK_LABEL:
                continue
            parsed = _workbook_date(raw_date, workbook.epoch)
            if parsed is not None:
                report_dates.append(parsed)
        if not report_dates:
            raise ValueError("NFT102 工作表中没有可识别的历史日报日期。")
        if product_columns == 0:
            raise ValueError("NFT102 表头中没有识别到13位库存编码。")
        latest = max(report_dates)
    finally:
        workbook.close()

    return Nft102UploadInspection(
        filename=safe_name,
        size_bytes=len(content),
        sha256=hashlib.sha256(content).hexdigest(),
        latest_report_date=latest,
        suggested_report_date=latest + timedelta(days=1),
        product_columns=product_columns,
    )


def persist_nft102_baseline(
    project_root: Path,
    inspection: Nft102UploadInspection,
    content: bytes,
    *,
    saved_at: datetime | None = None,
) -> Path:
    """Archive one immutable operator-edited baseline and return its path."""
    if hashlib.sha256(content).hexdigest() != inspection.sha256:
        raise ValueError("上传内容在校验后发生变化，请重新选择文件。")
    timestamp = (saved_at or datetime.now()).strftime("%Y%m%d-%H%M%S-%f")
    folder = (
        project_root.resolve()
        / "data"
        / "nft102-baselines"
        / f"{timestamp}-{inspection.sha256[:8]}"
    )
    folder.mkdir(parents=True, exist_ok=False)
    destination = folder / inspection.filename
    temporary = folder / f".{inspection.filename}.tmp"
    try:
        temporary.write_bytes(content)
        temporary.replace(destination)
    finally:
        temporary.unlink(missing_ok=True)
    return destination


def generate_nft102_from_baseline(
    project_root: Path,
    baseline_path: Path,
    report_date: date,
    *,
    runner: Runner = subprocess.run,
    timeout_seconds: int = 240,
) -> Nft102GenerationResult:
    """Run the proven PowerShell workflow against an explicit uploaded baseline."""
    root = project_root.resolve()
    baseline = baseline_path.resolve()
    if not baseline.is_file() or root not in baseline.parents:
        raise ValueError("基准文件不存在或不在项目目录中。")
    script = root / "scripts" / "update_nft102_daily.ps1"
    if not script.is_file():
        raise ValueError(f"未找到生成脚本：{script}")

    output_folder = root / "outputs" / "nft102-daily" / report_date.isoformat()
    before = set(output_folder.glob("*.xlsx")) if output_folder.is_dir() else set()
    command: Sequence[str] = (
        "powershell.exe",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(script),
        "-TemplatePath",
        str(baseline),
        "-ReportDate",
        report_date.isoformat(),
    )
    try:
        completed = runner(
            command,
            cwd=root,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout_seconds,
            check=False,
        )
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError("生成超过 4 分钟仍未完成，请检查网络后重试。") from exc
    if completed.returncode != 0:
        raise RuntimeError("NFT102 日报生成失败，请检查网络、接口密钥和本地日志。")

    after = set(output_folder.glob("*.xlsx")) if output_folder.is_dir() else set()
    created = sorted(after - before, key=lambda path: path.stat().st_mtime, reverse=True)
    if not created:
        raise RuntimeError("生成脚本已结束，但没有找到新的电子表格。")
    workbook_path = created[0]
    audit_json = workbook_path.with_suffix(".核对报告.json")
    audit_text = workbook_path.with_suffix(".核对报告.txt")
    if not audit_json.is_file() or not audit_text.is_file():
        raise RuntimeError("电子表格已生成，但核对报告不完整，请不要交付运营使用。")
    return Nft102GenerationResult(
        baseline_path=baseline,
        workbook_path=workbook_path,
        audit_json_path=audit_json,
        audit_text_path=audit_text,
        report_date=report_date,
    )


def _safe_xlsx_filename(filename: str) -> str:
    name = Path(filename).name.strip()
    if Path(name).suffix.casefold() != ".xlsx":
        raise ValueError("只支持上传电子表格文件。")
    cleaned = SAFE_FILENAME.sub("_", name).strip(" .")
    if not cleaned or cleaned.casefold() == ".xlsx":
        raise ValueError("文件名无效。")
    return cleaned


def _workbook_date(value: Any, epoch: datetime) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch)
        if isinstance(converted, datetime):
            return converted.date()
        if isinstance(converted, date):
            return converted
        return None
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError:
            return None
    return None
