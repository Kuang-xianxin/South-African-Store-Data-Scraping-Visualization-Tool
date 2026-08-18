"""Audited XLSX import for cross-store product identities and RMB unit costs."""

from __future__ import annotations

import hashlib
import re
import unicodedata
from collections import Counter, defaultdict
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy import Engine, select
from sqlalchemy.orm import Session

from takealot_ops.storage.models import (
    CompanyProduct,
    CompanyProductCost,
    ErpStore,
    LogisticsProviderSnapshot,
    OfferCurrent,
    PlatformSkuMapping,
    ProductMasterImportBatch,
    ProductMasterImportRow,
)
from takealot_ops.storage.store_context import store_scope


MAX_WORKBOOK_BYTES = 50 * 1024 * 1024
MAX_SOURCE_ROWS = 100_000
MAX_HEADER_SCAN_ROWS = 20
MAX_HEADER_SCAN_COLUMNS = 200
_COST_QUANTUM = Decimal("0.0001")
_MAX_COST_RMB = Decimal("9999999999.9999")
_PLATFORM_SKU_SPLIT = re.compile(r"[\r\n,，;；]+")
_REQUIRED_FIELDS = ("platform_sku", "company_sku", "product_name")
_HEADER_ALIASES: Mapping[str, frozenset[str]] = {
    "platform_sku": frozenset(
        {
            "平台sku",
            "平台商品sku",
            "takealotsku",
            "takealot平台sku",
        }
    ),
    "company_sku": frozenset(
        {
            "公司sku",
            "公司自编sku",
            "自编sku",
            "自定义sku",
            "内部sku",
        }
    ),
    "product_name": frozenset(
        {
            "产品",
            "产品名称",
            "商品",
            "商品名称",
            "中文",
            "中文品名",
        }
    ),
    "cost_rmb": frozenset(
        {
            "成本rmb",
            "rmb成本",
            "单件成本rmb",
            "采购成本rmb",
            "产品成本rmb",
            "采购价rmb",
            "人民币成本",
            "成本人民币",
        }
    ),
    "source_store": frozenset({"店铺", "店铺代码", "storecode"}),
}


class ProductMasterError(ValueError):
    """Base error with a safe, user-facing import explanation."""


class ProductMasterInputError(ProductMasterError):
    """Raised when the workbook cannot be interpreted without guessing."""


class ProductMasterConflictError(ProductMasterError):
    """Raised when an import would silently reassign an existing platform SKU."""


@dataclass(frozen=True)
class ProductMasterSourceRow:
    source_row_number: int
    source_store: str | None
    platform_sku: str
    normalized_platform_sku: str
    company_sku: str
    normalized_company_sku: str
    product_name: str
    cost_rmb: Decimal | None


@dataclass(frozen=True)
class ParsedProductMaster:
    source_file_name: str
    source_sha256: str
    source_modified_at: datetime
    sheet_name: str
    header_row_number: int
    effective_date: date
    source_row_count: int
    rows: tuple[ProductMasterSourceRow, ...]
    warnings: tuple[str, ...]
    header_labels: Mapping[str, str]


@dataclass(frozen=True)
class ProductMasterPlanRow:
    source: ProductMasterSourceRow
    product_action: str
    mapping_action: str
    cost_action: str
    previous_company_sku: str | None
    offer_match_count: int
    matched_store_codes: tuple[str, ...]
    resolved_store_code: str | None
    resolved_offer_id: str | None
    resolved_productline_id: str | None
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class ProductMasterImportPlan:
    parsed: ParsedProductMaster
    rows: tuple[ProductMasterPlanRow, ...]
    existing_batch_id: int | None
    product_actions: Mapping[str, str]
    cost_actions: Mapping[str, str]
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class ProductMasterImportReport:
    mode: str
    batch_id: int | None
    source_file_name: str
    source_sha256: str
    sheet_name: str
    header_row_number: int
    effective_date: date
    source_row_count: int
    normalized_row_count: int
    product_count: int
    mapping_count: int
    cost_count: int
    matched_offer_count: int
    unmatched_offer_count: int
    multiple_offer_match_count: int
    product_actions: Mapping[str, int]
    mapping_actions: Mapping[str, int]
    cost_actions: Mapping[str, int]
    reassignments: tuple[Mapping[str, str], ...]
    warnings: tuple[str, ...]

    def as_dict(self) -> dict[str, Any]:
        return {
            "mode": self.mode,
            "batch_id": self.batch_id,
            "source_file_name": self.source_file_name,
            "source_sha256": self.source_sha256,
            "sheet_name": self.sheet_name,
            "header_row_number": self.header_row_number,
            "effective_date": self.effective_date.isoformat(),
            "source_row_count": self.source_row_count,
            "normalized_row_count": self.normalized_row_count,
            "product_count": self.product_count,
            "mapping_count": self.mapping_count,
            "cost_count": self.cost_count,
            "offer_matches": {
                "matched": self.matched_offer_count,
                "unmatched": self.unmatched_offer_count,
                "multiple": self.multiple_offer_match_count,
            },
            "actions": {
                "products": dict(self.product_actions),
                "mappings": dict(self.mapping_actions),
                "costs": dict(self.cost_actions),
            },
            "reassignments": [dict(item) for item in self.reassignments],
            "warnings": list(self.warnings),
        }


@dataclass(frozen=True)
class ProductMasterLink:
    """Resolved company identity and latest applicable RMB cost for one platform SKU."""

    platform_sku: str
    company_sku: str
    product_name: str
    cost_rmb: Decimal | None
    cost_effective_date: date | None


@dataclass(frozen=True)
class _HeaderSelection:
    sheet_name: str
    row_number: int
    columns: Mapping[str, int]
    labels: Mapping[str, str]


@dataclass(frozen=True)
class _OfferMatch:
    store_code: str
    offer_id: str
    productline_id: str | None


class ProductMasterImportService:
    """Preview and transactionally import one downloaded product-master workbook."""

    def __init__(self, engine: Engine) -> None:
        self._engine = engine

    def preview(
        self,
        workbook_path: Path,
        *,
        effective_date: date,
        sheet_name: str | None = None,
    ) -> ProductMasterImportReport:
        parsed = parse_product_master_workbook(
            workbook_path,
            effective_date=effective_date,
            sheet_name=sheet_name,
        )
        with Session(self._engine) as session:
            plan = _build_import_plan(session, parsed)
        mode = "already_imported" if plan.existing_batch_id is not None else "preview"
        return _report_from_plan(plan, mode=mode, batch_id=plan.existing_batch_id)

    def import_workbook(
        self,
        workbook_path: Path,
        *,
        effective_date: date,
        imported_by: str,
        sheet_name: str | None = None,
        allow_reassignments: bool = False,
    ) -> ProductMasterImportReport:
        parsed = parse_product_master_workbook(
            workbook_path,
            effective_date=effective_date,
            sheet_name=sheet_name,
        )
        actor = _clean_required_text(imported_by, field_name="导入人", max_length=100)
        with Session(self._engine, expire_on_commit=False) as session, session.begin():
            plan = _build_import_plan(session, parsed)
            if plan.existing_batch_id is not None:
                return _report_from_plan(
                    plan,
                    mode="already_imported",
                    batch_id=plan.existing_batch_id,
                )
            reassignments = [row for row in plan.rows if row.mapping_action == "reassigned"]
            if reassignments and not allow_reassignments:
                samples = "；".join(
                    f"{row.source.platform_sku}: {row.previous_company_sku} -> "
                    f"{row.source.company_sku}"
                    for row in reassignments[:5]
                )
                raise ProductMasterConflictError(
                    "发现平台 SKU 改绑，默认拒绝静默覆盖。请核对后使用 "
                    f"--allow-reassignments 显式确认：{samples}"
                )
            return _apply_import_plan(session, plan, imported_by=actor)


def load_product_master_links(
    session: Session,
    *,
    platform_skus: Sequence[str] | None = None,
    as_of_date: date | None = None,
) -> Mapping[str, ProductMasterLink]:
    """Return links keyed by normalized platform SKU, using cost effective-date history."""
    statement = select(PlatformSkuMapping, CompanyProduct).join(
        CompanyProduct,
        CompanyProduct.id == PlatformSkuMapping.company_product_id,
    )
    if platform_skus is not None:
        keys = sorted({_normalize_sku(value) for value in platform_skus if _display_text(value)})
        if not keys:
            return {}
        statement = statement.where(PlatformSkuMapping.normalized_platform_sku.in_(keys))
    mapping_pairs = list(session.execute(statement).all())
    product_ids = sorted({product.id for _, product in mapping_pairs})
    latest_costs: dict[int, CompanyProductCost] = {}
    if product_ids:
        cost_statement = (
            select(CompanyProductCost)
            .where(CompanyProductCost.company_product_id.in_(product_ids))
            .order_by(
                CompanyProductCost.company_product_id,
                CompanyProductCost.effective_date.desc(),
                CompanyProductCost.recorded_at.desc(),
                CompanyProductCost.id.desc(),
            )
        )
        if as_of_date is not None:
            cost_statement = cost_statement.where(
                CompanyProductCost.effective_date <= as_of_date
            )
        for cost in session.scalars(cost_statement):
            latest_costs.setdefault(cost.company_product_id, cost)
    return {
        mapping.normalized_platform_sku: ProductMasterLink(
            platform_sku=mapping.platform_sku,
            company_sku=product.company_sku,
            product_name=product.product_name,
            cost_rmb=(
                latest_costs[product.id].cost_rmb if product.id in latest_costs else None
            ),
            cost_effective_date=(
                latest_costs[product.id].effective_date
                if product.id in latest_costs
                else None
            ),
        )
        for mapping, product in mapping_pairs
    }


def normalize_product_sku(value: Any) -> str:
    """Return the shared exact-match key used by platform and company SKUs."""
    return _normalize_sku(value)


def product_master_link_fields(
    link: ProductMasterLink | None,
) -> dict[str, Any]:
    """Project one product-master link into JSON-safe, stable UI fields."""
    return {
        "company_sku": link.company_sku if link is not None else None,
        "company_product_name": link.product_name if link is not None else None,
        "cost_rmb": float(link.cost_rmb) if link is not None and link.cost_rmb is not None else None,
        "cost_effective_date": (
            link.cost_effective_date.isoformat()
            if link is not None and link.cost_effective_date is not None
            else None
        ),
    }


def enrich_product_master_records(
    session: Session,
    records: Sequence[Mapping[str, Any]],
    *,
    sku_field: str = "sku",
    as_of_date: date | None = None,
) -> list[dict[str, Any]]:
    """Attach company identity/cost to records without changing their ordering."""
    copied = [dict(record) for record in records]
    links = load_product_master_links(
        session,
        platform_skus=[str(record.get(sku_field) or "") for record in copied],
        as_of_date=as_of_date,
    )
    for record in copied:
        link = links.get(_normalize_sku(record.get(sku_field)))
        record.update(product_master_link_fields(link))
    return copied


def load_company_inventory_for_plid(
    engine: Engine,
    *,
    plid: str,
    store_codes: set[str],
) -> dict[str, Any]:
    """Resolve all mapped company-SKU stock for one own PLID from local snapshots.

    W8 is one shared overseas warehouse, so its newest successful snapshot is read
    once. Platform warehouse stages remain separated by authorized store and Offer;
    the function deliberately does not manufacture a cross-stage grand total.
    """
    normalized_plid = str(plid or "").strip()
    normalized_store_codes = {
        str(code).strip() for code in store_codes if str(code).strip()
    }
    if not normalized_plid or not normalized_store_codes:
        return _empty_company_inventory_payload(normalized_store_codes)

    with Session(engine) as session:
        store_names = {
            row.code: row.display_name
            for row in session.scalars(
                select(ErpStore).where(ErpStore.code.in_(normalized_store_codes))
            )
        }

    plid_offers: list[tuple[str, OfferCurrent]] = []
    for store_code in sorted(normalized_store_codes):
        with store_scope(store_code), Session(engine) as session:
            offers = list(
                session.scalars(
                    select(OfferCurrent).where(
                        OfferCurrent.productline_id == normalized_plid
                    )
                )
            )
        plid_offers.extend((store_code, offer) for offer in offers)

    with Session(engine) as session:
        seed_links = load_product_master_links(
            session,
            platform_skus=[str(offer.sku or "") for _, offer in plid_offers],
        )
        company_keys = {
            _normalize_sku(link.company_sku) for link in seed_links.values()
        }
        if not company_keys:
            payload = _empty_company_inventory_payload(normalized_store_codes)
            payload["message"] = "该自有链接的平台 SKU 尚未关联公司 SKU。"
            return payload
        mapping_pairs = list(
            session.execute(
                select(PlatformSkuMapping, CompanyProduct)
                .join(
                    CompanyProduct,
                    CompanyProduct.id == PlatformSkuMapping.company_product_id,
                )
                .where(CompanyProduct.normalized_company_sku.in_(company_keys))
            ).all()
        )
        all_platform_skus = [mapping.platform_sku for mapping, _ in mapping_pairs]
        all_links = load_product_master_links(
            session,
            platform_skus=all_platform_skus,
        )

    platform_keys = {
        mapping.normalized_platform_sku for mapping, _ in mapping_pairs
    }
    platform_offers: list[dict[str, Any]] = []
    for store_code in sorted(normalized_store_codes):
        with store_scope(store_code), Session(engine) as session:
            offers = list(
                session.scalars(
                    select(OfferCurrent).where(OfferCurrent.sku.is_not(None))
                )
            )
        for offer in offers:
            platform_key = _normalize_sku(offer.sku)
            if platform_key not in platform_keys:
                continue
            link = all_links.get(platform_key)
            if link is None or _normalize_sku(link.company_sku) not in company_keys:
                continue
            platform_offers.append(
                {
                    "store_code": store_code,
                    "store_name": store_names.get(store_code, store_code),
                    "offer_id": offer.offer_id,
                    "plid": offer.productline_id,
                    "platform_sku": offer.sku,
                    "company_sku": link.company_sku,
                    "status": offer.status,
                    "platform_available_stock": offer.takealot_available_stock,
                    "platform_stock_on_way": offer.takealot_stock_on_way,
                    "platform_stock_in_receiving": offer.takealot_stock_in_receiving,
                    "captured_at": _iso_datetime(offer.captured_at),
                }
            )

    w8_snapshot = _latest_shared_w8_snapshot(engine)
    company_groups: dict[str, dict[str, Any]] = {}
    for mapping, product in mapping_pairs:
        company_key = product.normalized_company_sku
        link = all_links.get(mapping.normalized_platform_sku)
        group = company_groups.setdefault(
            company_key,
            {
                **product_master_link_fields(link),
                "company_sku": product.company_sku,
                "company_product_name": product.product_name,
                "mapped_platform_skus": [],
            },
        )
        group["mapped_platform_skus"].append(mapping.platform_sku)

    items: list[dict[str, Any]] = []
    for company_key, group in sorted(
        company_groups.items(),
        key=lambda item: str(item[1]["company_sku"]).casefold(),
    ):
        company_platform_offers = [
            row
            for row in platform_offers
            if _normalize_sku(row.get("company_sku")) == company_key
        ]
        group["mapped_platform_skus"] = sorted(
            set(group["mapped_platform_skus"]),
            key=str.casefold,
        )
        items.append(
            {
                **group,
                "overseas_warehouse": _company_w8_inventory(
                    w8_snapshot,
                    company_key=company_key,
                ),
                "platform_warehouse": {
                    "offer_count": len(company_platform_offers),
                    "stages": {
                        "available": _covered_sum(
                            company_platform_offers,
                            "platform_available_stock",
                        ),
                        "on_way": _covered_sum(
                            company_platform_offers,
                            "platform_stock_on_way",
                        ),
                        "in_receiving": _covered_sum(
                            company_platform_offers,
                            "platform_stock_in_receiving",
                        ),
                    },
                    "offers": sorted(
                        company_platform_offers,
                        key=lambda row: (
                            str(row["store_name"]).casefold(),
                            str(row["platform_sku"]).casefold(),
                            str(row["offer_id"]),
                        ),
                    ),
                    "latest_captured_at": max(
                        (
                            str(row["captured_at"])
                            for row in company_platform_offers
                            if row.get("captured_at")
                        ),
                        default=None,
                    ),
                },
            }
        )

    return {
        "items": items,
        "store_codes": sorted(normalized_store_codes),
        "company_sku_count": len(items),
        "w8_shared_once": True,
        "stage_totals_are_additive": False,
        "message": (
            "海外仓为共享 W8 快照，只展示一次；平台仓按授权店铺和 Offer 展开。"
            "各库存阶段可能重叠，不生成跨阶段总和。"
        ),
    }


def _empty_company_inventory_payload(store_codes: set[str]) -> dict[str, Any]:
    return {
        "items": [],
        "store_codes": sorted(store_codes),
        "company_sku_count": 0,
        "w8_shared_once": True,
        "stage_totals_are_additive": False,
        "message": "当前范围没有可展示的公司 SKU 库存。",
    }


def _latest_shared_w8_snapshot(engine: Engine) -> Mapping[str, Any] | None:
    table = LogisticsProviderSnapshot.__table__
    with engine.connect() as connection:
        row = connection.execute(
            select(table.c.fetched_at, table.c.payload)
            .where(table.c.provider == "w8")
            .order_by(table.c.fetched_at.desc())
            .limit(1)
        ).mappings().first()
    return dict(row) if row is not None else None


def _company_w8_inventory(
    snapshot: Mapping[str, Any] | None,
    *,
    company_key: str,
) -> dict[str, Any]:
    if snapshot is None:
        return {
            "available": False,
            "matched": False,
            "snapshot_at": None,
            "warehouse": None,
            "record_count": 0,
            "message": "尚无长睿海外仓本地快照。",
            "stages": {},
        }
    payload = snapshot.get("payload")
    payload = payload if isinstance(payload, Mapping) else {}
    inventory_items = payload.get("inventory_items")
    inventory_rows = inventory_items if isinstance(inventory_items, list) else []
    detail_available = bool(payload.get("inventory_detail_available")) and isinstance(
        inventory_items,
        list,
    )
    if not detail_available:
        return {
            "available": False,
            "matched": False,
            "snapshot_at": _iso_datetime(snapshot.get("fetched_at")),
            "warehouse": payload.get("warehouse"),
            "record_count": 0,
            "message": "现有长睿快照早于 SKU 明细功能，请等待下一次同步。",
            "stages": {},
        }
    matches = [
        dict(item)
        for item in inventory_rows
        if isinstance(item, Mapping)
        and _normalize_sku(item.get("company_sku")) == company_key
    ]
    fields = (
        "stock_total",
        "usable_stock",
        "locked_stock",
        "outbound_allocated",
        "transit_stock",
        "defective_stock",
    )
    return {
        "available": True,
        "matched": bool(matches),
        "snapshot_at": _iso_datetime(snapshot.get("fetched_at")),
        "warehouse": payload.get("warehouse"),
        "record_count": len(matches),
        "message": (
            "按公司 SKU 精确匹配长睿库存。"
            if matches
            else "长睿最新快照中没有该公司 SKU，按当前快照记为 0。"
        ),
        "stages": {
            field: _covered_sum(matches, field) if matches else {"value": 0, "coverage": 0}
            for field in fields
        },
        "records": matches,
    }


def _covered_sum(rows: Sequence[Mapping[str, Any]], field: str) -> dict[str, int]:
    values = [row.get(field) for row in rows]
    covered = [int(value) for value in values if value is not None]
    return {"value": sum(covered), "coverage": len(covered)}


def _iso_datetime(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.isoformat()
    text = _display_text(value)
    return text or None


def parse_product_master_workbook(
    workbook_path: Path,
    *,
    effective_date: date,
    sheet_name: str | None = None,
) -> ParsedProductMaster:
    """Parse one XLSX without accepting ambiguous cost or identity columns."""
    source = workbook_path.expanduser().resolve()
    if not source.is_file():
        raise ProductMasterInputError(f"找不到导入文件：{source}")
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ProductMasterInputError("只支持 .xlsx 或 .xlsm；旧版 .xls 请先另存为 .xlsx")
    size = source.stat().st_size
    if size <= 0:
        raise ProductMasterInputError("导入文件为空")
    if size > MAX_WORKBOOK_BYTES:
        raise ProductMasterInputError(
            f"导入文件超过 {MAX_WORKBOOK_BYTES // (1024 * 1024)} MB 安全上限"
        )

    workbook = load_workbook(
        source,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        selection = _select_sheet_and_headers(workbook, sheet_name=sheet_name)
        worksheet = workbook[selection.sheet_name]
        rows, source_row_count, warnings = _parse_source_rows(
            worksheet,
            selection,
        )
    finally:
        workbook.close()

    source_modified_at = datetime.fromtimestamp(source.stat().st_mtime, tz=UTC).replace(
        tzinfo=None
    )
    return ParsedProductMaster(
        source_file_name=source.name,
        source_sha256=_sha256_file(source),
        source_modified_at=source_modified_at,
        sheet_name=selection.sheet_name,
        header_row_number=selection.row_number,
        effective_date=effective_date,
        source_row_count=source_row_count,
        rows=rows,
        warnings=warnings,
        header_labels=selection.labels,
    )


def _select_sheet_and_headers(workbook: Any, *, sheet_name: str | None) -> _HeaderSelection:
    if sheet_name is not None:
        clean_sheet = sheet_name.strip()
        if clean_sheet not in workbook.sheetnames:
            raise ProductMasterInputError(
                f"找不到工作表“{clean_sheet}”；现有：{'、'.join(workbook.sheetnames)}"
            )
        selection = _locate_headers(workbook[clean_sheet])
        if selection is None:
            raise ProductMasterInputError(
                f"工作表“{clean_sheet}”前 {MAX_HEADER_SCAN_ROWS} 行找不到平台 SKU、公司 SKU、产品三列"
            )
        return selection

    candidates = [
        selection
        for name in workbook.sheetnames
        if (selection := _locate_headers(workbook[name])) is not None
    ]
    if not candidates:
        raise ProductMasterInputError(
            "所有工作表都找不到明确的平台 SKU、公司 SKU、产品三列；"
            "不会用普通“SKU”或“采购费用”猜列"
        )
    if len(candidates) > 1:
        names = "、".join(candidate.sheet_name for candidate in candidates)
        raise ProductMasterInputError(
            f"多个工作表都符合导入结构（{names}）；请用 --sheet 明确指定"
        )
    return candidates[0]


def _locate_headers(worksheet: Any) -> _HeaderSelection | None:
    if not worksheet.max_row or not worksheet.max_column:
        try:
            worksheet.calculate_dimension(force=True)
        except (TypeError, ValueError):
            return None
    max_row = min(int(worksheet.max_row or 0), MAX_HEADER_SCAN_ROWS)
    max_column = min(int(worksheet.max_column or 0), MAX_HEADER_SCAN_COLUMNS)
    if max_row < 1 or max_column < 1:
        return None
    for row_number, values in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=max_row,
            max_col=max_column,
            values_only=True,
        ),
        start=1,
    ):
        matches: dict[str, list[tuple[int, str]]] = defaultdict(list)
        for column_number, value in enumerate(values, start=1):
            label = _display_text(value)
            canonical = _canonical_header(label)
            if not canonical:
                continue
            for field, aliases in _HEADER_ALIASES.items():
                if canonical in aliases:
                    matches[field].append((column_number, label))
        if not all(matches.get(field) for field in _REQUIRED_FIELDS):
            continue
        duplicate_fields = [field for field, entries in matches.items() if len(entries) > 1]
        if duplicate_fields:
            raise ProductMasterInputError(
                f"工作表“{worksheet.title}”第 {row_number} 行存在重复语义列："
                f"{'、'.join(duplicate_fields)}"
            )
        return _HeaderSelection(
            sheet_name=str(worksheet.title),
            row_number=row_number,
            columns={field: entries[0][0] for field, entries in matches.items()},
            labels={field: entries[0][1] for field, entries in matches.items()},
        )
    return None


def _parse_source_rows(
    worksheet: Any,
    selection: _HeaderSelection,
) -> tuple[tuple[ProductMasterSourceRow, ...], int, tuple[str, ...]]:
    max_column = max(selection.columns.values())
    errors: list[str] = []
    warnings: list[str] = []
    source_rows = 0
    consecutive_blank_rows = 0
    parsed_rows: list[ProductMasterSourceRow] = []

    for source_row_number, values in enumerate(
        worksheet.iter_rows(
            min_row=selection.row_number + 1,
            max_col=max_column,
            values_only=True,
        ),
        start=selection.row_number + 1,
    ):
        raw = {
            field: values[column_number - 1]
            for field, column_number in selection.columns.items()
        }
        if all(_is_blank(value) for value in raw.values()):
            consecutive_blank_rows += 1
            if source_rows > 0 and consecutive_blank_rows >= 1000:
                break
            continue
        consecutive_blank_rows = 0
        source_rows += 1
        if source_rows > MAX_SOURCE_ROWS:
            raise ProductMasterInputError(
                f"非空业务行超过 {MAX_SOURCE_ROWS} 行安全上限"
            )
        try:
            platform_skus = _split_platform_skus(raw.get("platform_sku"))
            company_sku = _clean_required_text(
                raw.get("company_sku"),
                field_name="公司 SKU",
                max_length=255,
            )
            product_name = _clean_required_text(
                raw.get("product_name"),
                field_name="产品名称",
                max_length=2000,
            )
            source_store = _clean_optional_text(raw.get("source_store"), max_length=255)
            cost_rmb = _parse_cost(raw.get("cost_rmb"))
        except ProductMasterInputError as exc:
            errors.append(f"第 {source_row_number} 行：{exc}")
            continue

        for platform_sku in platform_skus:
            parsed_rows.append(
                ProductMasterSourceRow(
                    source_row_number=source_row_number,
                    source_store=source_store,
                    platform_sku=platform_sku,
                    normalized_platform_sku=_normalize_sku(platform_sku),
                    company_sku=company_sku,
                    normalized_company_sku=_normalize_sku(company_sku),
                    product_name=product_name,
                    cost_rmb=cost_rmb,
                )
            )

    if errors:
        rendered = "；".join(errors[:20])
        suffix = f"；另有 {len(errors) - 20} 条" if len(errors) > 20 else ""
        raise ProductMasterInputError(f"表格预检失败：{rendered}{suffix}")
    if not parsed_rows:
        raise ProductMasterInputError("没有可导入的数据行")

    deduplicated: dict[str, ProductMasterSourceRow] = {}
    duplicate_rows: list[tuple[int, int, str]] = []
    for row in parsed_rows:
        existing = deduplicated.get(row.normalized_platform_sku)
        if existing is None:
            deduplicated[row.normalized_platform_sku] = row
            continue
        if existing.normalized_company_sku != row.normalized_company_sku:
            errors.append(
                f"平台 SKU {row.platform_sku} 在第 {existing.source_row_number}、"
                f"{row.source_row_number} 行对应不同公司 SKU"
            )
            continue
        if _normalize_name(existing.product_name) != _normalize_name(row.product_name):
            errors.append(
                f"平台 SKU {row.platform_sku} 在第 {existing.source_row_number}、"
                f"{row.source_row_number} 行对应不同产品名称"
            )
            continue
        if (
            existing.cost_rmb is not None
            and row.cost_rmb is not None
            and existing.cost_rmb != row.cost_rmb
        ):
            errors.append(
                f"平台 SKU {row.platform_sku} 在第 {existing.source_row_number}、"
                f"{row.source_row_number} 行成本冲突"
            )
            continue
        if existing.cost_rmb is None and row.cost_rmb is not None:
            deduplicated[row.normalized_platform_sku] = row
        duplicate_rows.append(
            (existing.source_row_number, row.source_row_number, row.platform_sku)
        )

    rows = tuple(deduplicated.values())
    _validate_company_consistency(rows, errors)
    if errors:
        rendered = "；".join(errors[:20])
        suffix = f"；另有 {len(errors) - 20} 条" if len(errors) > 20 else ""
        raise ProductMasterInputError(f"表格预检失败：{rendered}{suffix}")

    if duplicate_rows:
        warnings.append(
            f"发现 {len(duplicate_rows)} 条完全重复的平台 SKU 行，已按平台 SKU 去重"
        )
    if "cost_rmb" not in selection.columns:
        warnings.append("未找到明确的 RMB 单件成本列；本次只导入产品与 SKU 映射")
    else:
        missing_cost_companies = {
            row.normalized_company_sku for row in rows if row.cost_rmb is None
        }
        if missing_cost_companies:
            warnings.append(
                f"{len(missing_cost_companies)} 个公司 SKU 的 RMB 单件成本为空；"
                "映射仍可导入，成本保持缺失"
            )
    return rows, source_rows, tuple(warnings)


def _validate_company_consistency(
    rows: Sequence[ProductMasterSourceRow],
    errors: list[str],
) -> None:
    names: dict[str, dict[str, list[int]]] = defaultdict(lambda: defaultdict(list))
    costs: dict[str, dict[Decimal, list[int]]] = defaultdict(lambda: defaultdict(list))
    display_skus: dict[str, str] = {}
    for row in rows:
        key = row.normalized_company_sku
        display_skus.setdefault(key, row.company_sku)
        names[key][_normalize_name(row.product_name)].append(row.source_row_number)
        if row.cost_rmb is not None:
            costs[key][row.cost_rmb].append(row.source_row_number)
    for key, name_variants in names.items():
        if len(name_variants) > 1:
            row_numbers = sorted(
                {number for values in name_variants.values() for number in values}
            )
            errors.append(
                f"公司 SKU {display_skus[key]} 在行 {row_numbers} 使用了不同产品名称"
            )
    for key, cost_variants in costs.items():
        if len(cost_variants) > 1:
            rendered = "、".join(str(value) for value in sorted(cost_variants))
            errors.append(
                f"公司 SKU {display_skus[key]} 在同一批次出现多个 RMB 成本：{rendered}"
            )


def _build_import_plan(
    session: Session,
    parsed: ParsedProductMaster,
) -> ProductMasterImportPlan:
    company_keys = sorted({row.normalized_company_sku for row in parsed.rows})
    platform_keys = sorted({row.normalized_platform_sku for row in parsed.rows})
    existing_batch = session.scalar(
        select(ProductMasterImportBatch).where(
            ProductMasterImportBatch.source_sha256 == parsed.source_sha256,
            ProductMasterImportBatch.sheet_name == parsed.sheet_name,
            ProductMasterImportBatch.effective_date == parsed.effective_date,
        )
    )

    products = list(
        session.scalars(
            select(CompanyProduct).where(
                CompanyProduct.normalized_company_sku.in_(company_keys)
            )
        )
    )
    products_by_key = {product.normalized_company_sku: product for product in products}
    mapping_pairs = list(
        session.execute(
            select(PlatformSkuMapping, CompanyProduct)
            .join(CompanyProduct, CompanyProduct.id == PlatformSkuMapping.company_product_id)
            .where(PlatformSkuMapping.normalized_platform_sku.in_(platform_keys))
        ).all()
    )
    mappings_by_key = {
        mapping.normalized_platform_sku: (mapping, product)
        for mapping, product in mapping_pairs
    }

    product_ids = [product.id for product in products]
    cost_versions_by_product: dict[int, list[CompanyProductCost]] = defaultdict(list)
    if product_ids:
        cost_versions = list(
            session.scalars(
                select(CompanyProductCost)
                .where(CompanyProductCost.company_product_id.in_(product_ids))
                .order_by(
                    CompanyProductCost.effective_date.desc(),
                    CompanyProductCost.recorded_at.desc(),
                    CompanyProductCost.id.desc(),
                )
            )
        )
        for version in cost_versions:
            cost_versions_by_product[version.company_product_id].append(version)

    source_by_company: dict[str, ProductMasterSourceRow] = {}
    for source_row in parsed.rows:
        current = source_by_company.get(source_row.normalized_company_sku)
        if current is None or (
            current.cost_rmb is None and source_row.cost_rmb is not None
        ):
            source_by_company[source_row.normalized_company_sku] = source_row

    product_actions: dict[str, str] = {}
    cost_actions: dict[str, str] = {}
    for key, source in source_by_company.items():
        existing_product = products_by_key.get(key)
        if existing_product is None:
            product_actions[key] = "created"
        elif _normalize_name(existing_product.product_name) == _normalize_name(
            source.product_name
        ):
            product_actions[key] = "unchanged"
        else:
            product_actions[key] = "renamed"
        cost_actions[key] = _cost_action(
            existing_product,
            cost_versions_by_product.get(existing_product.id, [])
            if existing_product is not None
            else [],
            input_cost=source.cost_rmb,
            effective_date=parsed.effective_date,
        )

    offers_by_platform = _load_offer_matches(session, parsed.rows)
    plan_rows: list[ProductMasterPlanRow] = []
    unmatched_count = 0
    multiple_count = 0
    for source in parsed.rows:
        existing_mapping = mappings_by_key.get(source.normalized_platform_sku)
        previous_company_sku: str | None = None
        if existing_mapping is None:
            mapping_action = "created"
        else:
            _, mapped_product = existing_mapping
            if mapped_product.normalized_company_sku == source.normalized_company_sku:
                mapping_action = "unchanged"
            else:
                mapping_action = "reassigned"
                previous_company_sku = mapped_product.company_sku

        matches = offers_by_platform.get(source.normalized_platform_sku, ())
        matched_store_codes = tuple(sorted({match.store_code for match in matches}))
        row_warnings: list[str] = []
        resolved_store_code: str | None = None
        resolved_offer_id: str | None = None
        resolved_productline_id: str | None = None
        if len(matches) == 1:
            resolved_store_code = matches[0].store_code
            resolved_offer_id = matches[0].offer_id
            resolved_productline_id = matches[0].productline_id
        elif not matches:
            unmatched_count += 1
            row_warnings.append("当前 offer_current 尚未找到该平台 SKU")
        else:
            multiple_count += 1
            row_warnings.append("当前 offer_current 存在多个同平台 SKU，未固化单一 Offer")

        plan_rows.append(
            ProductMasterPlanRow(
                source=source,
                product_action=product_actions[source.normalized_company_sku],
                mapping_action=mapping_action,
                cost_action=cost_actions[source.normalized_company_sku],
                previous_company_sku=previous_company_sku,
                offer_match_count=len(matches),
                matched_store_codes=matched_store_codes,
                resolved_store_code=resolved_store_code,
                resolved_offer_id=resolved_offer_id,
                resolved_productline_id=resolved_productline_id,
                warnings=tuple(row_warnings),
            )
        )

    warnings = list(parsed.warnings)
    if unmatched_count:
        warnings.append(
            f"{unmatched_count} 个平台 SKU 尚未出现在当前 Offer；映射可保留，匹配证据为空"
        )
    if multiple_count:
        warnings.append(
            f"{multiple_count} 个平台 SKU 命中多个当前 Offer；映射保留但不猜单一店铺/Offer"
        )
    reassignment_count = sum(row.mapping_action == "reassigned" for row in plan_rows)
    if reassignment_count:
        warnings.append(
            f"{reassignment_count} 个平台 SKU 将改绑公司 SKU；正式导入需显式确认"
        )
    return ProductMasterImportPlan(
        parsed=parsed,
        rows=tuple(plan_rows),
        existing_batch_id=existing_batch.id if existing_batch is not None else None,
        product_actions=product_actions,
        cost_actions=cost_actions,
        warnings=tuple(warnings),
    )


def _load_offer_matches(
    session: Session,
    rows: Sequence[ProductMasterSourceRow],
) -> Mapping[str, tuple[_OfferMatch, ...]]:
    table = OfferCurrent.__table__
    display_values = sorted({row.platform_sku for row in rows})
    grouped: dict[str, list[_OfferMatch]] = defaultdict(list)
    for values in _chunks(display_values, 500):
        matches = session.connection().execute(
            select(
                table.c.store_code,
                table.c.sku,
                table.c.offer_id,
                table.c.productline_id,
            ).where(table.c.sku.in_(values))
        )
        for match in matches.mappings():
            platform_sku = _normalize_sku(match["sku"])
            grouped[platform_sku].append(
                _OfferMatch(
                    store_code=str(match["store_code"]),
                    offer_id=str(match["offer_id"]),
                    productline_id=(
                        str(match["productline_id"])
                        if match["productline_id"] is not None
                        else None
                    ),
                )
            )
    return {
        key: tuple(sorted(values, key=lambda item: (item.store_code, item.offer_id)))
        for key, values in grouped.items()
    }


def _cost_action(
    existing_product: CompanyProduct | None,
    versions: Sequence[CompanyProductCost],
    *,
    input_cost: Decimal | None,
    effective_date: date,
) -> str:
    if input_cost is None:
        return "absent"
    if existing_product is None or not versions:
        return "created"
    same_date = next(
        (version for version in versions if version.effective_date == effective_date),
        None,
    )
    if same_date is not None:
        return "unchanged" if same_date.cost_rmb == input_cost else "revised"
    return "unchanged" if versions[0].cost_rmb == input_cost else "created"


def _apply_import_plan(
    session: Session,
    plan: ProductMasterImportPlan,
    *,
    imported_by: str,
) -> ProductMasterImportReport:
    parsed = plan.parsed
    now = datetime.utcnow()
    cost_count = len(
        {
            row.normalized_company_sku
            for row in parsed.rows
            if row.cost_rmb is not None
        }
    )
    batch = ProductMasterImportBatch(
        source_file_name=parsed.source_file_name,
        source_sha256=parsed.source_sha256,
        source_modified_at=parsed.source_modified_at,
        sheet_name=parsed.sheet_name,
        header_row_number=parsed.header_row_number,
        effective_date=parsed.effective_date,
        imported_by=imported_by,
        imported_at=now,
        status="completed",
        source_row_count=parsed.source_row_count,
        normalized_row_count=len(parsed.rows),
        product_count=len(plan.product_actions),
        mapping_count=len(plan.rows),
        cost_count=cost_count,
        warning_count=len(plan.warnings),
        summary=None,
    )
    session.add(batch)
    session.flush()

    company_keys = sorted(plan.product_actions)
    products = list(
        session.scalars(
            select(CompanyProduct).where(
                CompanyProduct.normalized_company_sku.in_(company_keys)
            )
        )
    )
    products_by_key = {product.normalized_company_sku: product for product in products}
    source_by_company: dict[str, ProductMasterSourceRow] = {}
    for row in parsed.rows:
        current = source_by_company.get(row.normalized_company_sku)
        if current is None or (current.cost_rmb is None and row.cost_rmb is not None):
            source_by_company[row.normalized_company_sku] = row

    for key, source in source_by_company.items():
        product = products_by_key.get(key)
        if product is None:
            product = CompanyProduct(
                company_sku=source.company_sku,
                normalized_company_sku=key,
                product_name=source.product_name,
                first_source_batch_id=batch.id,
                last_source_batch_id=batch.id,
                created_at=now,
                updated_at=now,
            )
            session.add(product)
            products_by_key[key] = product
        else:
            product.company_sku = source.company_sku
            product.product_name = source.product_name
            product.last_source_batch_id = batch.id
            product.updated_at = now
    session.flush()

    platform_keys = sorted({row.source.normalized_platform_sku for row in plan.rows})
    mappings = list(
        session.scalars(
            select(PlatformSkuMapping).where(
                PlatformSkuMapping.normalized_platform_sku.in_(platform_keys)
            )
        )
    )
    mappings_by_key = {mapping.normalized_platform_sku: mapping for mapping in mappings}
    for plan_row in plan.rows:
        source = plan_row.source
        product = products_by_key[source.normalized_company_sku]
        mapping = mappings_by_key.get(source.normalized_platform_sku)
        if mapping is None:
            mapping = PlatformSkuMapping(
                platform_sku=source.platform_sku,
                normalized_platform_sku=source.normalized_platform_sku,
                company_product_id=product.id,
                resolved_store_code=plan_row.resolved_store_code,
                resolved_offer_id=plan_row.resolved_offer_id,
                resolved_productline_id=plan_row.resolved_productline_id,
                first_source_batch_id=batch.id,
                last_source_batch_id=batch.id,
                last_source_row_number=source.source_row_number,
                created_at=now,
                updated_at=now,
            )
            session.add(mapping)
            mappings_by_key[source.normalized_platform_sku] = mapping
        else:
            mapping.platform_sku = source.platform_sku
            mapping.company_product_id = product.id
            mapping.resolved_store_code = plan_row.resolved_store_code
            mapping.resolved_offer_id = plan_row.resolved_offer_id
            mapping.resolved_productline_id = plan_row.resolved_productline_id
            mapping.last_source_batch_id = batch.id
            mapping.last_source_row_number = source.source_row_number
            mapping.updated_at = now

    for key, action in plan.cost_actions.items():
        source = source_by_company[key]
        if source.cost_rmb is None or action not in {"created", "revised"}:
            continue
        session.add(
            CompanyProductCost(
                company_product_id=products_by_key[key].id,
                effective_date=parsed.effective_date,
                cost_rmb=source.cost_rmb,
                source_batch_id=batch.id,
                source_row_number=source.source_row_number,
                recorded_at=now,
            )
        )

    for plan_row in plan.rows:
        source = plan_row.source
        session.add(
            ProductMasterImportRow(
                batch_id=batch.id,
                source_row_number=source.source_row_number,
                source_store=source.source_store,
                platform_sku=source.platform_sku,
                normalized_platform_sku=source.normalized_platform_sku,
                company_sku=source.company_sku,
                normalized_company_sku=source.normalized_company_sku,
                product_name=source.product_name,
                cost_rmb=source.cost_rmb,
                matched_store_codes=list(plan_row.matched_store_codes) or None,
                previous_company_sku=plan_row.previous_company_sku,
                product_action=plan_row.product_action,
                mapping_action=plan_row.mapping_action,
                cost_action=plan_row.cost_action,
                warnings=list(plan_row.warnings) or None,
                recorded_at=now,
            )
        )

    report = _report_from_plan(plan, mode="imported", batch_id=batch.id)
    batch.summary = report.as_dict()
    session.flush()
    return report


def _report_from_plan(
    plan: ProductMasterImportPlan,
    *,
    mode: str,
    batch_id: int | None,
) -> ProductMasterImportReport:
    matched = sum(row.offer_match_count == 1 for row in plan.rows)
    unmatched = sum(row.offer_match_count == 0 for row in plan.rows)
    multiple = sum(row.offer_match_count > 1 for row in plan.rows)
    reassignments = tuple(
        {
            "platform_sku": row.source.platform_sku,
            "from_company_sku": row.previous_company_sku or "",
            "to_company_sku": row.source.company_sku,
        }
        for row in plan.rows
        if row.mapping_action == "reassigned"
    )
    cost_count = len(
        {
            row.source.normalized_company_sku
            for row in plan.rows
            if row.source.cost_rmb is not None
        }
    )
    return ProductMasterImportReport(
        mode=mode,
        batch_id=batch_id,
        source_file_name=plan.parsed.source_file_name,
        source_sha256=plan.parsed.source_sha256,
        sheet_name=plan.parsed.sheet_name,
        header_row_number=plan.parsed.header_row_number,
        effective_date=plan.parsed.effective_date,
        source_row_count=plan.parsed.source_row_count,
        normalized_row_count=len(plan.rows),
        product_count=len(plan.product_actions),
        mapping_count=len(plan.rows),
        cost_count=cost_count,
        matched_offer_count=matched,
        unmatched_offer_count=unmatched,
        multiple_offer_match_count=multiple,
        product_actions=dict(Counter(plan.product_actions.values())),
        mapping_actions=dict(Counter(row.mapping_action for row in plan.rows)),
        cost_actions=dict(Counter(plan.cost_actions.values())),
        reassignments=reassignments,
        warnings=plan.warnings,
    )


def _split_platform_skus(value: Any) -> tuple[str, ...]:
    rendered = _display_text(value)
    if not rendered:
        raise ProductMasterInputError("平台 SKU 为空")
    values: list[str] = []
    seen: set[str] = set()
    for part in _PLATFORM_SKU_SPLIT.split(rendered):
        clean = _clean_required_text(part, field_name="平台 SKU", max_length=255)
        normalized = _normalize_sku(clean)
        if normalized not in seen:
            seen.add(normalized)
            values.append(clean)
    if not values:
        raise ProductMasterInputError("平台 SKU 为空")
    return tuple(values)


def _parse_cost(value: Any) -> Decimal | None:
    if _is_blank(value):
        return None
    rendered = _display_text(value)
    clean = re.sub(r"(?i)rmb|cny", "", rendered)
    clean = clean.replace("￥", "").replace("¥", "").replace(",", "").strip()
    try:
        parsed = Decimal(clean)
    except (InvalidOperation, ValueError) as exc:
        raise ProductMasterInputError(f"RMB 单件成本不是有效数字：{rendered}") from exc
    if not parsed.is_finite() or parsed <= 0:
        raise ProductMasterInputError("RMB 单件成本必须大于 0")
    quantized = parsed.quantize(_COST_QUANTUM, rounding=ROUND_HALF_UP)
    if quantized > _MAX_COST_RMB:
        raise ProductMasterInputError("RMB 单件成本超出数据库范围")
    return quantized


def _clean_required_text(value: Any, *, field_name: str, max_length: int) -> str:
    clean = _display_text(value)
    if not clean:
        raise ProductMasterInputError(f"{field_name}为空")
    if len(clean) > max_length:
        raise ProductMasterInputError(f"{field_name}超过 {max_length} 个字符")
    return clean


def _clean_optional_text(value: Any, *, max_length: int) -> str | None:
    clean = _display_text(value)
    if not clean:
        return None
    if len(clean) > max_length:
        raise ProductMasterInputError(f"文本超过 {max_length} 个字符")
    return clean


def _display_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return format(value, "f")
    return re.sub(r"[ \t]+", " ", str(value).strip())


def _canonical_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).casefold()
    return re.sub(r"[\s_\-—:：/\\()\[\]{}]+", "", normalized)


def _normalize_sku(value: Any) -> str:
    return unicodedata.normalize("NFKC", _display_text(value)).strip().casefold()


def _normalize_name(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).casefold()
    return re.sub(r"\s+", " ", normalized).strip()


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _chunks(values: Sequence[str], size: int) -> Iterable[Sequence[str]]:
    for index in range(0, len(values), size):
        yield values[index : index + size]
