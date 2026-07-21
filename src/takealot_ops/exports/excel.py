"""Create a professional OpenPyXL operations workbook."""

from __future__ import annotations

import math
from collections.abc import Mapping
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, cast

import pandas as pd
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.chart import LineChart, Reference  # type: ignore[import-untyped]
from openpyxl.formatting.rule import CellIsRule, FormulaRule  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from takealot_ops.metrics.service import DashboardDataset


SHEET_NAMES = (
    "运营总览",
    "单品分析",
    "商品数据",
    "异常商品",
    "每日汇总",
    "销售明细",
    "流量快照",
    "指标说明",
    "数据质量",
)

_NAVY = "173B68"
_BLUE = "2563EB"
_TEAL = "0F766E"
_PALE_BLUE = "EAF2FF"
_PALE_GREEN = "DCFCE7"
_PALE_RED = "FEE2E2"
_PALE_AMBER = "FEF3C7"
_TEXT = "172033"
_MUTED = "64748B"
_WHITE = "FFFFFF"
_LINE = "DCE4EF"
_HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color=_WHITE)
_BODY_FONT = Font(name="Microsoft YaHei", size=10, color=_TEXT)
_THIN_BOTTOM = Border(bottom=Side(style="thin", color=_LINE))

_PRODUCT_COLUMNS = [
    ("metric_date", "日期"),
    ("offer_id", "Offer ID"),
    ("sku", "SKU"),
    ("product_title_current", "商品名称（当前）"),
    ("selling_price_current", "当前售价（ZAR）"),
    ("ordered_units", "订购件数"),
    ("effective_units", "有效件数"),
    ("ordered_revenue", "订购销售额（ZAR）"),
    ("page_views_30_days", "近30天浏览量"),
    ("page_views_30_day_average", "近30天日均浏览量"),
    ("page_views_window_net_change", "30天浏览量窗口净变化"),
    ("conversion_percentage_30_days", "近30天转化率（%）"),
    ("conversion_percentage_previous_30_days", "前30天转化率（%）"),
    ("conversion_change_points", "转化率变化（百分点）"),
    ("total_stock_current", "当前库存"),
    ("offer_status_current", "商品状态（当前）"),
]

_TRAFFIC_COLUMNS = [
    ("metric_date", "日期"),
    ("offer_id", "Offer ID"),
    ("sku", "SKU"),
    ("product_title_current", "商品名称（当前）"),
    ("page_views_30_days", "近30天浏览量"),
    ("page_views_30_day_average", "近30天日均浏览量"),
    ("page_views_window_net_change", "30天浏览量窗口净变化"),
    ("conversion_percentage_30_days", "近30天转化率（%）"),
    ("conversion_percentage_previous_30_days", "前30天转化率（%）"),
    ("conversion_change_points", "转化率变化（百分点）"),
    ("offer_status_current", "商品状态（当前）"),
]

_OFFER_COLUMNS = [
    ("offer_id", "Offer ID"),
    ("tsin_id", "TSIN ID"),
    ("sku", "SKU"),
    ("barcode", "条码"),
    ("title", "商品名称"),
    ("selling_price", "当前售价（ZAR）"),
    ("rrp", "建议零售价（ZAR）"),
    ("benchmark_price", "基准价（ZAR）"),
    ("status_label", "商品状态"),
    ("page_views_30_days", "近30天浏览量"),
    ("conversion_percentage_30_days", "近30天转化率（%）"),
    ("conversion_percentage_previous_30_days", "前30天转化率（%）"),
    ("quantity_returned_30_days", "近30天退货件数"),
    ("total_wishlist", "累计收藏数"),
    ("wishlist_30_days", "近30天收藏数"),
    ("listing_quality", "Listing Quality 原值"),
    ("discount_percentage", "折扣率（%）"),
    ("total_stock", "当前库存"),
    ("updated_at", "商品更新时间"),
    ("captured_at", "采集时间"),
    ("productline_id", "Product Line ID"),
    ("image_url", "商品图片地址"),
    ("status", "API 商品状态代码"),
]

_ANOMALY_COLUMNS = [
    ("event_date", "日期"),
    ("offer_id", "Offer ID"),
    ("product_title_current", "商品名称（当前）"),
    ("anomaly_label", "异常类型"),
    ("severity_label", "级别"),
    ("explanation_label", "异常说明"),
    ("baseline_daily_units", "基准日均件数"),
    ("detail_ordered_units", "当日订购件数"),
    ("detail_offer_status", "异常记录中的状态"),
    ("sale_statuses", "涉及销售状态"),
    ("created_at", "创建时间"),
]

_QUALITY_COLUMNS = [
    ("event_date", "日期"),
    ("offer_id", "Offer ID"),
    ("product_title_current", "商品名称（当前）"),
    ("event_label", "质量问题"),
    ("severity_label", "级别"),
    ("problem_description", "问题说明"),
    ("sale_statuses", "涉及销售状态"),
    ("created_at", "创建时间"),
    ("event_id", "事件 ID"),
]

_OFFER_STATUS_LABELS = {
    "buyable": "可购买",
    "not_buyable": "不可购买",
    "disabled_by_seller": "卖家已停用",
    "disabled_by_takealot": "Takealot 已停用",
}

_ANOMALY_LABELS = {
    "low_views_high_conversion": "低浏览、高转化",
    "non_buyable": "商品不可购买",
    "sales_drop": "销量下降",
    "sales_spike": "销量突增",
    "unknown_sale_status": "销售状态未配置",
}

_ANOMALY_EXPLANATIONS = {
    "low_views_high_conversion": "浏览量较低但转化率较高，建议检查曝光机会。",
    "non_buyable": "商品状态不是可购买状态。",
    "sales_drop": "当日订购件数低于历史基准。",
    "sales_spike": "当日订购件数高于历史基准。",
    "unknown_sale_status": "销售状态尚未配置，有效件数暂不计算。",
}

_EVENT_LABELS = {
    "unknown_sale_status": "销售状态未配置",
    "missing_sku": "SKU 缺失",
}

_SEVERITY_LABELS = {"warning": "提醒", "critical": "严重"}

_DETAIL_LABELS = {
    "baseline_daily_units": "基准日均件数",
    "ordered_units": "当日订购件数",
    "offer_status": "商品状态",
    "sale_statuses": "销售状态",
}

_TEXT_IDENTIFIER_KEYS = {
    "offer_id",
    "tsin_id",
    "sku",
    "barcode",
    "event_id",
    "productline_id",
    "listing_quality",
}

_OFFER_STATUS_ORDER = {
    "buyable": 0,
    "not_buyable": 1,
    "disabled_by_seller": 2,
    "disabled_by_takealot": 3,
}


def export_excel(dataset: DashboardDataset, destination: Path) -> Path:
    """Write a macro-free workbook derived exclusively from ``dataset``."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    overview = workbook.active
    overview.title = SHEET_NAMES[0]
    for name in SHEET_NAMES[1:]:
        workbook.create_sheet(name)

    store_daily = _store_daily_for_excel(dataset)
    product_daily = _product_daily_for_excel(dataset)
    offer_current = _offer_current_for_excel(dataset.offer_current)
    anomalies = _anomalies_for_excel(dataset)
    quality_events = _quality_events_for_excel(dataset)

    _build_overview(overview, dataset, store_daily)
    _build_product_analysis(workbook["单品分析"], product_daily)
    _build_frame_sheet(
        workbook["商品数据"],
        "商品数据（当前快照）",
        "全部字段来自最近一次 Offer 采集；空白表示 API 未返回，不补零、不推断。",
        offer_current,
        _OFFER_COLUMNS,
    )
    _build_frame_sheet(
        workbook["异常商品"],
        "异常商品",
        "异常详情已拆成可筛选字段；空白表示该类异常没有对应指标。",
        anomalies,
        _ANOMALY_COLUMNS,
    )
    _build_frame_sheet(
        workbook["每日汇总"],
        "每日汇总",
        "按商品每日数据汇总的店铺日级指标。",
        store_daily,
        [
            ("metric_date", "日期"),
            ("ordered_units", "订购件数"),
            ("effective_units", "有效件数"),
            ("ordered_revenue", "订购销售额（ZAR）"),
        ],
    )
    _build_frame_sheet(
        workbook["销售明细"],
        "商品每日销售明细/汇总",
        "本表是商品/日期汇总，不代表订单行；未知销售状态对应的有效件数留白。",
        product_daily,
        _PRODUCT_COLUMNS,
    )
    _build_frame_sheet(
        workbook["流量快照"],
        "流量快照",
        "流量均为 30 天滚动窗口口径；空白表示 API 或历史快照未提供。",
        product_daily,
        _TRAFFIC_COLUMNS,
    )
    _build_metric_notes(workbook["指标说明"])
    _build_frame_sheet(
        workbook["数据质量"],
        "数据质量",
        "质量事件已转换为运营可读字段；事件 ID 仅用于追溯。",
        quality_events,
        _QUALITY_COLUMNS,
    )
    workbook["商品数据"].freeze_panes = "F4"
    workbook["商品数据"].sheet_view.zoomScale = 85
    _wrap_data_column(workbook["异常商品"], _ANOMALY_COLUMNS, "product_title_current")
    workbook["异常商品"].freeze_panes = "D4"
    workbook["销售明细"].freeze_panes = "D4"
    workbook["销售明细"].sheet_view.zoomScale = 85
    workbook["流量快照"].freeze_panes = "D4"
    workbook["流量快照"].sheet_view.zoomScale = 85
    _wrap_data_column(workbook["数据质量"], _QUALITY_COLUMNS, "product_title_current")
    workbook["数据质量"].freeze_panes = "D4"
    _add_conditional_formatting(workbook)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(destination)
    return destination


def _build_overview(
    sheet: Worksheet, dataset: DashboardDataset, store_daily: pd.DataFrame
) -> None:
    _base_sheet(sheet)
    sheet.merge_cells("A1:N1")
    sheet["A1"] = "Takealot 运营总览"
    _style_title(sheet["A1"])
    sheet.merge_cells("A2:N2")
    sheet["A2"] = "只读日报 · KPI 来源于同一 DashboardDataset · 空白表示未知"
    _style_subtitle(sheet["A2"])
    cards = [
        (
            "A3",
            "A4",
            "近7天订购件数",
            _sum_or_none(dataset.store_daily, "ordered_units"),
            "#,##0",
            _PALE_BLUE,
        ),
        (
            "C3",
            "C4",
            "近30天浏览量（商品汇总）",
            _latest_sum_or_none(dataset.product_daily, "metric_date", "page_views_30_days"),
            "#,##0",
            "ECFDF5",
        ),
        (
            "E3",
            "E4",
            "近7天订购销售额",
            _sum_or_none(dataset.store_daily, "ordered_revenue"),
            '"R" #,##0.00',
            _PALE_BLUE,
        ),
        ("G3", "G4", "异常记录数", len(dataset.anomalies), "#,##0", "FFF7ED"),
    ]
    card_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for label_cell, value_cell, label, value, number_format, fill_color in cards:
        first_column = sheet[label_cell].column
        last_column = first_column + 1
        sheet.merge_cells(
            start_row=sheet[label_cell].row,
            start_column=first_column,
            end_row=sheet[label_cell].row,
            end_column=last_column,
        )
        sheet.merge_cells(
            start_row=sheet[value_cell].row,
            start_column=first_column,
            end_row=sheet[value_cell].row,
            end_column=last_column,
        )
        sheet[label_cell] = label
        sheet[label_cell].font = Font(name="Microsoft YaHei", size=10, bold=True, color=_MUTED)
        sheet[label_cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet[value_cell] = _excel_value(value)
        sheet[value_cell].font = Font(name="Microsoft YaHei", size=18, bold=True, color=_NAVY)
        sheet[value_cell].alignment = Alignment(horizontal="center", vertical="center")
        sheet[value_cell].number_format = number_format
        for row in sheet.iter_rows(
            min_row=sheet[label_cell].row,
            max_row=sheet[value_cell].row,
            min_col=first_column,
            max_col=last_column,
        ):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = card_border
    sheet.merge_cells("A5:N5")
    sheet["A5"] = (
        "口径提示：近30天浏览量为各商品 page_views_30_days 的汇总，不代表独立访客；"
        "有效销量需完成销售状态规则确认。"
    )
    sheet["A5"].fill = PatternFill("solid", fgColor=_PALE_AMBER)
    sheet["A5"].font = Font(name="Microsoft YaHei", size=9, color="92400E")
    sheet["A5"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet["A5"].border = Border(
        left=Side(style="thin", color="F59E0B"),
        right=Side(style="thin", color="F59E0B"),
        top=Side(style="thin", color="F59E0B"),
        bottom=Side(style="thin", color="F59E0B"),
    )
    sheet["A6"] = "近7天订购件数趋势（可审计）"
    sheet["A6"].font = Font(name="Microsoft YaHei", size=11, bold=True, color=_NAVY)
    columns = [
        ("metric_date", "日期"),
        ("ordered_units", "订购件数"),
        ("effective_units", "有效件数"),
        ("ordered_revenue", "订购销售额（ZAR）"),
    ]
    last_row = _write_table(sheet, store_daily, columns, header_row=8)
    sheet["E8"] = "图表日期标签"
    sheet["E8"].fill = PatternFill("solid", fgColor=_NAVY)
    sheet["E8"].font = _HEADER_FONT
    sheet["E8"].alignment = Alignment(vertical="center", wrap_text=True)
    for row_index in range(9, last_row + 1):
        sheet.cell(row_index, 5, f'=TEXT(A{row_index},"yyyy-mm-dd")')
        sheet.cell(row_index, 5).font = _BODY_FONT
        sheet.cell(row_index, 5).border = _THIN_BOTTOM
    sheet.freeze_panes = "A9"
    sheet.auto_filter.ref = f"A8:E{last_row}"
    if last_row >= 9:
        chart = LineChart()
        chart.title = "近7天每日订购件数趋势"
        chart.style = 13
        chart.height = 7.2
        chart.width = 14.5
        chart.y_axis.title = "件数"
        chart.x_axis.title = "日期"
        chart.add_data(Reference(sheet, min_col=2, min_row=8, max_row=last_row), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=5, min_row=9, max_row=last_row))
        chart.legend = None
        # Keep the chart below the A9 freeze boundary to avoid a visual split while scrolling.
        sheet.add_chart(chart, "G9")
    for column, width in {"A": 14, "B": 14, "C": 14, "D": 20, "E": 17, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14, "N": 14}.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[1].height = 34
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].height = 24
    sheet.row_dimensions[4].height = 38
    sheet.row_dimensions[5].height = 30
    sheet.row_dimensions[6].height = 26


def _build_product_analysis(sheet: Worksheet, frame: pd.DataFrame) -> None:
    _base_sheet(sheet)
    sheet.merge_cells("A1:Z1")
    sheet["A1"] = "单品分析"
    _style_title(sheet["A1"])
    sheet.merge_cells("A2:Z2")
    sheet["A2"] = (
        "商品名称、售价和状态来自当前 Offer 快照；历史未采集字段保持空白。"
        "图表默认选择首个 Offer ID。"
    )
    _style_subtitle(sheet["A2"])
    ordered = frame.sort_values(["offer_id", "metric_date"], na_position="last") if not frame.empty else frame
    selected: str | None = None
    if not ordered.empty and not ordered["offer_id"].dropna().empty:
        selected = str(ordered["offer_id"].dropna().astype(str).iloc[0])
    sheet["A3"] = "图表商品"
    sheet["B3"] = selected
    selected_title: object = None
    if selected is not None and "product_title_current" in ordered:
        titles = ordered.loc[
            ordered["offer_id"].astype(str) == selected, "product_title_current"
        ].dropna()
        if not titles.empty:
            selected_title = titles.iloc[0]
    sheet["D3"] = "商品名称"
    sheet["E3"] = _excel_value(selected_title)
    sheet["A3"].font = Font(name="Microsoft YaHei", bold=True, color=_MUTED)
    sheet["B3"].font = Font(name="Microsoft YaHei", bold=True, color=_NAVY)
    sheet["D3"].font = Font(name="Microsoft YaHei", bold=True, color=_MUTED)
    sheet["E3"].font = Font(name="Microsoft YaHei", bold=True, color=_NAVY)
    last_row = _write_table(sheet, ordered, _PRODUCT_COLUMNS, header_row=5)
    sheet.freeze_panes = "D6"
    sheet.sheet_view.zoomScale = 85
    last_data_column = len(_PRODUCT_COLUMNS)
    sheet.auto_filter.ref = f"A5:{get_column_letter(last_data_column)}{last_row}"
    if selected is not None:
        selected_count = int((ordered["offer_id"].astype(str) == selected).sum())
        chart_last_row = 5 + selected_count
        helper_column = last_data_column + 2
        helper_letter = get_column_letter(helper_column)
        sheet.cell(5, helper_column, "图表日期标签")
        sheet.cell(5, helper_column).fill = PatternFill("solid", fgColor=_NAVY)
        sheet.cell(5, helper_column).font = _HEADER_FONT
        sheet.cell(5, helper_column).alignment = Alignment(vertical="center", wrap_text=True)
        for row_index in range(6, chart_last_row + 1):
            sheet.cell(row_index, helper_column, f'=TEXT(A{row_index},"yyyy-mm-dd")')
            sheet.cell(row_index, helper_column).font = _BODY_FONT
            sheet.cell(row_index, helper_column).border = _THIN_BOTTOM
        chart = LineChart()
        chart.title = f"商品 {selected} 订购件数趋势"
        chart.style = 13
        chart.height = 7.2
        chart.width = 14.5
        chart.y_axis.title = "件数"
        chart.x_axis.title = "日期"
        ordered_units_column = _column_index(_PRODUCT_COLUMNS, "ordered_units")
        chart.add_data(
            Reference(
                sheet,
                min_col=ordered_units_column,
                min_row=5,
                max_row=chart_last_row,
            ),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(
                sheet,
                min_col=helper_column,
                min_row=6,
                max_row=chart_last_row,
            )
        )
        chart.legend = None
        sheet.add_chart(chart, f"{get_column_letter(helper_column + 2)}4")
        sheet.column_dimensions[helper_letter].width = 16
    _set_widths(sheet, _PRODUCT_COLUMNS)
    sheet.column_dimensions[get_column_letter(last_data_column + 1)].width = 3
    sheet.column_dimensions[get_column_letter(last_data_column + 3)].width = 3


def _build_frame_sheet(
    sheet: Worksheet,
    title: str,
    subtitle: str,
    frame: pd.DataFrame,
    columns: list[tuple[str, str]],
) -> None:
    _base_sheet(sheet)
    last_column = max(1, len(columns))
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet["A1"] = title
    _style_title(sheet["A1"])
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    sheet["A2"] = subtitle
    _style_subtitle(sheet["A2"])
    last_row = _write_table(sheet, frame, columns, header_row=3)
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(last_column)}{last_row}"
    _set_widths(sheet, columns)


def _build_metric_notes(sheet: Worksheet) -> None:
    rows = pd.DataFrame(
        [
            {"metric": "ordered_units", "label": "订购件数", "definition": "所有销售状态的订购数量合计。"},
            {
                "metric": "effective_units",
                "label": "有效件数",
                "definition": "仅计入已确认有效的状态；存在未确认状态时留白。",
            },
            {"metric": "ordered_revenue", "label": "订购销售额", "definition": "销售价乘以订购数量；币种为 ZAR。"},
            {"metric": "page_views_30_days", "label": "近30天浏览量", "definition": "API 返回的 30 天滚动窗口值。"},
            {"metric": "page_views_30_day_average", "label": "近30天日均浏览量", "definition": "近30天浏览量除以 30 的派生平均值。"},
            {"metric": "page_views_window_net_change", "label": "30天浏览量窗口净变化", "definition": "相邻快照的滚动窗口净变化，仅作趋势参考。"},
            {"metric": "conversion_change_points", "label": "转化率变化（百分点）", "definition": "近 30 天与前 30 天转化率之差。"},
            {
                "metric": "current_offer_fields",
                "label": "当前商品字段",
                "definition": "商品名称、当前售价和商品状态来自最近一次 Offer 快照。",
            },
            {
                "metric": "listing_quality",
                "label": "Listing Quality 原值",
                "definition": "API 未说明当前枚举含义，因此保留原值，不自行解释。",
            },
            {
                "metric": "missing_values",
                "label": "缺失值",
                "definition": "空白表示 API 或历史快照未提供，不补零、不推断。",
            },
        ]
    )
    _build_frame_sheet(
        sheet,
        "指标说明",
        "指标口径与使用限制；报告不包含写入或刷新功能。",
        rows,
        [("metric", "字段"), ("label", "展示名称"), ("definition", "说明")],
    )
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 62


def _write_table(
    sheet: Worksheet,
    frame: pd.DataFrame,
    columns: list[tuple[str, str]],
    *,
    header_row: int,
) -> int:
    for column_index, (_, label) in enumerate(columns, start=1):
        cell = sheet.cell(header_row, column_index, label)
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 30
    for row_index, record in enumerate(frame.to_dict(orient="records"), start=header_row + 1):
        wrapped_row = False
        for column_index, (key, _) in enumerate(columns, start=1):
            raw_value = record.get(key)
            cell = sheet.cell(row_index, column_index, _excel_value(raw_value))
            numeric_identifier = False
            if key in _TEXT_IDENTIFIER_KEYS and cell.value is not None:
                if (
                    isinstance(raw_value, str)
                    and raw_value.isdigit()
                    and 11 <= len(raw_value) <= 15
                    and not raw_value.startswith("0")
                ):
                    cell.value = int(raw_value)
                    numeric_identifier = True
                else:
                    cell.value = str(cell.value)
                    cell.quotePrefix = True
            cell.font = _BODY_FONT
            cell.border = _THIN_BOTTOM
            wrap_text = key in {
                "explanation_label",
                "problem_description",
                "sale_statuses",
            }
            wrapped_row = wrapped_row or (wrap_text and cell.value is not None)
            cell.alignment = Alignment(
                horizontal=_horizontal_alignment(key),
                vertical="center",
                wrap_text=wrap_text,
            )
            cell.number_format = "0" if numeric_identifier else _number_format(key)
        if (row_index - header_row) % 2 == 0:
            for column_index in range(1, len(columns) + 1):
                sheet.cell(row_index, column_index).fill = PatternFill("solid", fgColor="F8FAFC")
        sheet.row_dimensions[row_index].height = 32 if wrapped_row else 22
    return max(header_row, header_row + len(frame))


def _base_sheet(sheet: Worksheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.outlinePr.summaryBelow = True


def _style_title(cell: Any) -> None:
    cell.fill = PatternFill("solid", fgColor=_NAVY)
    cell.font = Font(name="Microsoft YaHei", size=18, bold=True, color=_WHITE)
    cell.alignment = Alignment(vertical="center")
    cell.parent.row_dimensions[cell.row].height = 34


def _style_subtitle(cell: Any) -> None:
    cell.fill = PatternFill("solid", fgColor=_PALE_BLUE)
    cell.font = Font(name="Microsoft YaHei", size=10, color=_MUTED)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.parent.row_dimensions[cell.row].height = 28


def _set_widths(sheet: Worksheet, columns: list[tuple[str, str]]) -> None:
    for index, (key, label) in enumerate(columns, start=1):
        width = max(12, min(42, len(label) * 2 + 4))
        if key in {"definition", "explanation_label", "problem_description"}:
            width = 40
        elif key in {"product_title_current", "title"}:
            width = 48
        elif key == "image_url":
            width = 40
        elif key == "sale_statuses":
            width = 32
        elif key in {"offer_id", "event_id"}:
            width = 22 if key == "offer_id" else 38
        elif key in {"tsin_id", "productline_id", "barcode", "sku"}:
            width = 22
        elif key in {"created_at", "updated_at", "captured_at"}:
            width = 22
        elif key in {"offer_status_current", "status_label", "status"}:
            width = 20
        elif key == "anomaly_label":
            width = 20
        elif key == "severity_label":
            width = 10
        sheet.column_dimensions[get_column_letter(index)].width = width


def _wrap_data_column(
    sheet: Worksheet,
    columns: list[tuple[str, str]],
    key: str,
    *,
    first_data_row: int = 4,
) -> None:
    column_index = next(
        (index for index, (column_key, _) in enumerate(columns, start=1) if column_key == key),
        None,
    )
    if column_index is None:
        return
    for row_index in range(first_data_row, sheet.max_row + 1):
        cell = sheet.cell(row_index, column_index)
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal,
            vertical="center",
            wrap_text=True,
        )
        sheet.row_dimensions[row_index].height = max(
            sheet.row_dimensions[row_index].height or 0,
            32,
        )


def _add_conditional_formatting(workbook: Workbook) -> None:
    green = PatternFill("solid", fgColor=_PALE_GREEN)
    red = PatternFill("solid", fgColor=_PALE_RED)
    amber = PatternFill("solid", fgColor=_PALE_AMBER)
    product = workbook["单品分析"]
    if product.max_row >= 6:
        change_letter = _column_letter(_PRODUCT_COLUMNS, "page_views_window_net_change")
        status_letter = _column_letter(_PRODUCT_COLUMNS, "offer_status_current")
        product.conditional_formatting.add(
            f"{change_letter}6:{change_letter}{product.max_row}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=green),
        )
        product.conditional_formatting.add(
            f"{change_letter}6:{change_letter}{product.max_row}",
            CellIsRule(operator="lessThan", formula=["0"], fill=red),
        )
        product.conditional_formatting.add(
            f"{status_letter}6:{status_letter}{product.max_row}",
            FormulaRule(formula=[f'{status_letter}6="可购买"'], fill=green),
        )
        product.conditional_formatting.add(
            f"{status_letter}6:{status_letter}{product.max_row}",
            FormulaRule(
                formula=[
                    f'AND({status_letter}6<>"",{status_letter}6<>"可购买")'
                ],
                fill=red,
            ),
        )
    traffic = workbook["流量快照"]
    if traffic.max_row >= 4:
        change_letter = _column_letter(_TRAFFIC_COLUMNS, "page_views_window_net_change")
        traffic.conditional_formatting.add(
            f"{change_letter}4:{change_letter}{traffic.max_row}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=green),
        )
        traffic.conditional_formatting.add(
            f"{change_letter}4:{change_letter}{traffic.max_row}",
            CellIsRule(operator="lessThan", formula=["0"], fill=red),
        )
    offers = workbook["商品数据"]
    if offers.max_row >= 4:
        status_letter = _column_letter(_OFFER_COLUMNS, "status_label")
        offers.conditional_formatting.add(
            f"{status_letter}4:{status_letter}{offers.max_row}",
            FormulaRule(formula=[f'{status_letter}4="可购买"'], fill=green),
        )
        offers.conditional_formatting.add(
            f"{status_letter}4:{status_letter}{offers.max_row}",
            FormulaRule(
                formula=[
                    f'AND({status_letter}4<>"",{status_letter}4<>"可购买")'
                ],
                fill=red,
            ),
        )
    anomalies = workbook["异常商品"]
    if anomalies.max_row >= 4:
        severity_letter = _column_letter(_ANOMALY_COLUMNS, "severity_label")
        anomalies.conditional_formatting.add(
            f"{severity_letter}4:{severity_letter}{anomalies.max_row}",
            FormulaRule(formula=[f'{severity_letter}4="严重"'], fill=red),
        )
        anomalies.conditional_formatting.add(
            f"{severity_letter}4:{severity_letter}{anomalies.max_row}",
            FormulaRule(formula=[f'{severity_letter}4="提醒"'], fill=amber),
        )
    quality = workbook["数据质量"]
    if quality.max_row >= 4:
        severity_letter = _column_letter(_QUALITY_COLUMNS, "severity_label")
        quality.conditional_formatting.add(
            f"{severity_letter}4:{severity_letter}{quality.max_row}",
            FormulaRule(formula=[f'{severity_letter}4="严重"'], fill=red),
        )
        quality.conditional_formatting.add(
            f"{severity_letter}4:{severity_letter}{quality.max_row}",
            FormulaRule(formula=[f'{severity_letter}4="提醒"'], fill=amber),
        )


def _excel_value(value: object) -> object:
    if value is None or _is_missing(value):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value
    if isinstance(value, Mapping) or isinstance(value, (list, tuple, set)):
        return _readable_detail(value)
    if hasattr(value, "item"):
        return _excel_value(value.item())
    return value


def _store_daily_for_excel(dataset: DashboardDataset) -> pd.DataFrame:
    frame = dataset.store_daily.copy(deep=True)
    unknown_dates, _ = _unknown_status_scope(dataset.quality_events)
    if not frame.empty and "effective_units" in frame and unknown_dates:
        frame.loc[frame["metric_date"].isin(unknown_dates), "effective_units"] = pd.NA
    return frame


def _product_daily_for_excel(dataset: DashboardDataset) -> pd.DataFrame:
    frame = dataset.product_daily.copy(deep=True)
    _, unknown_pairs = _unknown_status_scope(dataset.quality_events)
    if not frame.empty and "effective_units" in frame and unknown_pairs:
        pairs = list(zip(frame["metric_date"], frame["offer_id"], strict=False))
        mask = pd.Series(
            [pair in unknown_pairs for pair in pairs], index=frame.index, dtype=bool
        )
        frame.loc[mask, "effective_units"] = pd.NA

    current_columns = {
        "title": "product_title_current",
        "selling_price": "selling_price_current",
        "total_stock": "total_stock_current",
        "status": "offer_status_code_current",
    }
    if dataset.offer_current.empty or "offer_id" not in dataset.offer_current:
        for target in current_columns.values():
            frame[target] = pd.NA
    else:
        available = [
            column
            for column in ["offer_id", *current_columns]
            if column in dataset.offer_current
        ]
        current = dataset.offer_current[available].copy(deep=True)
        current = current.rename(columns=current_columns)
        frame = frame.merge(current, on="offer_id", how="left", validate="many_to_one")
        for target in current_columns.values():
            if target not in frame:
                frame[target] = pd.NA
    frame["offer_status_current"] = frame["offer_status_code_current"].map(
        _offer_status_label
    )
    return frame


def _offer_current_for_excel(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy(deep=True)
    if result.empty:
        for key, _ in _OFFER_COLUMNS:
            if key not in result:
                result[key] = pd.Series(dtype="object")
        return result
    result["status_label"] = result["status"].map(_offer_status_label)
    result["_status_order"] = result["status"].map(_OFFER_STATUS_ORDER).fillna(99)
    return result.sort_values(
        ["_status_order", "title", "offer_id"], na_position="last"
    ).drop(columns=["_status_order"])


def _anomalies_for_excel(dataset: DashboardDataset) -> pd.DataFrame:
    frame = _with_offer_titles(dataset.anomalies, dataset.offer_current)
    if frame.empty:
        for key, _ in _ANOMALY_COLUMNS:
            if key not in frame:
                frame[key] = pd.Series(dtype="object")
        return frame
    frame["anomaly_label"] = frame["anomaly_type"].map(
        lambda value: _label_or_original(value, _ANOMALY_LABELS)
    )
    frame["severity_label"] = frame["severity"].map(
        lambda value: _label_or_original(value, _SEVERITY_LABELS)
    )
    frame["explanation_label"] = pd.Series(
        [_anomaly_explanation(row) for _, row in frame.iterrows()],
        index=frame.index,
        dtype="object",
    )
    frame["baseline_daily_units"] = frame["details"].map(
        lambda value: _mapping_item(value, "baseline_daily_units")
    )
    frame["detail_ordered_units"] = frame["details"].map(
        lambda value: _mapping_item(value, "ordered_units")
    )
    frame["detail_offer_status"] = frame["details"].map(
        lambda value: _offer_status_label(_mapping_item(value, "offer_status"))
    )
    frame["sale_statuses"] = frame["details"].map(
        lambda value: _joined_values(_mapping_item(value, "sale_statuses"))
    )
    return frame


def _quality_events_for_excel(dataset: DashboardDataset) -> pd.DataFrame:
    frame = _with_offer_titles(dataset.quality_events, dataset.offer_current)
    if frame.empty:
        for key, _ in _QUALITY_COLUMNS:
            if key not in frame:
                frame[key] = pd.Series(dtype="object")
        return frame
    frame["event_label"] = frame["event_type"].map(
        lambda value: _label_or_original(value, _EVENT_LABELS)
    )
    frame["severity_label"] = frame["severity"].map(
        lambda value: _label_or_original(value, _SEVERITY_LABELS)
    )
    frame["sale_statuses"] = frame["details"].map(
        lambda value: _joined_values(_mapping_item(value, "sale_statuses"))
    )
    frame["problem_description"] = pd.Series(
        [_quality_description(row) for _, row in frame.iterrows()],
        index=frame.index,
        dtype="object",
    )
    return frame


def _with_offer_titles(frame: pd.DataFrame, offers: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy(deep=True)
    if result.empty:
        result["product_title_current"] = pd.Series(dtype="object")
        return result
    if offers.empty or not {"offer_id", "title"}.issubset(offers.columns):
        result["product_title_current"] = pd.NA
        return result
    titles = offers[["offer_id", "title"]].rename(
        columns={"title": "product_title_current"}
    )
    return result.merge(titles, on="offer_id", how="left", validate="many_to_one")


def _unknown_status_scope(
    quality_events: pd.DataFrame,
) -> tuple[set[object], set[tuple[object, object]]]:
    required = {"event_date", "event_type", "offer_id"}
    if quality_events.empty or not required.issubset(quality_events.columns):
        return set(), set()
    unknown = quality_events.loc[
        quality_events["event_type"] == "unknown_sale_status",
        ["event_date", "offer_id"],
    ]
    dates = set(unknown["event_date"].dropna())
    pairs = {
        (row.event_date, row.offer_id)
        for row in unknown.itertuples(index=False)
        if row.event_date is not None and row.offer_id is not None
    }
    return cast(set[object], dates), cast(set[tuple[object, object]], pairs)


def _anomaly_explanation(row: pd.Series) -> object:
    anomaly_type = row.get("anomaly_type")
    if isinstance(anomaly_type, str) and anomaly_type in _ANOMALY_EXPLANATIONS:
        return _ANOMALY_EXPLANATIONS[anomaly_type]
    return row.get("explanation")


def _quality_description(row: pd.Series) -> object:
    event_type = row.get("event_type")
    if event_type == "unknown_sale_status":
        return "相关销售状态尚未配置，有效件数暂不计算。"
    details = _readable_detail(row.get("details"))
    return details or row.get("event_label")


def _label_or_original(value: object, labels: Mapping[str, str]) -> object:
    if value is None or _is_missing(value):
        return None
    text = str(value)
    return labels.get(text, text)


def _offer_status_label(value: object) -> object:
    return _label_or_original(value, _OFFER_STATUS_LABELS)


def _mapping_item(value: object, key: str) -> object:
    return value.get(key) if isinstance(value, Mapping) else None


def _joined_values(value: object) -> object:
    if value is None or _is_missing(value):
        return None
    if isinstance(value, (list, tuple, set)):
        values = [str(item) for item in value if item is not None and not _is_missing(item)]
        return "、".join(values) or None
    return str(value)


def _readable_detail(value: object) -> object:
    if value is None or _is_missing(value):
        return None
    if isinstance(value, Mapping):
        parts = []
        for key, item in value.items():
            rendered = _joined_values(item)
            if rendered is not None:
                parts.append(f"{_DETAIL_LABELS.get(str(key), str(key))}：{rendered}")
        return "；".join(parts) or None
    if isinstance(value, (list, tuple, set)):
        return _joined_values(value)
    return str(value)


def _sum_or_none(frame: pd.DataFrame, column: str) -> int | float | None:
    if frame.empty or column not in frame:
        return None
    value = frame[column].sum(min_count=1)
    if _is_missing(value):
        return None
    if hasattr(value, "item"):
        value = value.item()
    return value if isinstance(value, (int, float)) else float(value)


def _latest_sum_or_none(
    frame: pd.DataFrame, date_column: str, value_column: str
) -> int | float | None:
    if frame.empty or date_column not in frame or value_column not in frame:
        return None
    dates = frame[date_column].dropna()
    if dates.empty:
        return None
    value = frame.loc[frame[date_column] == dates.max(), value_column].sum(min_count=1)
    if _is_missing(value):
        return None
    if hasattr(value, "item"):
        value = value.item()
    return value if isinstance(value, (int, float)) else float(value)


def _column_index(columns: list[tuple[str, str]], key: str) -> int:
    for index, (column_key, _) in enumerate(columns, start=1):
        if column_key == key:
            return index
    raise ValueError(f"unknown Excel column key: {key}")


def _column_letter(columns: list[tuple[str, str]], key: str) -> str:
    return cast(str, get_column_letter(_column_index(columns, key)))


def _horizontal_alignment(key: str) -> str:
    if key in {
        "metric_date",
        "event_date",
        "created_at",
        "updated_at",
        "captured_at",
    }:
        return "center"
    if key in {
        "ordered_units",
        "effective_units",
        "ordered_revenue",
        "selling_price",
        "selling_price_current",
        "rrp",
        "benchmark_price",
        "page_views_30_days",
        "page_views_30_day_average",
        "page_views_window_net_change",
        "conversion_percentage_30_days",
        "conversion_percentage_previous_30_days",
        "conversion_change_points",
        "quantity_returned_30_days",
        "total_wishlist",
        "wishlist_30_days",
        "discount_percentage",
        "total_stock",
        "total_stock_current",
        "baseline_daily_units",
        "detail_ordered_units",
    }:
        return "right"
    return "left"


def _number_format(key: str) -> str:
    if key in {"metric_date", "event_date"}:
        return "yyyy-mm-dd"
    if key in {"created_at", "updated_at", "captured_at"}:
        return "yyyy-mm-dd hh:mm"
    if key in {
        "ordered_revenue",
        "selling_price",
        "selling_price_current",
        "rrp",
        "benchmark_price",
    }:
        return '"R" #,##0.00'
    if key in _TEXT_IDENTIFIER_KEYS:
        return "@"
    if key in {
        "ordered_units",
        "effective_units",
        "page_views_30_days",
        "page_views_window_net_change",
        "total_stock",
        "total_stock_current",
        "quantity_returned_30_days",
        "total_wishlist",
        "wishlist_30_days",
        "detail_ordered_units",
    }:
        return "#,##0"
    if key == "baseline_daily_units":
        return "#,##0.0"
    if key in {
        "page_views_30_day_average",
        "conversion_percentage_30_days",
        "conversion_percentage_previous_30_days",
        "conversion_change_points",
        "discount_percentage",
    }:
        return '0.0"%"' if key != "page_views_30_day_average" else "#,##0.0"
    return "General"


def _is_missing(value: object) -> bool:
    return isinstance(value, float) and math.isnan(value) or value is pd.NA or value is pd.NaT
