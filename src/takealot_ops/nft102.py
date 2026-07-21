"""Build an honest, SKU-mapped payload for the NFT102 visitor workbook."""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from sqlalchemy import select
from sqlalchemy.orm import Session

from takealot_ops.storage.models import CollectionRun, OfferSnapshot, SaleItem


SHEET_NAME = "NFT102"
SKU_PATTERN = re.compile(r"(?<!\d)(\d{13})(?!\d)")


@dataclass(frozen=True)
class ProductColumn:
    column_index: int
    column_letter: str
    header: str
    sku: str | None


@dataclass(frozen=True)
class ColumnUpdate:
    column_index: int
    column_letter: str
    header: str
    sku: str | None
    active: bool
    traffic_value: int | None
    order_value: int | None
    reason: str | None


def read_product_columns(template_path: Path) -> tuple[list[ProductColumn], int]:
    """Read NFT102 product headers without saving or mutating the workbook."""
    workbook = load_workbook(template_path, read_only=True, data_only=False)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"工作簿中不存在 {SHEET_NAME} 工作表")
        worksheet = workbook[SHEET_NAME]
        columns: list[ProductColumn] = []
        for column_index in range(3, worksheet.max_column + 1):
            raw_header = worksheet.cell(1, column_index).value
            header = "" if raw_header is None else str(raw_header).strip()
            match = SKU_PATTERN.search(header)
            columns.append(
                ProductColumn(
                    column_index=column_index,
                    column_letter=get_column_letter(column_index),
                    header=header,
                    sku=match.group(1) if match else None,
                )
            )
        return columns, worksheet.max_column
    finally:
        workbook.close()


def build_update_payload(
    session: Session,
    template_path: Path,
    report_date: date,
    *,
    sales_data_complete: bool,
) -> dict[str, Any]:
    """Map the NFT102 columns to today's Offer snapshot and yesterday's Sales."""
    columns, max_column = read_product_columns(template_path)
    sales_date = report_date - timedelta(days=1)

    snapshots = list(
        session.scalars(
            select(OfferSnapshot).where(OfferSnapshot.snapshot_date == report_date)
        )
    )
    snapshot_by_sku = {row.sku: row for row in snapshots if row.sku}
    traffic_complete = (
        session.scalar(
            select(CollectionRun.run_id)
            .where(
                CollectionRun.run_type == "offers",
                CollectionRun.scope_date == report_date,
                CollectionRun.status == "success",
            )
            .limit(1)
        )
        is not None
    )

    sales = list(session.scalars(select(SaleItem).where(SaleItem.sales_day == sales_date)))
    units_by_sku: Counter[str] = Counter()
    for sale in sales:
        if sale.sku:
            units_by_sku[sale.sku] += sale.quantity

    workbook_skus = [column.sku for column in columns if column.sku]
    duplicate_skus = {sku for sku, count in Counter(workbook_skus).items() if count > 1}
    active_column_for_sku: dict[str, int] = {}
    for column in columns:
        if column.sku:
            # The workbook's newer/active duplicate is the right-most column.
            active_column_for_sku[column.sku] = column.column_index

    updates: list[ColumnUpdate] = []
    for column in columns:
        reason: str | None = None
        active = True
        snapshot = snapshot_by_sku.get(column.sku) if column.sku else None
        if column.sku is None:
            active = False
            reason = "表头没有13位SKU，无法安全匹配"
        elif active_column_for_sku[column.sku] != column.column_index:
            active = False
            reason = "SKU在表中重复；为避免订单重复计算，仅使用最右侧列"
        elif snapshot is None:
            reason = "当天Offer快照中没有匹配到SKU"

        traffic_value: int | None = None
        order_value: int | None = None
        if active and snapshot is not None:
            traffic_value = snapshot.page_views_30_days
            if sales_data_complete:
                order_value = units_by_sku.get(column.sku or "", 0)
            elif column.sku is not None and column.sku in units_by_sku:
                order_value = units_by_sku[column.sku]

        updates.append(
            ColumnUpdate(
                column_index=column.column_index,
                column_letter=column.column_letter,
                header=column.header,
                sku=column.sku,
                active=active,
                traffic_value=traffic_value,
                order_value=order_value,
                reason=reason,
            )
        )

    matched = [item for item in updates if item.active and item.sku in snapshot_by_sku]
    return {
        "sheet_name": SHEET_NAME,
        "source_template": str(template_path.resolve()),
        "report_date": report_date.isoformat(),
        "sales_date": sales_date.isoformat(),
        "max_column": max_column,
        "max_column_letter": get_column_letter(max_column),
        "traffic_complete": traffic_complete,
        "sales_complete": sales_data_complete,
        "columns": [asdict(item) for item in updates],
        "summary": {
            "product_columns": len(columns),
            "matched_active_columns": len(matched),
            "traffic_values": sum(item.traffic_value is not None for item in matched),
            "order_values": sum(item.order_value is not None for item in matched),
            "columns_without_sku": sum(item.sku is None for item in updates),
            "duplicate_sku_columns_skipped": sum(
                item.sku in duplicate_skus and not item.active for item in updates
            ),
            "unmatched_sku_columns": sum(
                item.active and item.sku is not None and item.sku not in snapshot_by_sku
                for item in updates
            ),
            "ordered_units_total": sum(units_by_sku.values()),
            "ordered_units_mapped": sum(
                item.order_value or 0 for item in updates if item.active
            ),
        },
        "field_policy": {
            "访客总数": "Takealot page_views_30_days（近30天滚动浏览量）",
            "当天访客数": "接口不提供精确单日访客数，留空",
            "当天订单数": "表格日期前一天的 Sales quantity；采集完整时无订单写0",
            "平台库存数量": "当前尚未采集仓库库存扩展字段，留空",
        },
    }
